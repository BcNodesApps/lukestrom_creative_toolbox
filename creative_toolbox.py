import importlib.util
import calendar
import ctypes
import ctypes.wintypes as wintypes
import io
import json
import math
import os
import random
import re
import shutil
import struct
import subprocess
import sys
import threading
import tkinter as tk
import time as time_module
import uuid
import urllib.parse
import urllib.request
import webbrowser
from datetime import date, datetime, time, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

try:
    from PIL import Image, ImageDraw, ImageTk
except Exception:
    Image = None
    ImageDraw = None
    ImageTk = None

try:
    import psutil
except Exception:
    psutil = None

try:
    import numpy as np
except Exception:
    np = None

try:
    import sounddevice as sd
except Exception:
    sd = None

try:
    import soundcard as sc
except Exception:
    sc = None

try:
    import comtypes
    from comtypes import CLSCTX_ALL
    from pycaw.pycaw import AudioUtilities, IAudioMeterInformation
    PYCAW_IMPORT_ERROR = ""
except Exception:
    PYCAW_IMPORT_ERROR = ""
    try:
        PYCAW_IMPORT_ERROR = str(sys.exc_info()[1])
    except Exception:
        pass
    comtypes = None
    CLSCTX_ALL = None
    AudioUtilities = None
    IAudioMeterInformation = None


APP_VERSION = "V4.6"
APP_TITLE = f"LukeStrom Creative Tool {APP_VERSION}"
BASE_DIR = Path(r"C:\appdevelopment\toolbox\codex")
APP_ICON_FILENAME = "260414 logo lukestrom round.png"
TILE_BACKGROUND_DIR = Path(r"D:\OneDrive\Production\creations")
TILE_BACKGROUND_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
TILE_FONT_EXTS = {".ttf", ".otf"}
TILE_BACKGROUND_SCAN_SECONDS = 0.65
TILE_BACKGROUND_SCAN_LIMIT = 600
DEFAULT_TILE_PICTURE_INTERVAL_SECONDS = 20
DEFAULT_TILE_FONT_INTERVAL_SECONDS = 15
DEFAULT_TILE_FONT_SIZE_MIN_PERCENT = 100
DEFAULT_TILE_FONT_SIZE_MAX_PERCENT = 300
DEFAULT_TILE_ARTWORK_OPACITY_PERCENT = 50
DEFAULT_VU_ARTWORK_OPACITY_PERCENT = 15
CACHECLIP_DIR = Path(r"D:\OneDrive\Production\creations\misc\video\mediacache\CacheClip")
UNIVERSE_DIR = Path(r"D:\OneDrive\Production\creations")
EXTERNAL_APP_ICON_PATH = Path(r"D:\OneDrive\Production\uploads") / APP_ICON_FILENAME
SETTINGS_DIR = Path(os.getenv("APPDATA", str(Path.home()))) / "Creative Toolbox"
SETTINGS_FILE = SETTINGS_DIR / "settings.json"
ICON_CACHE_DIR = SETTINGS_DIR / "icons"
DEFAULT_METRICS_WORKBOOK_PATH = Path(r"D:\OneDrive\Production\creations\misc\app\metrics.xlsx")
DEFAULT_MONTHLY_RECOMMENDATIONS_PATH = Path(r"D:\OneDrive\Production\creations\misc\app\monthly reports\monthly recommendations.txt")
TILE_BACKGROUND_CACHE = {"dir": None, "paths": []}
PERFORMANCE_IO_STATE = {"time": None, "disk": None, "net": None, "gpu_time": 0, "gpu": None}
CPU_TIME_STATE = {"idle": None, "kernel": None, "user": None}
ONEDRIVE_ACTIVITY_STATE = {"time": None, "read": None, "write": None}
LIGHT_THEME = {
    "app_bg": "#f5f3ee",
    "panel_bg": "#ffffff",
    "text": "#202124",
    "muted": "#5f6368",
    "border": "#d8d3c8",
    "entry_bg": "#ffffff",
    "entry_fg": "#111111",
    "nav_bg": "#ffffff",
    "nav_active": "#ececec",
    "notes_bg": "#ece7dc",
    "notes_active": "#ddd5c7",
    "chart_bg": "#ffffff",
    "grid": "#eeeeee",
}
DARK_THEME = {
    "app_bg": "#151515",
    "panel_bg": "#242424",
    "text": "#f3f0e8",
    "muted": "#b8b2a8",
    "border": "#3a3a3a",
    "entry_bg": "#1b1b1b",
    "entry_fg": "#f3f0e8",
    "nav_bg": "#2b2b2b",
    "nav_active": "#3a3a3a",
    "notes_bg": "#3a342a",
    "notes_active": "#4a4235",
    "chart_bg": "#202020",
    "grid": "#363636",
}
TILE_TITLE_FONTS = [
    "Segoe UI",
    "Arial",
    "Calibri",
    "Candara",
    "Century Gothic",
    "Comic Sans MS",
    "Cooper Black",
    "Curlz MT",
    "Elephant",
    "Forte",
    "Franklin Gothic Heavy",
    "Freestyle Script",
    "Gabriola",
    "Georgia",
    "Gill Sans Ultra Bold",
    "Harlow Solid Italic",
    "Harrington",
    "Impact",
    "Jokerman",
    "Kristen ITC",
    "Lucida Handwriting",
    "Magneto",
    "Matura MT Script Capitals",
    "Papyrus",
    "Ravie",
    "Segoe Print",
    "Segoe Script",
    "Showcard Gothic",
    "Snap ITC",
    "Stencil",
    "Tempus Sans ITC",
    "Trebuchet MS",
    "Verdana",
    "Viner Hand ITC",
    "Vivaldi",
    "Wide Latin",
]

DOWNLOADS_PATH = Path.home() / "Downloads"
ARTIST_NAME = "LukeStrom"
YOUTUBE_HANDLE = "@lukestrommusic"
POST_TIME = time(20, 0)
POST_TIME_TEXT = "20:00"

URL_DISTROKID = "https://distrokid.com/mymusic/"
URL_YOUTUBE = "https://studio.youtube.com/channel/UCAUCC7uw_shkAmX7kPMy5IQ/videos/upload?d=ud&filter=%5B%5D&sort=%7B%22columnType%22%3A%22date%22%2C%22sortOrder%22%3A%22DESCENDING%22%7D"
URL_FACEBOOK = "https://business.facebook.com/latest/posts/published_posts/?business_id=678482264451473&asset_id=1126335547219736"
URL_TIKTOK = "https://www.tiktok.com/tiktokstudio/content"
URL_METRICOOL = "https://app.metricool.com/planner/calendar?blogId=6089075&userId=4699196"
URL_METRICOOL_SUMMARY = "https://app.metricool.com/evolution/brandSummary?blogId=6089075&userId=4699196"
URL_GOOGLE_REMOTE_DESKTOP = "https://remotedesktop.google.com/access"
URL_LINKTREE = "https://linktr.ee/lukestrommusic"
URL_MTRBIO = "https://t.mtrbio.com/lukestrom"
URL_YANDEX = "https://yandex.com/"
CHATGPT_LINKS = [
    ("Reaper", "https://chatgpt.com/g/g-p-68d65621edf0819180a8e1b3964afba1-reaper/project"),
    ("Music marketing", "https://chatgpt.com/g/g-p-6928a52f02dc81919c5b4a68247730b3-music-marketing/project"),
    ("Marketing analysis", "https://chatgpt.com/g/g-p-69242e5fe6ec819191b4772acb2cc0de-music-marketing-analysis/project"),
    ("Songwriting", "https://chatgpt.com/g/g-p-6953cce5398c8191b90b6957929fd529-songwriting/project"),
    ("Instrument tech", "https://chatgpt.com/g/g-p-693132401eac8191a3e4f60fc044ceb9-instrument-tech/project"),
    ("IT & operations", "https://chatgpt.com/g/g-p-690630d67b648191a9537d412e1dd23b-it-operations/project"),
    ("Video/animation", "https://chatgpt.com/g/g-p-684552aadb348191a668f41fe5eb8582-videography-animation-davinci/project"),
]

HOME_LINKS = [
    ("DistroKid", URL_DISTROKID),
    ("Metricool", URL_METRICOOL),
    ("Meta", URL_FACEBOOK),
    ("YouTube", URL_YOUTUBE),
    ("TikTok", URL_TIKTOK),
    ("Linktree", URL_LINKTREE),
]

WEB_SHORTCUTS = [
    {"title": "DistroKid", "url": URL_DISTROKID, "icon_domain": "distrokid.com", "color": "#1db954"},
    {"title": "Metricool", "url": URL_METRICOOL, "icon_domain": "metricool.com", "color": "#2f6f73"},
    {"title": "Meta", "url": URL_FACEBOOK, "icon_domain": "facebook.com", "color": "#3267d6"},
    {"title": "YouTube", "url": URL_YOUTUBE, "icon_domain": "youtube.com", "color": "#d93025"},
    {"title": "TikTok", "url": URL_TIKTOK, "icon_domain": "tiktok.com", "color": "#111111"},
    {"title": "Linktree", "url": URL_LINKTREE, "icon_domain": "linktr.ee", "color": "#43e660"},
    {"title": "MTR Bio", "url": URL_MTRBIO, "icon_domain": "t.mtrbio.com", "color": "#6f42c1"},
    {"title": "Remote Desktop", "url": URL_GOOGLE_REMOTE_DESKTOP, "icon_domain": "remotedesktop.google.com", "color": "#5f6368"},
    {"title": "Yandex", "url": URL_YANDEX, "icon_domain": "yandex.com", "color": "#fc3f1d"},
]

DEFAULT_HASHTAGS = """#lukestrom
#meloverse
#originalmusic
#rockmusic
#breda"""

MELOVERSE_WORDS = """
MeloVerse Lukestrom Miles MouMou Nicco Mercedes Jeff Gaylord Caesar Jolene Cruisin Rain Horizon Highway Journey
Freedom Sunset Dawn Twilight Asphalt Desert Ocean Harbor Lighthouse Forest Trailer Workshop Grease Chrome Steel
Porsche Cadillac Harley Vespa Telecaster Gretsch Marshall Amplifier Vinyl Cassette Guitar Solo Chorus Echo Harmony
Melody Atmosphere Silence Static Motion Drift Wander Escape Longing Nostalgia Solitude Hope Regret Mystery Tension
Yearning Distance Memory Dream Reflection Shadow Moonlight Fireworks Neon Motel Diner Coffee Letter Photograph Window
Mirror Skyline Breda Nightingale Platform Reunion Goodbye Waiting Promise Fortune Thunder Storm Breeze Dust Gravel
Cigarette Denim Leather Ponytail Compass Compassion Story Legend Tomorrow Home
""".split()

FIRST_POST_OFFSETS = [0, 3, 8, 15]
REPOST_START_OFFSETS = [21, 46, 76, 111]


def resource_root():
    return Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))


def resource_path(*parts):
    return resource_root().joinpath(*parts)


def first_existing_path(*paths):
    for path in paths:
        if path.exists():
            return path
    return paths[0]


def load_settings():
    try:
        if SETTINGS_FILE.exists():
            return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def save_settings(settings):
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        SETTINGS_FILE.write_text(json.dumps(settings, indent=2), encoding="utf-8")
    except Exception:
        pass


def get_tile_background_dir():
    if load_settings().get("tile_background_disabled", False):
        return None
    configured = load_settings().get("tile_background_dir")
    if configured:
        path = Path(configured)
        if path.exists():
            return path
    return TILE_BACKGROUND_DIR


def get_tile_font_dir():
    if load_settings().get("basic_font_only", False):
        return None
    configured = load_settings().get("tile_font_dir")
    if configured:
        path = Path(configured)
        if path.exists():
            return path
    return None


def get_metrics_workbook_path():
    configured = load_settings().get("metrics_workbook_path")
    if configured:
        return Path(configured)
    return DEFAULT_METRICS_WORKBOOK_PATH


def get_monthly_recommendations_path():
    configured = load_settings().get("monthly_recommendations_path")
    if configured:
        return Path(configured)
    return DEFAULT_MONTHLY_RECOMMENDATIONS_PATH


MONTH_NAME_LOOKUP = {name.lower(): name.lower() for name in calendar.month_name if name}
MONTH_NAME_LOOKUP.update({name.lower(): calendar.month_name[index].lower() for index, name in enumerate(calendar.month_abbr) if name})


def monthly_report_path_from_heading(text):
    text = text.strip()
    patterns = [
        r"^([A-Za-z]+)\s+(\d{2,4})$",
        r"^(\d{2,4})\s+([A-Za-z]+)$",
    ]
    for pattern in patterns:
        match = re.match(pattern, text)
        if not match:
            continue
        first, second = match.groups()
        if first.isdigit():
            year_text, month_text = first, second
        else:
            month_text, year_text = first, second
        month = MONTH_NAME_LOOKUP.get(month_text.lower())
        if not month:
            continue
        year = year_text[-2:]
        return get_monthly_recommendations_path().parent / f"{year}{month}.pdf"
    return None


def get_interval_setting(key, default_seconds):
    try:
        value = int(load_settings().get(key, default_seconds))
    except Exception:
        value = default_seconds
    return max(5, min(60, value))


def get_tile_picture_interval_seconds():
    return get_interval_setting("tile_picture_interval_seconds", DEFAULT_TILE_PICTURE_INTERVAL_SECONDS)


def get_tile_font_interval_seconds():
    return get_interval_setting("tile_font_interval_seconds", DEFAULT_TILE_FONT_INTERVAL_SECONDS)


def get_percent_setting(key, default_percent, minimum=0, maximum=100):
    try:
        value = int(load_settings().get(key, default_percent))
    except Exception:
        value = default_percent
    return max(minimum, min(maximum, value))


def get_tile_artwork_opacity_percent():
    return get_percent_setting("tile_artwork_opacity_percent", DEFAULT_TILE_ARTWORK_OPACITY_PERCENT)


def get_vu_artwork_opacity_percent():
    return get_percent_setting("vu_artwork_opacity_percent", DEFAULT_VU_ARTWORK_OPACITY_PERCENT)


def get_tile_font_size_range():
    settings = load_settings()
    try:
        minimum = int(settings.get("tile_font_size_min_percent", DEFAULT_TILE_FONT_SIZE_MIN_PERCENT))
        maximum = int(settings.get("tile_font_size_max_percent", DEFAULT_TILE_FONT_SIZE_MAX_PERCENT))
    except Exception:
        minimum = DEFAULT_TILE_FONT_SIZE_MIN_PERCENT
        maximum = DEFAULT_TILE_FONT_SIZE_MAX_PERCENT
    minimum = max(100, min(300, minimum))
    maximum = max(100, min(300, maximum))
    if minimum > maximum:
        minimum, maximum = maximum, minimum
    return minimum, maximum


def use_basic_fonts_only():
    return bool(load_settings().get("basic_font_only", False))


def get_dark_mode():
    return bool(load_settings().get("dark_mode", False))


def theme_colors(dark_mode=None):
    if dark_mode is None:
        dark_mode = get_dark_mode()
    return DARK_THEME if dark_mode else LIGHT_THEME


REGISTERED_FONT_PATHS = set()


def font_family_from_file(path):
    data = path.read_bytes()
    table_count = struct.unpack(">H", data[4:6])[0]
    name_offset = None
    for index in range(table_count):
        offset = 12 + index * 16
        tag = data[offset:offset + 4]
        if tag == b"name":
            name_offset = struct.unpack(">I", data[offset + 8:offset + 12])[0]
            break
    if name_offset is None:
        return None

    count = struct.unpack(">H", data[name_offset + 2:name_offset + 4])[0]
    storage_offset = name_offset + struct.unpack(">H", data[name_offset + 4:name_offset + 6])[0]
    best = None
    for index in range(count):
        offset = name_offset + 6 + index * 12
        platform_id, _encoding_id, language_id, name_id, length, string_offset = struct.unpack(">HHHHHH", data[offset:offset + 12])
        if name_id not in (1, 16):
            continue
        raw = data[storage_offset + string_offset:storage_offset + string_offset + length]
        try:
            text = raw.decode("utf-16-be" if platform_id in (0, 3) else "macroman").strip("\x00").strip()
        except Exception:
            text = raw.decode("latin-1", errors="ignore").strip("\x00").strip()
        if not text:
            continue
        if name_id == 16 or language_id in (0x0409, 0):
            return text
        best = best or text
    return best


def register_font_file(path):
    if os.name != "nt" or str(path) in REGISTERED_FONT_PATHS:
        return
    try:
        ctypes.windll.gdi32.AddFontResourceExW(str(path), 0x10, 0)
        REGISTERED_FONT_PATHS.add(str(path))
    except Exception:
        pass


def custom_tile_fonts():
    font_dir = get_tile_font_dir()
    if font_dir is None:
        return []
    fonts = []
    for path in font_dir.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in TILE_FONT_EXTS:
            continue
        try:
            register_font_file(path)
            family = font_family_from_file(path)
            if family and family not in fonts:
                fonts.append(family)
        except Exception:
            pass
    return fonts


def tile_title_fonts():
    if use_basic_fonts_only():
        return ["Segoe UI", "Arial", "Calibri", "Verdana", "Trebuchet MS", "Georgia"]
    fonts = custom_tile_fonts() + TILE_TITLE_FONTS
    return list(dict.fromkeys(fonts))


def export_metrics_template(output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "data entry"
    ws.append(["week", "metri", None, "tt views", "tt followers", None, "ig views", "ig followers", None, "yt views", "yt followers", None, "fb views", "fb followers"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    for col in "DEGHJKMN":
        ws.column_dimensions[col].width = 14
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F6F73")

    monthly = wb.create_sheet("Monthly")
    monthly.append(MONTHLY_HEADERS)
    widths = [12] + [20] * 16 + [42, 42]
    for index, width in enumerate(widths, start=1):
        monthly.column_dimensions[chr(64 + index) if index <= 26 else "Z"].width = width
    for cell in monthly[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7B5F2A")

    wb.save(output_path)


SONG_ANALYZER_PATH = first_existing_path(
    resource_path("13-song_analyzer", "song_analyzer.py"),
    BASE_DIR / "13-song_analyzer" / "song_analyzer.py",
)
RELEASE_NOTES_FILE = first_existing_path(
    resource_path("8-youtube_dl", "release notes.txt"),
    BASE_DIR / "8-youtube_dl" / "release notes.txt",
)
APP_RELEASE_NOTES = """# Creative Toolbox release notes

## Current version

Creative Toolbox is now a single-window creator dashboard for music, reels, planning, downloads, metrics, system actions, and quick creator links.

## V4.6
- Monthly charts now display a single-month platform snapshot when only one Monthly row exists.
- Once two or more months are available, the monthly charts automatically switch back to trend lines.

## V4.5
- Audio L/R and Upload/Download are now placed as fixed VU meter pairs, so they cannot split onto separate rows.
- The two regular Windows File Explorer windows opened by the MeloVerse Explorer are moved side by side instead of appearing on top of each other.

## V4.4
- Upload/Download and Audio L/R VU meters now stay side by side by switching the meter grid to paired columns when needed.
- The LukeStrom MeloVerse Explorer now opens two regular Windows File Explorer windows for the creations folder instead of the custom in-app explorer.

## V4.3
- VU was renamed to Tools, and the former Tools page was renamed to Shortcuts.
- A refresh icon was added next to the hamburger menu for quick page refreshes.
- The old Refresh system-action tile was removed from Shortcuts.
- The LukeStrom MeloVerse Explorer now opens the main creations folder.
- Monthly worksheet import now reads by column names and accepts common aliases such as views/impressions and fb/Facebook.
- Weekly observations were added as an automatic dashboard panel without changing the Weekly worksheet.
- The Tools meter page hides the audio source selector and instead shows compact system action buttons with hover help.
- Upload/Download and Audio L/R meters are kept in side-by-side pairs when possible.

## V4.2
- Monthly worksheet detection is now case-insensitive, so monthly, Monthly, and MONTHLY all work.
- Universe opens in the app window again, now defaulting to a cleaner list-style file browser.

## V4.1
- Fixed the VU page startup by initializing the meter availability/dropdown state.
- VU meters no longer become permanently unchecked after a temporary n/a reading.

## V4.0
- VU meters now use all available vertical space and reflow dynamically when meters are selected or deselected.
- If a performance meter has no available value, it is hidden from the grid and disabled in the VU meters menu.
- The fullscreen toggle now maximizes/restores on the current screen instead of using exclusive fullscreen.
- Hack Windows was removed from the Tools UI to avoid unnecessary CPU load.
- Metrics now treats weekly views as period-based weekly impressions, not cumulative lifetime totals.
- Cumulative Total Views and the Total Views Over Time chart were removed.
- A Monthly worksheet is supported for monthly impressions, followers, posts, interactions, Dashboard Summary, and Next Month Focus.
- Monthly KPI cards and monthly platform line charts were added.
- Monthly Summary now reads from the Monthly worksheet instead of a separate recommendations text file.

## V3.9
- Fixed the embedded Audio L/R meter engine so the VU page can render correctly.
- The VU grid now expands with the page instead of behaving like a shallow toolbar.
- Campaign interval number fields now follow dark mode colors.
- Metrics growth percentages now show whole numbers unless the change is smaller than 1%.

## V3.8
- Audio L and Audio R are now selectable meters in the same VU grid as Processor, Memory, Disk, Network, GPU, Core temp, and OneDrive activity.
- The separate Audio VU section was removed from the VU page layout.
- Campaign Planner interval sliders now include numeric day fields for precise short intervals.
- Campaign Planner interval sliders were widened for easier use.

## V3.7
- The Home tile and navigation item were renamed from Audio VU to VU.
- VU is now the dedicated meter page, combining system meters, OneDrive activity meters, and the Audio VU meters.
- The VU page has a checkable VU meters menu, including Audio VU.
- Tools now focuses on tools, system actions, web shortcuts, and file/browser helpers instead of showing meters.
- A Fullscreen option was added to the hamburger menu.
- ChatGPT was added to Web shortcuts and opens a dedicated subpage with project tiles.
- Campaign Planner intervals can now be edited with sliders.
- Metrics now shows growth percentages versus last week and a Platform MVP card.
- The misleading Windows ACPI temperature fallback was removed so fixed 28C readings no longer pretend to be real sensor data.
- Audio VU now shows its audio source selector.

## V3.6
- The Tools VU meter grid now compacts after every dropdown change so deselected meters cannot leave empty spaces between visible meters.

## V3.5
- Hack Windows now uses generated random MeloVerse mini-icons instead of repeating the app feather icon.
- Hack Windows now flashes MeloVerse vocabulary words during the terminal stream.
- ASCII engine/data bursts appear more often and also appear immediately when a Hack Windows terminal opens.
- Hack Windows titles now mix artwork filenames, MeloVerse words, and varied trace labels.

## V3.4
- OneDrive no longer shows "Files: none detected" inside the OneDrive system tile.
- Upload activity and Download activity VU meters now show their detected file path below the meter.
- Hack Windows now randomizes window titles with varied trace labels instead of always using system trace.
- Hack Windows can use artwork images as window icons.
- Hack Windows output now includes live-looking process names, inbound/outbound data strings, flow rates, and occasional ASCII engine bursts.

## V3.3
- Tools now uses a VU meters dropdown with checkable meter options.
- Processor, Memory, Disk, Network, GPU, Core temp, Upload activity, and Download activity are all checked by default.
- Upload and Download activity meters are now part of the main VU meter set instead of a separate OneDrive toggle.
- The Home tile callback was made tolerant so missing Tkinter event objects no longer cause callback errors.
- Hack Windows windows now stay visible for 15 to 30 seconds, use larger sizes, and use artwork filenames as randomized window titles.

## V3.2
- Tools now includes a Hack Windows tile.
- Hack Windows starts a safe theatrical terminal storm with cmd-style pop-up windows.
- The terminal windows show fast scrolling live-looking system, sync, route, checksum, and creative matrix info in random colors.
- Up to three terminal windows can be visible at once, each with a random size and 5 to 15 second lifetime.
- Clicking Home stops the Hack Windows process and closes all active terminal windows.

## V3.1
- Tools now has an optional OneDrive activity meter toggle in System actions.
- When enabled, Upload activity and Download activity VU meters appear next to the OneDrive tile.
- The OneDrive activity meters show sync activity strength, not exact per-file progress.
- The most likely recent OneDrive file path is shown below each activity meter when detected.
- Other System actions move down automatically while the OneDrive activity meters are visible.

## V3.0
- Universe now shows folders and files, supports list and big-icon views, opens files, copies selected items, and can open the current folder in File Explorer.
- Tools now uses a single plus/minus control to switch between 3 and 6 visible VU meters.
- System actions and Web shortcuts returned to clean solid panel backgrounds.
- OneDrive tries to show compact recent file names alongside transfer activity.
- The hamburger button now uses the proper three-line symbol again.
- The Windows temperature fallback was restored for laptops where CoreTemp/OpenHardwareMonitor is not available.

## V2.9
- Tools now uses inline 3/6 selectors to choose how many performance VU meters are visible.
- System actions and Web shortcuts now use one artwork background behind each full pane.
- The Universe tile opens the LukeStrom universe folder inside the app with a Back button.
- MTR Bio was added to Web shortcuts.
- Outlook close is silent, without a confirmation popup.
- Unavailable Tools VU meters now show only full-opacity artwork instead of a dead meter.
- Campaign captions now include the LukeStrom MeloVerse and Full Songs MTR Bio block.
- The unreliable fixed Windows temperature fallback was removed so fake frozen laptop temperatures are not shown.
- Canvas shapes use consistent 10px corner roundness.

## V2.8
- The performance VU expand control moved out of the main menu bar and into the Tools page.
- Monthly recommendation headings are clickable and open the matching monthly PDF report when available.

## V2.7
- A separate all-meters performance VU window was added before being replaced by the inline 3/6 Tools control.
- GPU detection gained extra Windows counter fallbacks for better laptop compatibility.
- Monthly recommendations gained report-link parsing for headings such as june 26 and may 26.

## V2.6
- Metrics was redesigned as a responsive dashboard for clearer balance across screen sizes.
- Metrics now uses a stable KPI row, compact entry/table column, and adaptive chart/recommendation grid.

## V2.5
- Metrics layout now keeps Monthly recommendations visible in the lower-right area.
- The platform chart and recommendations pane now share the bottom-right row.

## V2.4
- Rounded corners were made a little larger across tiles, artwork masks, icon blocks, and meter frames.
- Reel Design now shows and exports the fixed reel guideline.
- Metrics now has a Monthly recommendations pane on the lower-right side.
- The monthly recommendations text file path can be changed from the hamburger menu under Metrics.

## V2.3
- Tiles, icon blocks, meter frames, and artwork masks now use subtle rounded corners.
- Square frame outlines around tile cards were removed so the rounded canvas shapes define the tile edges.

## V2.2
- OneDrive tile text is simplified to status plus compact transfer speed so it no longer clogs.
- Real tile icons now sit inside circular icon holders for a cleaner, consistent look.

## V2.1
- Tools tiles now try to use real favicon/logo image icons instead of drawn placeholder icons.
- Icons are cached under the app settings folder so they load faster after the first run.

## V2.0
- Tools now scrolls as one full page when the fixed-size tiles do not fit on screen.
- System actions now use the same tile size and grid rhythm as Web shortcuts.
- System action tiles and Web shortcut tiles stopped using text initials.

## V1.9
- OneDrive system action tile is now double-width so sync status no longer overlaps.
- OneDrive activity text is compacted to one speed line plus one shortened recent-file line.

## V1.8
- Tools performance meters default to Processor, Memory, and Disk for better laptop compatibility.
- Performance collection is more defensive so one unavailable metric no longer blanks all meters.
- Web shortcuts only show the vertical scrollbar when content is taller than the visible area.
- OneDrive tile is taller, wraps activity text, and shows recent OneDrive files as a practical sync hint.
- Audio VU no longer tries to name the playback app, because Windows can report stale audio sessions.

## V1.7
- The bold duplicate app title was removed from the top navigation bar; the version remains in the window title.
- Home now starts faster by loading tile artwork after the first screen is shown.
- Media cache deletion now shows determinate progress in the Tools page.
- OneDrive activity text now spells out upload and download rates.
- Audio VU now displays digital dBFS values instead of only percentages.

## V1.6
- Audio VU now uses a single thin meter border instead of a heavy frame.
- Audio VU labels were shortened to L and R.
- Audio VU now has a red peak marker that resets every 5 seconds.

## V1.5
- Audio VU bottom diagnostics and source text were removed from the visible page for a cleaner always-on meter view.
- Audio VU meter panels were simplified to one subtle frame instead of nested frames.
- Audio VU artwork now fills the full meter face when artwork is available.

## V1.4
- Audio VU Windows peak scaling was softened so YouTube/browser audio no longer jumps to 100% immediately and stays there.
- Audio VU needle movement now uses separate attack and release smoothing for a more musical analog response.

## V1.3
- Audio VU now tries multiple pycaw/Windows peak-meter access routes for better compatibility across Python and laptop audio setups.
- Audio VU diagnostic messages now include more useful detail when Windows exposes the audio device differently.

## Navigation
- The app now shows the visible build label LukeStrom Creative Tool V2.6 in the window title.
- The top bar now stays focused on app navigation only: Home, Song, Reel, Campaign, Audio VU, Tools, and Metrics.
- Audio VU was added as a Home tile for music-reactive meters.
- YouTube Downloader moved from Home/top navigation into Tools.
- Top navigation is now clickable text instead of button blocks.
- Release notes moved into the hamburger menu under Info.
- A new How to page was added under Info.
- Home tiles and Tools tiles are clickable directly.

## Visual Style
- Home tiles can use artwork from a configurable folder, including subfolders.
- Tile titles can rotate through installed fonts and custom font folders.
- Tile title size can vary between the selected minimum and maximum.
- VU meters can use the same artwork folder as a soft 15% opacity background.
- VU meter artwork opacity was reduced to 15% for better readability.
- Tile artwork opacity and VU artwork opacity can now be adjusted under Artwork.
- Artwork scanning is cached and time-limited so startup feels faster.
- Dark mode, No picture, Basic font, image intervals, font intervals, and font size can be changed from the hamburger menu.

## Tools
- Tools opens inside the main window.
- OneDrive, Outlook, and Media Cache are clickable system action tiles.
- Refresh is now also a system action tile.
- The OneDrive tile changes between start and stop depending on current status.
- The OneDrive tile now tries to show current read/write activity while syncing.
- The Media Cache tile shows file count and total size before deletion.
- Deleting Media Cache now shows progress inside the Tools page.
- The old Open all core links button is now a web shortcut tile.
- The Summary web shortcut tile was removed.
- Web shortcuts are shown as clickable tiles and open in Chrome when available.
- Web shortcuts now sit in a scrollable area inside Tools.
- VU meters show live system information with smoother analog-style needle movement.
- VU meter peak text was replaced by a small peak marker line.
- Tools shows three VU meters by default: Processor, GPU, and Memory.
- Clicking a VU meter can still switch it to CPU, GPU, Memory, Disk, Network, or Core temp.
- VU peak markers now stick out over the top side of the scale and reset after 1 minute.
- Performance collection avoids overlapping measurements so the app stays more responsive.
- Audio VU prefers the Windows output peak meter through pycaw for YouTube/Chrome audio, with soundcard and sounddevice as fallbacks.
- Audio VU starts automatically when opened and no longer needs Start/Stop buttons.

## Hamburger Menu
- The hamburger menu is grouped into Artwork, Fonts, Metrics, and Info.
- Info sits at the bottom of the hamburger menu and contains Release notes and How to.
- Dark mode can be switched on or off from the hamburger menu.
- Pictures can be disabled with No picture.
- Fonts can be simplified with Basic font.
- Artwork settings include picture folder and picture interval.
- Artwork settings include separate opacity controls for Home tiles and VU meters.
- Font settings include font folder, font interval, and font size.
- Metrics settings include Excel location and export of a clean starter template.

## Metrics
- Metrics uses a configurable Excel workbook.
- Weekly entry fields are aligned per platform, with views and followers on the same row.
- Week and Sunday update date are managed at the bottom of the Add week panel.
- Tables show newest entries first and can be scrolled for older data.
- Charts show total views over time, total followers over time, and latest week views by platform.
- The x-axis for timeline charts uses month labels.

## YouTube Downloader
- Full video and segment downloads are supported.
- Segment downloads automatically attempt the highest available quality.
- Finished filenames include the actual detected resolution.
- FFmpeg is used for local processing when available.

## Planning Tools
- Reel Design and Campaign Planner are separated into their own tiles.
- Reel Design opens the Downloads folder after creating the design text file.
- Campaign Planner can generate campaign files using configurable post and repost counts.
- Campaign Planner supports a custom campaign start date.
- Campaign Planner can use default intervals or custom post/repost day offsets.
- Campaign Planner groups post schedule output per post file, including repost dates.
- Campaign Planner no longer writes fixed post times into the schedule text.
- Song descriptions start from "A song about..." by default.

## Packaging
- The app is prepared for a one-file Windows exe build.
- External artwork and metrics files stay configurable after packaging.
"""
APP_HOW_TO = """# Creative Toolbox how to

## Home
Use Home as your starting point. Click any tile to open that tool in the same window. The changing images and fonts are only there for atmosphere; the tiles themselves are the app navigation.

## Song Analyzer
Use this when you want a quick report for a song file. Choose an audio file, run the analysis, then save the result as a PDF if you want to keep or share it.

## Reel Design
Use this when you want creative notes for a reel. Enter the song/title details and the number of posts you want. The app creates a text file and opens the Downloads folder when it is done.

## Campaign Planner
Use this when you want a release or posting plan. Enter the song details, number of posts, number of reposts, and the campaign start date. Default intervals are used unless you uncheck the default intervals option and enter your own day spacing. The output is grouped per post, with all repost dates directly under the original post.

## YouTube Downloader
Open this from Tools. Use it to download a full YouTube video or only a segment. Paste the link. Leave Start and End empty for the full video, or enter times like 00:30 and 00:45 for a segment. The downloader automatically tries the highest available quality and puts the actual resolution in the filename.

## Tools
Use Tools when you want live meters and quick system actions in one screen. The grid can show processor, memory, disk, network, GPU, temperature when available, OneDrive upload/download activity, and Audio L/R. Use the VU meters menu to choose which meters are visible. The small action buttons can start or stop OneDrive, close Outlook, delete the media cache, open YouTube Downloader, or open the MeloVerse Explorer.

## Shortcuts
Use Shortcuts for creator links and quick utilities. OneDrive, Outlook, Media Cache, YouTube Downloader, and the LukeStrom MeloVerse Explorer are system action tiles. Web shortcut tiles open creator sites in Chrome. The ChatGPT tile opens your project links.

## Metrics
Use Metrics to keep weekly operational stats and monthly strategic stats. Weekly answers what happened this week: enter impressions and followers per platform, then save the week to Excel. Monthly answers how the project is developing over time: fill the Monthly worksheet with Metricool report values and ChatGPT analysis. The dashboard shows weekly impressions, weekly followers, platform MVP, monthly KPI cards, monthly trend charts, Dashboard Summary, and Next Month Focus.

## Hamburger Menu
Use the hamburger menu for settings and help. Artwork controls the picture folder, image interval, tile opacity, and VU opacity. Fonts controls custom fonts, basic font mode, font interval, and title size. Metrics controls the Excel file location and can export a starter template. Info sits at the bottom and contains Release notes and this How to.

## Suggested Workflow
Start with Song Analyzer or Reel Design while creating content. Use Campaign Planner when a song or reel batch is ready to publish. Open YouTube Downloader from Shortcuts when you need source clips. Use Tools when you want visual meters reacting to your system and music. Keep Shortcuts open for web links and file utilities. Update Metrics once a week after your Sunday 20:00 stats check.
"""
APP_ICON_PATH = first_existing_path(
    EXTERNAL_APP_ICON_PATH,
    resource_path(APP_ICON_FILENAME),
)


def one_drive_candidates():
    candidates = []
    for env_name in ("LOCALAPPDATA", "PROGRAMFILES", "PROGRAMFILES(X86)"):
        value = os.environ.get(env_name)
        if value:
            candidates.append(Path(value) / "Microsoft" / "OneDrive" / "OneDrive.exe")
    candidates.extend([
        Path.home() / "AppData" / "Local" / "Microsoft" / "OneDrive" / "OneDrive.exe",
        Path(r"C:\Program Files\Microsoft OneDrive\OneDrive.exe"),
        Path(r"C:\Program Files (x86)\Microsoft OneDrive\OneDrive.exe"),
    ])
    return candidates


def one_drive_exe():
    for candidate in one_drive_candidates():
        if candidate.exists():
            return candidate
    return None


def chrome_candidates():
    candidates = []
    for env_name in ("PROGRAMFILES", "PROGRAMFILES(X86)", "LOCALAPPDATA"):
        value = os.environ.get(env_name)
        if value:
            candidates.extend([
                Path(value) / "Google" / "Chrome" / "Application" / "chrome.exe",
                Path(value) / "Google" / "Chrome Beta" / "Application" / "chrome.exe",
            ])
    candidates.extend([
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
    ])
    return candidates


def chrome_exe():
    for candidate in chrome_candidates():
        if candidate.exists():
            return candidate
    found = shutil.which("chrome") or shutil.which("chrome.exe")
    return Path(found) if found else None


def is_onedrive_running():
    try:
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq OneDrive.exe"],
            capture_output=True,
            text=True,
            timeout=5,
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )
        return "OneDrive.exe" in result.stdout
    except Exception:
        return False


def one_drive_activity_text():
    snapshot = one_drive_activity_snapshot()
    status = snapshot.get("status", "Running")
    if status == "Stopped":
        return "Stopped"
    if status == "starting":
        return "Running | activity starting"
    if status == "unknown":
        return "Running | activity unknown"
    return f"Running | D {format_bytes(snapshot.get('download_rate', 0))}/s | U {format_bytes(snapshot.get('upload_rate', 0))}/s"


def one_drive_activity_snapshot():
    if not is_onedrive_running():
        ONEDRIVE_ACTIVITY_STATE["time"] = None
        ONEDRIVE_ACTIVITY_STATE["read"] = None
        ONEDRIVE_ACTIVITY_STATE["write"] = None
        return {"status": "Stopped", "download_rate": 0, "upload_rate": 0}
    try:
        read_bytes = 0
        write_bytes = 0
        if psutil is not None:
            for process in psutil.process_iter(["name"]):
                if (process.info.get("name") or "").lower() == "onedrive.exe":
                    counters = process.io_counters()
                    read_bytes += counters.read_bytes
                    write_bytes += counters.write_bytes
        else:
            read_bytes = int(powershell_float("(Get-Process OneDrive -ErrorAction SilentlyContinue | Measure-Object IOReadBytes -Sum).Sum", timeout=2) or 0)
            write_bytes = int(powershell_float("(Get-Process OneDrive -ErrorAction SilentlyContinue | Measure-Object IOWriteBytes -Sum).Sum", timeout=2) or 0)
            if read_bytes == 0 and write_bytes == 0:
                return {"status": "unknown", "download_rate": 0, "upload_rate": 0}
        now = time_module.monotonic()
        previous_time = ONEDRIVE_ACTIVITY_STATE.get("time")
        previous_read = ONEDRIVE_ACTIVITY_STATE.get("read")
        previous_write = ONEDRIVE_ACTIVITY_STATE.get("write")
        ONEDRIVE_ACTIVITY_STATE["time"] = now
        ONEDRIVE_ACTIVITY_STATE["read"] = read_bytes
        ONEDRIVE_ACTIVITY_STATE["write"] = write_bytes
        if previous_time is None or previous_read is None or previous_write is None:
            return {"status": "starting", "download_rate": 0, "upload_rate": 0}
        elapsed = max(0.2, now - previous_time)
        read_rate = max(0, int((read_bytes - previous_read) / elapsed))
        write_rate = max(0, int((write_bytes - previous_write) / elapsed))
        recent_paths = onedrive_recent_sync_paths(limit=2, seconds=600)
        upload_file = str(recent_paths[0]) if recent_paths and write_rate > 0 else "No active file detected"
        download_file = str(recent_paths[0]) if recent_paths and read_rate > 0 else "No active file detected"
        return {
            "status": "running",
            "download_rate": read_rate,
            "upload_rate": write_rate,
            "download_file": str(recent_paths[0]) if recent_paths and read_rate > 0 else "",
            "upload_file": str(recent_paths[0]) if recent_paths and write_rate > 0 else "",
            "download_label": download_file,
            "upload_label": upload_file,
        }
    except Exception:
        return {"status": "unknown", "download_rate": 0, "upload_rate": 0}


def onedrive_rate_to_activity_percent(rate):
    return max(0, min(100, int(rate * 100 / (2 * 1024 * 1024))))


def onedrive_recent_sync_paths(limit=3, seconds=180):
    roots = []
    for env_name in ("OneDrive", "OneDriveCommercial", "OneDriveConsumer"):
        value = os.environ.get(env_name)
        if value:
            roots.append(Path(value))
    roots.append(Path(r"D:\OneDrive"))
    root = next((path for path in roots if str(path) and path.exists()), None)
    if root is None:
        return []
    cutoff = time_module.time() - seconds
    found = []
    try:
        for path in root.rglob("*"):
            if len(found) >= 300:
                break
            try:
                if path.is_file():
                    mtime = path.stat().st_mtime
                    if mtime >= cutoff:
                        found.append((mtime, path))
            except Exception:
                pass
    except Exception:
        return []
    found.sort(reverse=True)
    return [path for _mtime, path in found[:limit]]


def onedrive_recent_sync_files(limit=3, seconds=180):
    return [path.name for path in onedrive_recent_sync_paths(limit, seconds)]


def onedrive_activity_lines():
    activity = one_drive_activity_text()
    if activity.startswith("Running | "):
        activity = activity.replace("Running | ", "", 1)
    elif activity.startswith("Running"):
        activity = "Activity starting"
    recent_files = onedrive_recent_sync_files(limit=2, seconds=600)
    if recent_files:
        names = ", ".join(ellipsize_middle(name, 26) for name in recent_files)
        return activity, f"Files: {names}"
    return activity, ""


def ellipsize_middle(text, max_chars=42):
    text = str(text)
    if len(text) <= max_chars:
        return text
    keep = max(8, (max_chars - 3) // 2)
    return f"{text[:keep]}...{text[-keep:]}"


def parse_time_to_seconds(value):
    value = value.strip()
    if not value:
        return None

    parts = value.split(":")
    if not all(part.isdigit() for part in parts):
        raise ValueError("Use seconds, MM:SS or HH:MM:SS.")

    numbers = [int(part) for part in parts]
    if len(numbers) == 1:
        return numbers[0]
    if len(numbers) == 2:
        minutes, seconds = numbers
        return minutes * 60 + seconds
    if len(numbers) == 3:
        hours, minutes, seconds = numbers
        return hours * 3600 + minutes * 60 + seconds

    raise ValueError("Use seconds, MM:SS or HH:MM:SS.")


def format_seconds_for_filename(seconds):
    if seconds is None:
        return "end"
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    if hours:
        return f"{hours:02d}-{minutes:02d}-{secs:02d}"
    return f"{minutes:02d}-{secs:02d}"


def recent_downloads(since_timestamp):
    if not DOWNLOADS_PATH.exists():
        return []

    exts = {".mp4", ".mkv", ".webm", ".m4a", ".mp3", ".part"}
    files = [
        path for path in DOWNLOADS_PATH.iterdir()
        if path.is_file() and path.suffix.lower() in exts and path.stat().st_mtime >= since_timestamp
    ]
    return sorted(files, key=lambda path: path.stat().st_mtime, reverse=True)


def find_tile_backgrounds():
    background_dir = get_tile_background_dir()
    if background_dir is None or Image is None or ImageTk is None or not background_dir.exists():
        TILE_BACKGROUND_CACHE["dir"] = None
        TILE_BACKGROUND_CACHE["paths"] = []
        return []

    cache_key = str(background_dir)
    if TILE_BACKGROUND_CACHE["dir"] == cache_key and TILE_BACKGROUND_CACHE["paths"]:
        return list(TILE_BACKGROUND_CACHE["paths"])

    paths = []
    started = time_module.monotonic()
    try:
        for path in background_dir.rglob("*"):
            if path.is_file() and path.suffix.lower() in TILE_BACKGROUND_EXTS:
                paths.append(path)
                if len(paths) >= TILE_BACKGROUND_SCAN_LIMIT:
                    break
            if time_module.monotonic() - started >= TILE_BACKGROUND_SCAN_SECONDS:
                break
    except Exception:
        paths = []

    TILE_BACKGROUND_CACHE["dir"] = cache_key
    TILE_BACKGROUND_CACHE["paths"] = paths
    return list(paths)


def image_cover(path, width=360, height=210, blend_color="#ffffff", image_opacity=0.50, corner_radius=0):
    image = Image.open(path).convert("RGB")
    width = max(120, int(width))
    height = max(90, int(height))
    image_ratio = image.width / image.height
    target_ratio = width / height

    if image_ratio > target_ratio:
        new_height = height
        new_width = int(height * image_ratio)
    else:
        new_width = width
        new_height = int(width / image_ratio)

    image = image.resize((new_width, new_height), Image.LANCZOS)
    left = max(0, (new_width - width) // 2)
    top = max(0, (new_height - height) // 2)
    image = image.crop((left, top, left + width, top + height))
    wash = Image.new("RGB", image.size, blend_color)
    image = Image.blend(wash, image, image_opacity)
    if corner_radius and ImageDraw is not None:
        image = image.convert("RGBA")
        mask = Image.new("L", image.size, 0)
        draw = ImageDraw.Draw(mask)
        draw.rounded_rectangle((0, 0, image.width - 1, image.height - 1), radius=int(corner_radius), fill=255)
        image.putalpha(mask)
    return ImageTk.PhotoImage(image)


def image_for_tile(path, width=360, height=210):
    return image_cover(path, width, height, "#ffffff", get_tile_artwork_opacity_percent() / 100, corner_radius=10)


def image_for_meter(path, width=360, height=210, blend_color="#ffffff"):
    return image_cover(path, width, height, blend_color, get_vu_artwork_opacity_percent() / 100, corner_radius=10)


def image_for_meter_full(path, width=360, height=210):
    return image_cover(path, width, height, "#ffffff", 1.0, corner_radius=10)


def image_for_pane(path, width=720, height=240, blend_color="#ffffff"):
    return image_cover(path, width, height, blend_color, 0.22, corner_radius=10)


def rounded_rect(canvas, x1, y1, x2, y2, radius=10, **kwargs):
    radius = max(0, min(radius, abs(x2 - x1) / 2, abs(y2 - y1) / 2))
    if radius <= 0:
        return canvas.create_rectangle(x1, y1, x2, y2, **kwargs)
    points = [
        x1 + radius, y1,
        x2 - radius, y1,
        x2, y1,
        x2, y1 + radius,
        x2, y2 - radius,
        x2, y2,
        x2 - radius, y2,
        x1 + radius, y2,
        x1, y2,
        x1, y2 - radius,
        x1, y1 + radius,
        x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, splinesteps=12, **kwargs)


ICON_IMAGE_CACHE = {}


def favicon_url_for_domain(domain, size=64):
    clean = str(domain or "").strip()
    if not clean:
        return ""
    return f"https://www.google.com/s2/favicons?domain={urllib.parse.quote(clean)}&sz={int(size)}"


def safe_icon_cache_name(icon_url):
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", icon_url)[:180] + ".png"


def load_tile_icon_image(icon_url, size=56):
    if Image is None or ImageTk is None or not icon_url:
        return None
    key = (icon_url, int(size))
    if key in ICON_IMAGE_CACHE:
        return ICON_IMAGE_CACHE[key]

    try:
        ICON_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    cache_path = ICON_CACHE_DIR / safe_icon_cache_name(icon_url)
    data = None
    if cache_path.exists():
        try:
            data = cache_path.read_bytes()
        except Exception:
            data = None
    if data is None:
        try:
            request = urllib.request.Request(icon_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(request, timeout=3) as response:
                data = response.read(512000)
            try:
                cache_path.write_bytes(data)
            except Exception:
                pass
        except Exception:
            return None
    try:
        image = Image.open(io.BytesIO(data)).convert("RGBA")
        image.thumbnail((size, size), Image.LANCZOS)
        canvas_image = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        x = (size - image.width) // 2
        y = (size - image.height) // 2
        canvas_image.alpha_composite(image, (x, y))
        photo = ImageTk.PhotoImage(canvas_image)
        ICON_IMAGE_CACHE[key] = photo
        return photo
    except Exception:
        return None


def safe_media_name(text):
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", text)
    cleaned = " ".join(cleaned.split()).strip()
    return cleaned[:150] or "youtube_segment"


def find_ffmpeg_exe():
    bundled = resource_path("ffmpeg", "ffmpeg.exe")
    if bundled.exists():
        return str(bundled)
    bundled = Path(r"C:\ffmpeg\bin\ffmpeg.exe")
    if bundled.exists():
        return str(bundled)
    found = shutil.which("ffmpeg")
    return found


def resolution_from_info(info):
    width = info.get("width")
    height = info.get("height")

    requested = info.get("requested_downloads") or []
    for item in requested:
        width = width or item.get("width")
        height = height or item.get("height")
        if width and height:
            break

    if width and height:
        return f"{int(width)}x{int(height)}"
    if height:
        return f"{int(height)}p"
    return "unknown-resolution"


def release_note_headings(content):
    headings = []
    for line_number, line in enumerate(content.splitlines(), start=1):
        text = line.strip()
        if not text:
            continue
        is_heading = (
            text.startswith("#")
            or text.lower().startswith(("v", "version", "release", "changes", "new ", "fixed", "added"))
            or text.isupper()
        )
        if is_heading and len(text) <= 90:
            headings.append((text.lstrip("# ").strip(), line_number))
    return headings[:30]


def load_metrics_rows():
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl is needed for metrics: {exc}")

    workbook_path = get_metrics_workbook_path()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Metrics workbook not found:\n{workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["data entry"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            continue
        rows.append({
            "week": row[0],
            "date": row[1],
            "tt_views": row[3] or 0,
            "tt_followers": row[4] or 0,
            "ig_views": row[6] or 0,
            "ig_followers": row[7] or 0,
            "yt_views": row[9] or 0,
            "yt_followers": row[10] or 0,
            "fb_views": row[12] or 0,
            "fb_followers": row[13] or 0,
        })
    return rows


def metrics_totals(rows):
    totals = []
    for row in rows:
        total_views = row["tt_views"] + row["ig_views"] + row["yt_views"] + row["fb_views"]
        total_followers = row["tt_followers"] + row["ig_followers"] + row["yt_followers"] + row["fb_followers"]
        totals.append({**row, "total_views": total_views, "total_impressions": total_views, "total_followers": total_followers})
    return totals


def completed_week_number(update_sunday):
    week_number = int(update_sunday.isocalendar().week) - 1
    if week_number < 1:
        week_number = int((update_sunday - timedelta(days=7)).isocalendar().week)
    return week_number


def metric_month_label(value):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%b")
    return ""


def append_metrics_row(values):
    from openpyxl import load_workbook

    workbook_path = get_metrics_workbook_path()
    wb = load_workbook(workbook_path)
    ws = wb["data entry"]
    next_row = ws.max_row + 1
    ws.cell(next_row, 1).value = values["week"]
    ws.cell(next_row, 2).value = values["date"]
    ws.cell(next_row, 4).value = values["tt_views"]
    ws.cell(next_row, 5).value = values["tt_followers"]
    ws.cell(next_row, 7).value = values["ig_views"]
    ws.cell(next_row, 8).value = values["ig_followers"]
    ws.cell(next_row, 10).value = values["yt_views"]
    ws.cell(next_row, 11).value = values["yt_followers"]
    ws.cell(next_row, 13).value = values["fb_views"]
    ws.cell(next_row, 14).value = values["fb_followers"]
    wb.save(workbook_path)


MONTHLY_HEADERS = [
    "Month",
    "Facebook Impressions", "Instagram Impressions", "TikTok Impressions", "YouTube Impressions",
    "Facebook Followers", "Instagram Followers", "TikTok Followers", "YouTube Followers",
    "Facebook Posts", "Instagram Posts", "TikTok Posts", "YouTube Posts",
    "Facebook Interactions", "Instagram Interactions", "TikTok Interactions", "YouTube Interactions",
    "Dashboard Summary", "Next Month Focus",
]


def ensure_monthly_worksheet(wb):
    monthly_name = next((name for name in wb.sheetnames if name.lower() == "monthly"), None)
    ws = wb[monthly_name] if monthly_name else wb.create_sheet("Monthly")
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for index, header in enumerate(MONTHLY_HEADERS, start=1):
            ws.cell(1, index).value = header
    return ws


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def normalize_month_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    if not text:
        return ""
    parsed = None
    for fmt in ("%Y-%m", "%Y/%m", "%m-%Y", "%m/%Y", "%B %Y", "%b %Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            break
        except Exception:
            pass
    if parsed is not None:
        return parsed.strftime("%Y-%m")
    return text


MONTHLY_FIELD_ALIASES = {
    "month": ["month", "maand", "period", "periode"],
    "fb_impressions": ["facebook impressions", "facebook impression", "fb impressions", "fb impression", "facebook views", "fb views"],
    "ig_impressions": ["instagram impressions", "instagram impression", "ig impressions", "ig impression", "instagram views", "ig views"],
    "tt_impressions": ["tiktok impressions", "tiktok impression", "tt impressions", "tt impression", "tiktok views", "tt views"],
    "yt_impressions": ["youtube impressions", "youtube impression", "yt impressions", "yt impression", "youtube views", "yt views"],
    "fb_followers": ["facebook followers", "fb followers"],
    "ig_followers": ["instagram followers", "ig followers"],
    "tt_followers": ["tiktok followers", "tt followers"],
    "yt_followers": ["youtube followers", "yt followers"],
    "fb_posts": ["facebook posts", "fb posts"],
    "ig_posts": ["instagram posts", "ig posts"],
    "tt_posts": ["tiktok posts", "tt posts"],
    "yt_posts": ["youtube posts", "yt posts"],
    "fb_interactions": ["facebook interactions", "fb interactions", "facebook engagement", "fb engagement"],
    "ig_interactions": ["instagram interactions", "ig interactions", "instagram engagement", "ig engagement"],
    "tt_interactions": ["tiktok interactions", "tt interactions", "tiktok engagement", "tt engagement"],
    "yt_interactions": ["youtube interactions", "yt interactions", "youtube engagement", "yt engagement"],
    "dashboard_summary": ["dashboard summary", "summary", "samenvatting"],
    "next_month_focus": ["next month focus", "focus", "next focus", "volgende maand focus"],
}


def monthly_header_map(ws):
    headers = [normalize_header(cell.value) for cell in ws[1]]
    lookup = {header: index for index, header in enumerate(headers) if header}
    result = {}
    for key, aliases in MONTHLY_FIELD_ALIASES.items():
        for alias in aliases:
            normalized = normalize_header(alias)
            if normalized in lookup:
                result[key] = lookup[normalized]
                break
    return result


def load_monthly_rows():
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    workbook_path = get_metrics_workbook_path()
    if not workbook_path.exists():
        return []
    try:
        wb_edit = load_workbook(workbook_path)
        if not any(name.lower() == "monthly" for name in wb_edit.sheetnames):
            ensure_monthly_worksheet(wb_edit)
            wb_edit.save(workbook_path)
        wb_edit.close()
    except Exception:
        pass
    wb = load_workbook(workbook_path, data_only=True)
    ws = ensure_monthly_worksheet(wb)
    if ws.max_row <= 1:
        return []

    keys = [
        "month",
        "fb_impressions", "ig_impressions", "tt_impressions", "yt_impressions",
        "fb_followers", "ig_followers", "tt_followers", "yt_followers",
        "fb_posts", "ig_posts", "tt_posts", "yt_posts",
        "fb_interactions", "ig_interactions", "tt_interactions", "yt_interactions",
        "dashboard_summary", "next_month_focus",
    ]
    header_map = monthly_header_map(ws)
    if "month" not in header_map:
        header_map = {key: index for index, key in enumerate(keys)}

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not values:
            continue
        raw_month = values[header_map["month"]] if header_map["month"] < len(values) else None
        month = normalize_month_value(raw_month)
        if not month:
            continue
        row = {}
        for key in keys:
            index = header_map.get(key)
            value = values[index] if index is not None and index < len(values) else None
            if key in ("month", "dashboard_summary", "next_month_focus"):
                row[key] = month if key == "month" else ("" if value is None else str(value))
            else:
                try:
                    row[key] = int(float(value or 0))
                except Exception:
                    row[key] = 0
        rows.append(row)
    rows.sort(key=lambda item: item["month"])
    return rows


def format_ics_datetime(dt):
    return dt.strftime("%Y%m%dT%H%M%S")


def safe_filename(text):
    return "_".join(text.strip().lower().split())


def display_song_filename(text):
    return " ".join(text.strip().split())


def clean_multiline(text):
    return text.strip() if text.strip() else "-"


def hashtags_to_single_line(hashtags):
    return " ".join(hashtags.split())


def build_caption(song_title, song_description, hashtags):
    return (
        f"{ARTIST_NAME} - {song_title}\n\n"
        f"{song_description.strip()}\n\n"
        f"LukeStrom\n\n"
        f"🌌 Welcome to the MeloVerse\n\n"
        f"🎵 Full Songs⭐⭐⭐⭐⭐ → {URL_MTRBIO}\n\n"
        f"{hashtags_to_single_line(hashtags)}"
    )


def get_output_filenames(song_title):
    today = date.today()
    safe_title = safe_filename(song_title)
    date_prefix = today.strftime("%y%m%d")
    return {
        "ics_filename": f"{date_prefix}_{safe_title}_catalog.ics",
        "schedule_txt_filename": f"{date_prefix}_{safe_title}_post_schedule.txt",
        "design_txt_filename": f"{date_prefix}_{safe_title}_reel_design.txt",
    }


def create_ics(events):
    now_stamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Lukestrom//Post Planner//EN", "CALSCALE:GREGORIAN", "METHOD:PUBLISH"]

    for event in events:
        start_dt = event["datetime"]
        end_dt = start_dt + timedelta(minutes=30)
        description = event["description"].replace("\n", "\\n")
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uuid.uuid4()}@lukestrom.local",
            f"DTSTAMP:{now_stamp}",
            f"DTSTART:{format_ics_datetime(start_dt)}",
            f"DTEND:{format_ics_datetime(end_dt)}",
            f"SUMMARY:{event['title']}",
            f"DESCRIPTION:{description}",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    return "\n".join(lines)


def day_offset_for_post(post_index, post_offsets=None):
    offsets = post_offsets or FIRST_POST_OFFSETS
    if post_index <= len(offsets):
        return offsets[post_index - 1]
    return offsets[-1] + ((post_index - len(offsets)) * 7)


def day_offset_for_repost(repost_index, post_index, post_offsets=None, repost_offsets=None):
    reposts = repost_offsets or REPOST_START_OFFSETS
    if repost_index <= len(reposts):
        base = reposts[repost_index - 1]
    else:
        base = reposts[-1] + ((repost_index - len(reposts)) * 35)
    return base + day_offset_for_post(post_index, post_offsets)


def build_campaign_items(song_title, post_count=4, repost_count=4, start_date=None, post_offsets=None, repost_offsets=None):
    filename_song = display_song_filename(song_title)
    items = []
    start_date = start_date or date.today()

    for post_number in range(1, post_count + 1):
        items.append({
            "label": f"Post {post_number}",
            "post_number": post_number,
            "date": start_date + timedelta(days=day_offset_for_post(post_number, post_offsets)),
            "filename": f"post {post_number} - {filename_song}.mp4",
        })

    for repost_number in range(1, repost_count + 1):
        for post_number in range(1, post_count + 1):
            items.append({
                "label": f"Repost {repost_number} - Post {post_number}",
                "post_number": post_number,
                "date": start_date + timedelta(days=day_offset_for_repost(repost_number, post_number, post_offsets, repost_offsets)),
                "filename": f"post {post_number} - {filename_song}.mp4",
            })

    return items


def build_post_schedule_text(song_title, song_description, campaign_items, hashtags):
    caption = build_caption(song_title, song_description, hashtags)
    item_count = len(campaign_items)
    lines = [
        f"Song: {song_title}", f"Artist: {ARTIST_NAME}", "", "POST SCHEDULE", "",
        "General instructions:",
        f"- This file contains {item_count} content moments and {item_count} scheduling tasks.",
        "- Metricool handles cross-posting.",
        "- Each block is one post file with all initial and repost dates grouped together.",
        "- Choose the best time in Metricool while scheduling.",
        "- The first 2 seconds and opening text are already baked into the video.",
        "- Use the filename, dates, and caption exactly as shown.",
        "", "URLs:", f"DistroKid: {URL_DISTROKID}", f"Meta Business Suite: {URL_FACEBOOK}",
        f"YouTube Uploads: {URL_YOUTUBE}", f"TikTok Studio: {URL_TIKTOK}", "", "=" * 70, "",
    ]

    post_numbers = sorted({item["post_number"] for item in campaign_items})
    for post_number in post_numbers:
        items = [item for item in campaign_items if item["post_number"] == post_number]
        initial_items = [item for item in items if item["label"] == f"Post {post_number}"]
        repost_items = [item for item in items if item["label"].startswith("Repost ")]
        ordered_items = initial_items + repost_items
        if not ordered_items:
            continue
        filename = ordered_items[0]["filename"]
        for item in ordered_items:
            if item["label"].startswith("Repost "):
                label = item["label"].replace(" - ", " ").lower().capitalize()
            else:
                label = item["label"]
            lines.append(f"{label} - Date: {item['date'].strftime('%d/%m/%Y')}")
        lines.extend([
            "",
            f"Filename: {filename}",
            "Caption:",
            caption,
            "",
            "-" * 70,
            "",
        ])
    return "\n".join(lines)


def build_reel_design_text(song_title, creative_items):
    lines = [
        f"Song: {song_title}", f"Artist: {ARTIST_NAME}", "", "REEL DESIGN", "",
        "Use this file to create the four reel/video files before scheduling the campaign.",
        "Only Post 1 to Post 4 need creative notes.",
        "Reposts use the same video files.",
        "",
        "Reel guideline:",
        "- one striking visual",
        "- immediate branding",
        "- quick montage",
        "- then settle into the story",
        "", "=" * 70, "",
    ]

    for item in creative_items:
        lines.extend([
            item["label"], "", "Filename:", item["filename"], "",
            "First 2 seconds:", clean_multiline(item["first_seconds"]), "",
            "Opening text:", clean_multiline(item["opening_text"]), "", "=" * 70, "",
        ])
    return "\n".join(lines)


def build_events(song_title, schedule_txt_filename, design_txt_filename, item_count=20, start_date=None):
    start_date = start_date or date.today()
    return [
        {
            "datetime": datetime.combine(start_date, POST_TIME),
            "title": f"{song_title} - Publish on DistroKid",
            "description": f"Song: {song_title}\nArtist: {ARTIST_NAME}\nTask: Publish on DistroKid\n\nUpload the song to DistroKid.\nCheck title, artist name, credits, lyrics, artwork, and AI/vocal disclosure settings.\n\nURL:\n{URL_DISTROKID}",
        },
        {
            "datetime": datetime.combine(start_date, POST_TIME),
            "title": f"{song_title} - Schedule campaign",
            "description": f"Song: {song_title}\nArtist: {ARTIST_NAME}\nTask: Schedule full social media campaign\n\nOpen the schedule file:\n{schedule_txt_filename}\n\nCreative reference file:\n{design_txt_filename}\n\nThis campaign contains {item_count} content moments and {item_count} scheduling tasks.\nMetricool handles cross-posting.\nSchedule everything at 20:00.\n\nMeta Business Suite: {URL_FACEBOOK}\nYouTube Uploads: {URL_YOUTUBE}\nTikTok Studio: {URL_TIKTOK}",
        },
        {
            "datetime": datetime.combine(start_date + timedelta(days=1), POST_TIME),
            "title": f"{song_title} - Cleanup",
            "description": f"Song: {song_title}\nArtist: {ARTIST_NAME}\nTask: Cleanup\n\n- Verify all scheduled posts were created.\n- Check the schedule file again if needed: {schedule_txt_filename}\n- Store the creative reel design file: {design_txt_filename}\n- Archive project files clearly.\n- Remove temporary reel files if no longer needed.",
        },
    ]


def load_song_analyzer_module():
    if not SONG_ANALYZER_PATH.exists():
        raise FileNotFoundError(f"Song Analyzer not found:\n{SONG_ANALYZER_PATH}")

    module_dir = str(SONG_ANALYZER_PATH.parent)
    if module_dir not in sys.path:
        sys.path.insert(0, module_dir)

    spec = importlib.util.spec_from_file_location("toolbox_song_analyzer", SONG_ANALYZER_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class ScrollFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.auto_scrollbar = False
        self.canvas = tk.Canvas(self, highlightthickness=0, bg="#f5f3ee")
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.body = ttk.Frame(self.canvas)
        self.window = self.canvas.create_window((0, 0), window=self.body, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.body.bind("<Configure>", self._on_body_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

    def set_auto_scrollbar(self, enabled=True):
        self.auto_scrollbar = enabled
        self._update_scrollbar_visibility()

    def _on_body_configure(self, _event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self._update_scrollbar_visibility()

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.window, width=event.width)
        self._update_scrollbar_visibility()

    def _update_scrollbar_visibility(self):
        if not self.auto_scrollbar:
            return
        try:
            bbox = self.canvas.bbox("all")
            needed = bool(bbox and bbox[3] > self.canvas.winfo_height())
            mapped = bool(self.scrollbar.winfo_ismapped())
            if needed and not mapped:
                self.scrollbar.pack(side="right", fill="y")
            elif not needed and mapped:
                self.scrollbar.pack_forget()
        except Exception:
            pass


class CreativeToolbox(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.app_icon = None
        self.geometry("1180x780")
        self.minsize(980, 680)
        self.dark_mode = get_dark_mode()
        self.colors = theme_colors(self.dark_mode)
        self.configure(bg=self.colors["app_bg"])
        self.current_page = None
        self.hack_windows_running = False
        self.hack_windows = []
        self.hack_window_job = None
        self._setup_style()
        self._set_app_icon()
        self._build_shell()
        self.show_home()

    def _set_app_icon(self):
        if Image is None or ImageTk is None or not APP_ICON_PATH.exists():
            return
        try:
            image = Image.open(APP_ICON_PATH).resize((64, 64), Image.LANCZOS)
            self.app_icon = ImageTk.PhotoImage(image)
            self.iconphoto(True, self.app_icon)
        except Exception:
            self.app_icon = None

    def _setup_style(self):
        colors = self.colors
        self.style = ttk.Style(self)
        self.style.theme_use("clam")
        self.style.configure(".", font=("Segoe UI", 10), background=colors["app_bg"], foreground=colors["text"])
        self.style.configure("TFrame", background=colors["app_bg"])
        self.style.configure("Panel.TFrame", background=colors["panel_bg"])
        self.style.configure("Selected.TFrame", background=colors["nav_active"])
        self.style.configure("Title.TLabel", font=("Segoe UI", 22, "bold"), foreground=colors["text"], background=colors["app_bg"])
        self.style.configure("PageTitle.TLabel", font=("Segoe UI", 18, "bold"), foreground=colors["text"], background=colors["app_bg"])
        self.style.configure("Muted.TLabel", foreground=colors["muted"], background=colors["app_bg"])
        self.style.configure("CardTitle.TLabel", font=("Segoe UI", 15, "bold"), background=colors["panel_bg"], foreground=colors["text"])
        self.style.configure("CardText.TLabel", background=colors["panel_bg"], foreground=colors["muted"])
        self.style.configure("TButton", padding=(14, 8), font=("Segoe UI", 10))
        self.style.configure("Accent.TButton", padding=(16, 9), font=("Segoe UI", 10, "bold"))
        self.style.configure("TEntry", padding=6, fieldbackground=colors["entry_bg"], foreground=colors["entry_fg"])
        self.style.configure("TSpinbox", padding=4, fieldbackground=colors["entry_bg"], foreground=colors["entry_fg"], arrowcolor=colors["text"])
        self.style.map(
            "TSpinbox",
            fieldbackground=[("readonly", colors["entry_bg"]), ("!disabled", colors["entry_bg"])],
            foreground=[("readonly", colors["entry_fg"]), ("!disabled", colors["entry_fg"])],
        )
        self.style.configure("NavHome.TButton", padding=(14, 7), font=("Segoe UI", 9), background="#111111", foreground="#ffffff", borderwidth=0)
        self.style.map("NavHome.TButton", background=[("active", "#333333")], foreground=[("active", "#ffffff")])
        self.style.configure("NavLinks.TButton", padding=(12, 7), font=("Segoe UI", 9), background=colors["nav_bg"], foreground=colors["text"], borderwidth=0, focuscolor=colors["nav_bg"])
        self.style.map("NavLinks.TButton", background=[("active", colors["nav_active"])], foreground=[("active", colors["text"])])
        self.style.configure("NavNotes.TButton", padding=(12, 7), font=("Segoe UI", 9), background=colors["notes_bg"], foreground=colors["text"], borderwidth=0)
        self.style.map("NavNotes.TButton", background=[("active", colors["notes_active"])], foreground=[("active", colors["text"])])
        self.style.configure("NavText.TLabel", padding=(8, 6), font=("Segoe UI", 10), background=colors["app_bg"], foreground=colors["text"])
        self.style.configure("Version.TLabel", padding=(8, 6), font=("Segoe UI", 10, "bold"), background=colors["app_bg"], foreground=colors["text"])

    def _build_shell(self):
        self.topbar = ttk.Frame(self, padding=(22, 12, 22, 10))
        self.topbar.pack(fill="x")
        self.topbar.columnconfigure(1, weight=1)

        self.home_button = self.create_nav_text(self.topbar, "Home", self.show_home)
        self.home_button.grid(row=0, column=0, sticky="w", padx=(0, 6))

        tool_nav = ttk.Frame(self.topbar)
        tool_nav.grid(row=0, column=1, sticky="w")
        tool_buttons = [
            ("Song", self.show_song_analyzer),
            ("Reel", self.show_reel_design),
            ("Campaign", self.show_post_planner),
            ("Tools", self.show_audio_vu),
            ("Shortcuts", self.show_tools),
            ("Metrics", self.show_metrics),
        ]
        for text, command in tool_buttons:
            self.create_nav_text(tool_nav, text, command).pack(side="left", padx=(0, 10))

        right_nav = ttk.Frame(self.topbar)
        right_nav.grid(row=0, column=2, sticky="e")
        self.refresh_button = ttk.Button(right_nav, text="↻", width=3, command=self.refresh_active_page, style="NavLinks.TButton", takefocus=False)
        self.refresh_button.pack(side="left", padx=(0, 6))
        self.menu_button = ttk.Button(right_nav, text="☰", width=3, command=self.show_settings_menu, style="NavLinks.TButton", takefocus=False)
        self.menu_button.pack(side="left")

        self.content = ttk.Frame(self, padding=(22, 8, 22, 22))
        self.content.pack(fill="both", expand=True)

    def create_nav_text(self, parent, text, command):
        label = ttk.Label(parent, text=text, style="NavText.TLabel", cursor="hand2")
        label.bind("<Button-1>", lambda _event: command())
        label.bind("<Enter>", lambda _event: label.configure(font=("Segoe UI", 10, "underline")))
        label.bind("<Leave>", lambda _event: label.configure(font=("Segoe UI", 10)))
        return label

    def show_settings_menu(self):
        menu = tk.Menu(self, tearoff=False)
        dark_var = tk.BooleanVar(value=self.dark_mode)
        menu.add_checkbutton(label="Dark mode", variable=dark_var, command=self.toggle_dark_mode)
        fullscreen_var = tk.BooleanVar(value=self.state() == "zoomed")
        menu.add_checkbutton(label="Fullscreen", variable=fullscreen_var, command=self.toggle_fullscreen)
        menu.add_separator()
        menu.add_command(label="Artwork", state="disabled")
        menu.add_checkbutton(label="No picture", variable=tk.BooleanVar(value=load_settings().get("tile_background_disabled", False)), command=self.toggle_tile_pictures)
        menu.add_command(label="Pictures location", command=self.choose_background_folder)
        menu.add_command(label="Interval pics", command=lambda: self.choose_interval("tile_picture_interval_seconds", "Picture interval", get_tile_picture_interval_seconds()))
        menu.add_command(label="Tile opacity", command=lambda: self.choose_opacity("tile_artwork_opacity_percent", "Tile opacity", get_tile_artwork_opacity_percent()))
        menu.add_command(label="VU opacity", command=lambda: self.choose_opacity("vu_artwork_opacity_percent", "VU opacity", get_vu_artwork_opacity_percent()))
        menu.add_separator()
        menu.add_command(label="Fonts", state="disabled")
        menu.add_checkbutton(label="Basic font", variable=tk.BooleanVar(value=use_basic_fonts_only()), command=self.toggle_basic_font)
        menu.add_command(label="Font location", command=self.choose_font_folder)
        menu.add_command(label="Interval fonts", command=lambda: self.choose_interval("tile_font_interval_seconds", "Font interval", get_tile_font_interval_seconds()))
        menu.add_command(label="Font size", command=self.choose_font_size_range)
        menu.add_separator()
        menu.add_command(label="Metrics", state="disabled")
        menu.add_command(label="Excel location", command=self.choose_metrics_workbook)
        menu.add_command(label="Export Excel template", command=self.export_metrics_template)
        menu.add_separator()
        menu.add_command(label="Info", state="disabled")
        menu.add_command(label="Release notes", command=self.show_release_notes)
        menu.add_command(label="How to", command=self.show_how_to)
        x = self.menu_button.winfo_rootx()
        y = self.menu_button.winfo_rooty() + self.menu_button.winfo_height()
        menu.tk_popup(x, y)

    def toggle_tile_pictures(self):
        settings = load_settings()
        settings["tile_background_disabled"] = not settings.get("tile_background_disabled", False)
        save_settings(settings)
        if isinstance(self.current_page, HomePage):
            self.show_home()
        elif isinstance(self.current_page, ToolsPage):
            self.show_tools()
        elif isinstance(self.current_page, VuPage):
            self.show_audio_vu()

    def toggle_basic_font(self):
        settings = load_settings()
        settings["basic_font_only"] = not settings.get("basic_font_only", False)
        save_settings(settings)
        if isinstance(self.current_page, HomePage):
            self.show_home()

    def toggle_dark_mode(self):
        settings = load_settings()
        settings["dark_mode"] = not self.dark_mode
        save_settings(settings)
        self.dark_mode = not self.dark_mode
        self.colors = theme_colors(self.dark_mode)
        self.configure(bg=self.colors["app_bg"])
        self._setup_style()
        self.refresh_current_page()

    def toggle_fullscreen(self):
        if self.state() == "zoomed":
            self.state("normal")
        else:
            self.state("zoomed")

    def refresh_current_page(self):
        current = self.current_page
        if isinstance(current, SongAnalyzerPage):
            self.show_song_analyzer()
        elif isinstance(current, ReelDesignPage):
            self.show_reel_design()
        elif isinstance(current, CampaignPlannerPage):
            self.show_post_planner()
        elif isinstance(current, YouTubeDownloaderPage):
            self.show_youtube_downloader()
        elif isinstance(current, VuPage):
            self.show_audio_vu()
        elif isinstance(current, ToolsPage):
            self.show_tools()
        elif isinstance(current, MetricsPage):
            self.show_metrics()
        else:
            self.show_home()

    def refresh_active_page(self):
        current = self.current_page
        if hasattr(current, "refresh"):
            try:
                current.refresh()
                return
            except TypeError:
                pass
        if isinstance(current, VuPage):
            current.refresh_performance()
        else:
            self.refresh_current_page()

    def style_text_widget(self, widget):
        widget.configure(
            bg=self.colors["entry_bg"],
            fg=self.colors["entry_fg"],
            insertbackground=self.colors["entry_fg"],
            selectbackground="#4f6f8f",
            selectforeground="#ffffff",
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["border"],
        )

    def style_chart_canvas(self, canvas):
        canvas.configure(bg=self.colors["chart_bg"], highlightbackground=self.colors["border"])

    def choose_background_folder(self):
        initial_dir = get_tile_background_dir()
        selected = filedialog.askdirectory(
            title="Choose artwork folder",
            initialdir=str(initial_dir if initial_dir is not None and initial_dir.exists() else Path.home()),
        )
        if not selected:
            return

        settings = load_settings()
        settings["tile_background_dir"] = selected
        settings["tile_background_disabled"] = False
        save_settings(settings)

        if isinstance(self.current_page, HomePage):
            self.show_home()
        elif isinstance(self.current_page, ToolsPage):
            self.show_tools()
        elif isinstance(self.current_page, VuPage):
            self.show_audio_vu()

    def choose_font_folder(self):
        initial_dir = get_tile_font_dir() or Path.home()
        selected = filedialog.askdirectory(
            title="Choose font folder",
            initialdir=str(initial_dir if initial_dir.exists() else Path.home()),
        )
        if not selected:
            return

        settings = load_settings()
        settings["tile_font_dir"] = selected
        save_settings(settings)

        if isinstance(self.current_page, HomePage):
            self.show_home()

    def choose_interval(self, setting_key, title, current_value):
        window = tk.Toplevel(self)
        window.title(title)
        window.resizable(False, False)
        window.transient(self)

        shell = ttk.Frame(window, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True)
        suggested = 20 if setting_key == "tile_picture_interval_seconds" else 15
        initial_value = suggested
        label_var = tk.StringVar(value=f"{initial_value} seconds")

        ttk.Label(shell, text=title, style="CardTitle.TLabel").pack(anchor="w")
        scale = ttk.Scale(
            shell,
            from_=5,
            to=60,
            orient="horizontal",
            value=initial_value,
            length=260,
            command=lambda value: label_var.set(f"{int(float(value))} seconds"),
        )
        scale.pack(fill="x", pady=(14, 6))
        marker = tk.Canvas(shell, height=14, width=260, highlightthickness=0, bg=self.colors["panel_bg"])
        marker.pack(fill="x")
        marker_x = int((suggested - 5) / (60 - 5) * 260)
        marker.create_line(marker_x, 0, marker_x, 12, fill=self.colors["text"], width=2)
        ttk.Label(shell, textvariable=label_var, style="CardText.TLabel").pack(anchor="w")

        buttons = ttk.Frame(shell, style="Panel.TFrame")
        buttons.pack(fill="x", pady=(16, 0))

        def save_interval():
            seconds = int(float(scale.get()))
            settings = load_settings()
            settings[setting_key] = seconds
            save_settings(settings)
            window.destroy()
            if isinstance(self.current_page, HomePage):
                self.show_home()
            elif isinstance(self.current_page, VuPage):
                self.show_audio_vu()

        ttk.Button(buttons, text="OK", command=save_interval, style="Accent.TButton").pack(side="left", fill="x", expand=True)
        ttk.Button(buttons, text="Cancel", command=window.destroy).pack(side="left", fill="x", expand=True, padx=(10, 0))

        self.update_idletasks()
        width = 380
        height = 230
        x = self.winfo_rootx() + (self.winfo_width() - width) // 2
        y = self.winfo_rooty() + (self.winfo_height() - height) // 2
        window.geometry(f"{width}x{height}+{x}+{y}")

    def choose_opacity(self, setting_key, title, current_value):
        window = tk.Toplevel(self)
        window.title(title)
        window.resizable(False, False)
        window.transient(self)

        shell = ttk.Frame(window, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True)
        label_var = tk.StringVar(value=f"{current_value}%")

        ttk.Label(shell, text=title, style="CardTitle.TLabel").pack(anchor="w")
        scale = ttk.Scale(
            shell,
            from_=0,
            to=100,
            orient="horizontal",
            value=current_value,
            length=280,
            command=lambda value: label_var.set(f"{int(float(value))}%"),
        )
        scale.pack(fill="x", pady=(14, 6))
        ttk.Label(shell, textvariable=label_var, style="CardText.TLabel").pack(anchor="w")

        buttons = ttk.Frame(shell, style="Panel.TFrame")
        buttons.pack(fill="x", pady=(16, 0))

        def save_opacity():
            settings = load_settings()
            settings[setting_key] = int(float(scale.get()))
            save_settings(settings)
            window.destroy()
            if isinstance(self.current_page, (HomePage, ToolsPage, VuPage)):
                self.refresh_current_page()

        ttk.Button(buttons, text="OK", command=save_opacity, style="Accent.TButton").pack(side="left", fill="x", expand=True)
        ttk.Button(buttons, text="Cancel", command=window.destroy).pack(side="left", fill="x", expand=True, padx=(10, 0))

        self.update_idletasks()
        width = 380
        height = 210
        x = self.winfo_rootx() + (self.winfo_width() - width) // 2
        y = self.winfo_rooty() + (self.winfo_height() - height) // 2
        window.geometry(f"{width}x{height}+{x}+{y}")

    def choose_font_size_range(self):
        window = tk.Toplevel(self)
        window.title("Font size")
        window.resizable(False, False)
        window.transient(self)

        min_value, max_value = get_tile_font_size_range()
        shell = ttk.Frame(window, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True)
        ttk.Label(shell, text="Font size", style="CardTitle.TLabel").pack(anchor="w")

        min_label = tk.StringVar(value=f"Minimum: {min_value}%")
        max_label = tk.StringVar(value=f"Maximum: {max_value}%")

        ttk.Label(shell, textvariable=min_label, style="CardText.TLabel").pack(anchor="w", pady=(14, 0))
        min_scale = ttk.Scale(shell, from_=100, to=300, orient="horizontal", value=min_value, length=280, command=lambda value: min_label.set(f"Minimum: {int(float(value))}%"))
        min_scale.pack(fill="x", pady=(4, 8))

        ttk.Label(shell, textvariable=max_label, style="CardText.TLabel").pack(anchor="w")
        max_scale = ttk.Scale(shell, from_=100, to=300, orient="horizontal", value=max_value, length=280, command=lambda value: max_label.set(f"Maximum: {int(float(value))}%"))
        max_scale.pack(fill="x", pady=(4, 12))

        buttons = ttk.Frame(shell, style="Panel.TFrame")
        buttons.pack(fill="x", pady=(8, 0))

        def save_range():
            minimum = int(float(min_scale.get()))
            maximum = int(float(max_scale.get()))
            if minimum > maximum:
                minimum, maximum = maximum, minimum
            settings = load_settings()
            settings["tile_font_size_min_percent"] = minimum
            settings["tile_font_size_max_percent"] = maximum
            save_settings(settings)
            window.destroy()
            if isinstance(self.current_page, HomePage):
                self.show_home()

        ttk.Button(buttons, text="OK", command=save_range, style="Accent.TButton").pack(side="left", fill="x", expand=True)
        ttk.Button(buttons, text="Cancel", command=window.destroy).pack(side="left", fill="x", expand=True, padx=(10, 0))

        self.update_idletasks()
        width = 380
        height = 260
        x = self.winfo_rootx() + (self.winfo_width() - width) // 2
        y = self.winfo_rooty() + (self.winfo_height() - height) // 2
        window.geometry(f"{width}x{height}+{x}+{y}")

    def choose_metrics_workbook(self):
        current = get_metrics_workbook_path()
        selected = filedialog.askopenfilename(
            title="Choose metrics Excel file",
            initialdir=str(current.parent if current.parent.exists() else Path.home()),
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not selected:
            return

        settings = load_settings()
        settings["metrics_workbook_path"] = selected
        save_settings(settings)

        if isinstance(self.current_page, MetricsPage):
            self.show_metrics()

    def choose_monthly_recommendations(self):
        current = get_monthly_recommendations_path()
        selected = filedialog.askopenfilename(
            title="Choose monthly recommendations text file",
            initialdir=str(current.parent if current.parent.exists() else Path.home()),
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if not selected:
            return

        settings = load_settings()
        settings["monthly_recommendations_path"] = selected
        save_settings(settings)

        if isinstance(self.current_page, MetricsPage):
            self.show_metrics()

    def export_metrics_template(self):
        default_path = get_metrics_workbook_path()
        output_path = filedialog.asksaveasfilename(
            title="Export metrics Excel template",
            defaultextension=".xlsx",
            initialdir=str(default_path.parent if default_path.parent.exists() else Path.home()),
            initialfile=default_path.name,
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not output_path:
            return

        try:
            export_metrics_template(Path(output_path))
            messagebox.showinfo("Template exported", f"Metrics template exported:\n\n{output_path}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))

    def clear_content(self):
        for child in self.content.winfo_children():
            child.destroy()
        self.current_page = None

    def stop_hack_windows(self):
        self.hack_windows_running = False
        if self.hack_window_job is not None:
            try:
                self.after_cancel(self.hack_window_job)
            except Exception:
                pass
            self.hack_window_job = None
        for window in list(self.hack_windows):
            try:
                window.destroy()
            except Exception:
                pass
        self.hack_windows = []

    def start_hack_windows(self):
        if self.hack_windows_running:
            return
        self.hack_windows_running = True
        self.spawn_hack_window()

    def schedule_next_hack_window(self):
        if not self.hack_windows_running:
            return
        delay_ms = random.randint(15000, 30000)
        self.hack_window_job = self.after(delay_ms, self.spawn_hack_window)

    def spawn_hack_window(self):
        if not self.hack_windows_running:
            return
        self.hack_windows = [window for window in self.hack_windows if window.winfo_exists()]
        if len(self.hack_windows) < 3:
            window = tk.Toplevel(self)
            window.title(self.hack_window_title())
            self.set_hack_window_icon(window)
            width = random.randint(620, 980)
            height = random.randint(360, 620)
            screen_w = max(900, self.winfo_screenwidth())
            screen_h = max(700, self.winfo_screenheight())
            x = random.randint(20, max(20, screen_w - width - 40))
            y = random.randint(20, max(20, screen_h - height - 80))
            window.geometry(f"{width}x{height}+{x}+{y}")
            window.configure(bg="#050505")
            text = tk.Text(window, bg="#050505", fg=random.choice(["#00ff66", "#00d9ff", "#f5ff00", "#ff4fd8"]), insertbackground="#ffffff", font=("Consolas", random.choice([9, 10, 11])), wrap="none", borderwidth=0, highlightthickness=0)
            text.pack(fill="both", expand=True, padx=8, pady=8)
            text.tag_configure("flash", foreground="#050505", background=random.choice(["#00ff66", "#00d9ff", "#f5ff00", "#ff4fd8", "#ffffff"]))
            text.tag_configure("ascii", foreground=random.choice(["#ffffff", "#00d9ff", "#f5ff00"]), background="#050505")
            text.configure(state="normal")
            text.insert(tk.END, self.hack_ascii_block() + "\n", "ascii")
            text.insert(tk.END, self.hack_flash_word() + "\n", "flash")
            text.configure(state="disabled")
            window.protocol("WM_DELETE_WINDOW", window.destroy)
            self.hack_windows.append(window)
            life_ms = random.randint(15000, 30000)
            line_job = self.after(80, lambda w=window, t=text: self.write_hack_line(w, t))
            window.after(life_ms, lambda w=window: self.close_hack_window(w))
            window.hack_line_job = line_job
        self.schedule_next_hack_window()

    def set_hack_window_icon(self, window):
        if Image is None or ImageTk is None or ImageDraw is None:
            return
        try:
            word = random.choice(MELOVERSE_WORDS)
            bg = random.choice(["#00ff66", "#00d9ff", "#f5ff00", "#ff4fd8", "#ffffff", "#ff8c00"])
            image = Image.new("RGB", (32, 32), bg)
            draw = ImageDraw.Draw(image)
            initials = word[:2].upper()
            draw.rectangle((2, 2, 29, 29), outline="#050505", width=2)
            draw.text((8, 9), initials, fill="#050505")
            icon = ImageTk.PhotoImage(image)
            window.hack_icon_ref = icon
            window.iconphoto(True, icon)
        except Exception:
            pass

    def hack_window_title(self):
        paths = find_tile_backgrounds()
        suffix = random.choice(["kernel flow", "sync engine", "route trace", "process watch", "creative matrix", "data pulse", "window bus", "MeloVerse pulse"])
        if paths:
            name = random.choice(paths).stem
            return f"{name} :: {random.choice(MELOVERSE_WORDS)} :: {suffix}"
        return f"{random.choice(MELOVERSE_WORDS)}-{uuid.uuid4().hex[:6]} :: {suffix}"

    def close_hack_window(self, window):
        try:
            if hasattr(window, "hack_line_job"):
                self.after_cancel(window.hack_line_job)
        except Exception:
            pass
        try:
            window.destroy()
        except Exception:
            pass
        self.hack_windows = [item for item in self.hack_windows if item.winfo_exists()]

    def write_hack_line(self, window, text):
        if not self.hack_windows_running or not window.winfo_exists():
            return
        try:
            text.configure(state="normal")
            if random.random() < 0.22:
                text.tag_configure("ascii", foreground=random.choice(["#ffffff", "#00d9ff", "#f5ff00", "#ff4fd8"]), background="#050505")
                text.insert(tk.END, self.hack_ascii_block() + "\n", "ascii")
            if random.random() < 0.24:
                text.tag_configure("flash", foreground="#050505", background=random.choice(["#00ff66", "#00d9ff", "#f5ff00", "#ff4fd8", "#ffffff"]))
                text.insert(tk.END, self.hack_flash_word() + "\n", "flash")
            for _index in range(random.randint(1, 4)):
                text.insert(tk.END, self.hack_status_line() + "\n")
            text.see(tk.END)
            line_count = int(float(text.index("end-1c")))
            if line_count > 220:
                text.delete("1.0", "80.0")
            text.configure(state="disabled")
            window.hack_line_job = self.after(random.randint(35, 140), lambda w=window, t=text: self.write_hack_line(w, t))
        except Exception:
            pass

    def hack_flash_word(self):
        word = random.choice(MELOVERSE_WORDS).upper()
        width = random.randint(42, 72)
        pulse = random.choice(["SIGNAL", "TRACE", "FLOW", "MEMORY", "ROUTE", "ENGINE", "MOTION"])
        return f"{'=' * width}\n>>> {word:^28} :: {pulse} :: {uuid.uuid4().hex[:8].upper()} <<<\n{'=' * width}"

    def hack_ascii_block(self):
        blocks = [
            r"""
  [##########]  BUS LOCK
  [##====####]  STREAM TRACE
  [##########]  ACTIVE
""",
            r"""
      .-.
   .-(   )-.   PACKET CLOUD
  (___.__)__)  ROUTE MAPPED
""",
            r"""
  +----------------+
  | LUKESTROM BUS  |
  |  SIGNAL ALIVE  |
  +----------------+
""",
        ]
        return random.choice(blocks).strip("\n")

    def hack_status_line(self):
        try:
            cpu = psutil.cpu_percent(interval=None) if psutil is not None else random.randint(4, 92)
        except Exception:
            cpu = random.randint(4, 92)
        try:
            memory = psutil.virtual_memory().percent if psutil is not None else random.randint(15, 88)
        except Exception:
            memory = random.randint(15, 88)
        try:
            disk = psutil.disk_usage(str(Path.home().anchor or "C:\\")).percent if psutil is not None else random.randint(2, 75)
        except Exception:
            disk = random.randint(2, 75)
        process_name = "system"
        process_pid = random.randint(100, 9999)
        try:
            if psutil is not None:
                processes = [proc.info for proc in psutil.process_iter(["pid", "name"]) if proc.info.get("name")]
                if processes:
                    picked = random.choice(processes)
                    process_name = picked.get("name") or process_name
                    process_pid = picked.get("pid") or process_pid
        except Exception:
            pass
        inbound = uuid.uuid4().hex[:24].upper()
        outbound = uuid.uuid4().hex[:24].upper()
        templates = [
            "SYS  cpu={cpu:>5}  mem={memory:>5}  disk={disk:>5}  gpu={gpu:>5}",
            "NET  uplink={up}  downlink={down}  packets={packets:05d}  route=local",
            "SCAN {path}  checksum={checksum}  state={state}",
            "SYNC queue={queue:02d}  onedrive={onedrive}  cache={cache}",
            "PORT {port:05d}  handshake={handshake}  latency={latency}ms",
            "MTRX seed={seed}  frame={frame:04d}  mode=creative",
            "PROC pid={pid:<6} name={process:<26} cpu={cpu:>5} mem={memory:>5}",
            "IN   {inbound} -> bus:{port:05d} bytes={inbytes}",
            "OUT  bus:{port:05d} -> {outbound} bytes={outbytes}",
            "FLOW {process:<24} read={readrate}/s write={writerate}/s",
        ]
        def pct(value):
            return "n/a" if value is None else f"{value:.0f}%"
        recent = onedrive_recent_sync_paths(limit=1, seconds=900)
        path = ellipsize_middle(str(recent[0]) if recent else str(TILE_BACKGROUND_DIR), 54)
        return random.choice(templates).format(
            cpu=pct(cpu),
            memory=pct(memory),
            disk=pct(disk),
            gpu=random.choice(["n/a", f"{random.randint(0, 88)}%"]),
            up=format_bytes(random.randint(0, 4_000_000)) + "/s",
            down=format_bytes(random.randint(0, 4_000_000)) + "/s",
            packets=random.randint(0, 99999),
            path=path,
            checksum=uuid.uuid4().hex[:12].upper(),
            state=random.choice(["OK", "WATCH", "TRACE", "MAPPED", "LIVE"]),
            queue=random.randint(0, 24),
            onedrive="ON" if is_onedrive_running() else "OFF",
            cache=random.choice(["0 B", "mapped", "idle", "watch"]),
            port=random.randint(1024, 65535),
            handshake=random.choice(["ACK", "SYN", "READY", "LOCK", "FLOW"]),
            latency=random.randint(2, 180),
            seed=uuid.uuid4().hex[:8],
            frame=random.randint(0, 9999),
            pid=process_pid,
            process=ellipsize_middle(process_name, 26),
            inbound=inbound,
            outbound=outbound,
            inbytes=format_bytes(random.randint(128, 8_000_000)),
            outbytes=format_bytes(random.randint(128, 8_000_000)),
            readrate=format_bytes(random.randint(0, 80_000_000)),
            writerate=format_bytes(random.randint(0, 80_000_000)),
        )

    def set_page_title(self, title):
        self.title(APP_TITLE if title == APP_TITLE else f"{APP_TITLE} - {title}")

    def open_all_links(self):
        for _text, url in HOME_LINKS:
            webbrowser.open(url)

    def open_in_chrome(self, url):
        chrome = chrome_exe()
        if chrome is not None:
            try:
                subprocess.Popen([str(chrome), url], shell=False)
                return
            except Exception as exc:
                messagebox.showwarning("Chrome could not open", f"Opening in default browser instead.\n\n{exc}")
        webbrowser.open(url)

    def kill_outlook(self):
        subprocess.run(["taskkill", "/IM", "outlook.exe", "/F"], capture_output=True, text=True)

    def show_release_notes(self):
        try:
            file_content = RELEASE_NOTES_FILE.read_text(encoding="utf-8").strip()
        except FileNotFoundError:
            file_content = ""
        content = APP_RELEASE_NOTES.strip()
        if file_content:
            content = f"{content}\n\n# Tool-specific release notes\n\n{file_content}"

        window = tk.Toplevel(self)
        window.title("Release notes")
        window.geometry("940x560")

        shell = ttk.Frame(window, padding=10)
        shell.pack(fill="both", expand=True)
        shell.columnconfigure(1, weight=1)
        shell.rowconfigure(0, weight=1)

        toc = tk.Listbox(shell, width=28, font=("Segoe UI", 9), activestyle="none")
        toc.grid(row=0, column=0, sticky="ns", padx=(0, 10))

        notes = tk.Text(shell, wrap="word", font=("Segoe UI", 10), padx=12, pady=12)
        notes.grid(row=0, column=1, sticky="nsew")
        self.style_text_widget(notes)
        notes.insert("1.0", content)
        notes.configure(state="disabled")

        headings = release_note_headings(content)
        if headings:
            for heading, _line_number in headings:
                toc.insert(tk.END, heading)
        else:
            toc.insert(tk.END, "Release notes")

        def jump_to_heading(_event=None):
            selection = toc.curselection()
            if not selection or not headings:
                return
            _heading, line_number = headings[selection[0]]
            notes.configure(state="normal")
            notes.see(f"{line_number}.0")
            notes.tag_remove("toc_highlight", "1.0", tk.END)
            notes.tag_add("toc_highlight", f"{line_number}.0", f"{line_number}.end")
            notes.tag_configure("toc_highlight", background="#fff2b8")
            notes.configure(state="disabled")

        toc.bind("<<ListboxSelect>>", jump_to_heading)

    def show_how_to(self):
        content = APP_HOW_TO.strip()
        window = tk.Toplevel(self)
        window.title("How to")
        window.geometry("940x560")

        shell = ttk.Frame(window, padding=10)
        shell.pack(fill="both", expand=True)
        shell.columnconfigure(1, weight=1)
        shell.rowconfigure(0, weight=1)

        toc = tk.Listbox(shell, width=28, font=("Segoe UI", 9), activestyle="none")
        toc.grid(row=0, column=0, sticky="ns", padx=(0, 10))

        text = tk.Text(shell, wrap="word", font=("Segoe UI", 10), padx=12, pady=12)
        text.grid(row=0, column=1, sticky="nsew")
        self.style_text_widget(text)
        text.insert("1.0", content)
        text.configure(state="disabled")

        headings = release_note_headings(content)
        if headings:
            for heading, _line_number in headings:
                toc.insert(tk.END, heading)
        else:
            toc.insert(tk.END, "How to")

        def jump_to_heading(_event=None):
            selection = toc.curselection()
            if not selection or not headings:
                return
            _heading, line_number = headings[selection[0]]
            text.configure(state="normal")
            text.see(f"{line_number}.0")
            text.tag_remove("toc_highlight", "1.0", tk.END)
            text.tag_add("toc_highlight", f"{line_number}.0", f"{line_number}.end")
            text.tag_configure("toc_highlight", background="#fff2b8")
            text.configure(state="disabled")

        toc.bind("<<ListboxSelect>>", jump_to_heading)

    def show_home(self):
        self.stop_hack_windows()
        self.clear_content()
        self.set_page_title(APP_TITLE)
        self.current_page = HomePage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_post_planner(self):
        self.clear_content()
        self.set_page_title("Campaign Planner")
        self.current_page = CampaignPlannerPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_reel_design(self):
        self.clear_content()
        self.set_page_title("Reel Design")
        self.current_page = ReelDesignPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_youtube_downloader(self):
        self.clear_content()
        self.set_page_title("YouTube Downloader")
        self.current_page = YouTubeDownloaderPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_audio_vu(self):
        self.clear_content()
        self.set_page_title("Tools")
        self.current_page = VuPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_song_analyzer(self):
        self.clear_content()
        self.set_page_title("Song Analyzer")
        self.current_page = SongAnalyzerPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_metrics(self):
        self.clear_content()
        self.set_page_title("Metrics")
        self.current_page = MetricsPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_tools(self):
        self.clear_content()
        self.set_page_title("Shortcuts")
        self.current_page = ToolsPage(self.content, self)
        self.current_page.pack(fill="both", expand=True)

    def show_tools_vu_window(self):
        PerformanceVuWindow(self)

    def delete_cacheclip(self):
        if isinstance(self.current_page, ToolsPage):
            self.current_page.confirm_delete_cache()
            return
        if not CACHECLIP_DIR.exists():
            messagebox.showinfo("Cache not found", f"This cache folder does not exist:\n\n{CACHECLIP_DIR}")
            return
        confirmed = messagebox.askyesno(
            "Delete cache",
            f"Delete this entire cache folder?\n\n{CACHECLIP_DIR}\n\nThis cannot be undone.",
        )
        if not confirmed:
            return
        errors = []
        make_tree_writable(CACHECLIP_DIR)
        try:
            def handle_remove_error(func, path, _exc_info):
                os.chmod(path, 0o700)
                func(path)
            shutil.rmtree(CACHECLIP_DIR, onerror=handle_remove_error)
        except Exception as exc:
            errors.append(str(exc))

        if CACHECLIP_DIR.exists():
            result = powershell_remove_tree(CACHECLIP_DIR)
            if result.returncode != 0:
                details = (result.stderr or result.stdout or "").strip()
                if details:
                    errors.append(details)

        if CACHECLIP_DIR.exists():
            file_count, total_size = folder_stats(CACHECLIP_DIR)
            messagebox.showwarning(
                "Cache partly deleted",
                "Some cache files could not be deleted. They may be locked by OneDrive or another app.\n\n"
                f"Remaining: {file_count:,} files | {format_bytes(total_size)}\n\n"
                "Tip: stop OneDrive in Tools, close video/editing apps, then try Delete cache again.\n\n"
                + ("\n\n".join(errors[-2:]) if errors else ""),
            )
        else:
            messagebox.showinfo("Cache deleted", "Media cache folder deleted.")
        if isinstance(self.current_page, ToolsPage):
            self.current_page.refresh()

    def start_onedrive(self, status_var=None, on_change=None):
        exe = one_drive_exe()
        if exe is None:
            tried = "\n".join(str(path) for path in one_drive_candidates())
            messagebox.showerror("OneDrive not found", f"I could not find OneDrive. Checked locations:\n\n{tried}")
            return
        subprocess.Popen([str(exe)], shell=False)
        if status_var is not None:
            status_var.set("OneDrive starting.")
        if on_change is not None:
            self.after(1200, on_change)

    def stop_onedrive(self, status_var=None, on_change=None):
        exe = one_drive_exe()
        if exe is None:
            tried = "\n".join(str(path) for path in one_drive_candidates())
            messagebox.showerror("OneDrive not found", f"I could not find OneDrive. Checked locations:\n\n{tried}")
            return
        subprocess.Popen([str(exe), "/shutdown"], shell=False)
        if status_var is not None:
            status_var.set("OneDrive stopping.")
        if on_change is not None:
            self.after(1200, on_change)

    def run_with_loading(self, action):
        loading = tk.Toplevel(self)
        loading.title("")
        loading.resizable(False, False)
        loading.transient(self)
        loading.configure(bg=self.colors["panel_bg"])
        loading.overrideredirect(True)

        frame = ttk.Frame(loading, padding=(28, 18), style="Panel.TFrame")
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Loading...", font=("Segoe UI", 12, "bold")).pack()

        self.update_idletasks()
        width = 210
        height = 76
        x = self.winfo_rootx() + (self.winfo_width() - width) // 2
        y = self.winfo_rooty() + (self.winfo_height() - height) // 2
        loading.geometry(f"{width}x{height}+{x}+{y}")
        loading.lift()
        loading.update()

        def run_action():
            try:
                action()
            finally:
                loading.destroy()

        self.after(60, run_action)


def folder_stats(path):
    total_size = 0
    file_count = 0
    if not path.exists():
        return 0, 0
    for root, _dirs, files in os.walk(path):
        for filename in files:
            file_path = Path(root) / filename
            try:
                total_size += file_path.stat().st_size
                file_count += 1
            except Exception:
                pass
    return file_count, total_size


def make_tree_writable(path):
    if not path.exists():
        return
    for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            try:
                os.chmod(Path(root) / name, 0o700)
            except Exception:
                pass
        for name in dirs:
            try:
                os.chmod(Path(root) / name, 0o700)
            except Exception:
                pass
    try:
        os.chmod(path, 0o700)
    except Exception:
        pass


def powershell_remove_tree(path):
    command = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-Command",
        "Remove-Item -LiteralPath $args[0] -Recurse -Force -ErrorAction Stop",
        str(path),
    ]
    return subprocess.run(command, capture_output=True, text=True, timeout=120)


def powershell_float(script, timeout=5):
    try:
        completed = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            timeout=timeout,
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )
        if completed.returncode != 0:
            return None
        output = (completed.stdout or "").strip().splitlines()
        if not output:
            return None
        text = output[-1].strip().replace(",", ".")
        return float(text)
    except Exception:
        return None


def powershell_text(script, timeout=5):
    try:
        completed = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            timeout=timeout,
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )
        if completed.returncode != 0:
            return ""
        return (completed.stdout or "").strip()
    except Exception:
        return ""


def clamp_percent(value):
    if value is None:
        return None
    return max(0.0, min(100.0, float(value)))


class MemoryStatusEx(ctypes.Structure):
    _fields_ = [
        ("dwLength", ctypes.c_ulong),
        ("dwMemoryLoad", ctypes.c_ulong),
        ("ullTotalPhys", ctypes.c_ulonglong),
        ("ullAvailPhys", ctypes.c_ulonglong),
        ("ullTotalPageFile", ctypes.c_ulonglong),
        ("ullAvailPageFile", ctypes.c_ulonglong),
        ("ullTotalVirtual", ctypes.c_ulonglong),
        ("ullAvailVirtual", ctypes.c_ulonglong),
        ("ullAvailExtendedVirtual", ctypes.c_ulonglong),
    ]


def get_memory_usage_percent():
    if psutil is not None:
        try:
            return clamp_percent(psutil.virtual_memory().percent)
        except Exception:
            pass
    if os.name == "nt":
        status = MemoryStatusEx()
        status.dwLength = ctypes.sizeof(MemoryStatusEx)
        if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(status)):
            return clamp_percent(status.dwMemoryLoad)
    return None


def scaled_rate_percent(current_value, state_key, scale_per_second):
    now = time_module.monotonic()
    time_key = f"{state_key}_time"
    previous_time = PERFORMANCE_IO_STATE.get(time_key)
    previous_value = PERFORMANCE_IO_STATE.get(state_key)
    PERFORMANCE_IO_STATE[state_key] = current_value
    PERFORMANCE_IO_STATE[time_key] = now
    if previous_time is None or previous_value is None:
        return 0.0
    elapsed = max(0.2, now - previous_time)
    rate = max(0.0, (current_value - previous_value) / elapsed)
    return clamp_percent((rate / scale_per_second) * 100)


def get_disk_usage_percent():
    if psutil is None:
        value = powershell_float("(Get-CimInstance Win32_PerfFormattedData_PerfDisk_PhysicalDisk -Filter \"Name='_Total'\").PercentDiskTime", timeout=2)
        if value is None:
            value = powershell_float("(Get-Counter '\\PhysicalDisk(_Total)\\% Disk Time').CounterSamples.CookedValue", timeout=2)
        return clamp_percent(value)
    try:
        counters = psutil.disk_io_counters()
        if counters is None:
            return None
        total = float(counters.read_bytes + counters.write_bytes)
        return scaled_rate_percent(total, "disk", 80 * 1024 * 1024)
    except Exception:
        return None


def get_network_usage_percent():
    if psutil is None:
        value = powershell_float("$n=(Get-CimInstance Win32_PerfFormattedData_Tcpip_NetworkInterface | Measure-Object BytesTotalPersec -Sum).Sum; if ($n -ne $null) {$n}", timeout=2)
        if value is None:
            return None
        return clamp_percent((value / (12 * 1024 * 1024)) * 100)
    try:
        counters = psutil.net_io_counters()
        if counters is None:
            return None
        total = float(counters.bytes_sent + counters.bytes_recv)
        return scaled_rate_percent(total, "net", 12 * 1024 * 1024)
    except Exception:
        return None


def filetime_to_int(filetime):
    return (int(filetime.dwHighDateTime) << 32) + int(filetime.dwLowDateTime)


def get_cpu_usage_percent_from_system_times():
    if os.name != "nt":
        return None
    idle = wintypes.FILETIME()
    kernel = wintypes.FILETIME()
    user = wintypes.FILETIME()
    if not ctypes.windll.kernel32.GetSystemTimes(ctypes.byref(idle), ctypes.byref(kernel), ctypes.byref(user)):
        return None
    idle_value = filetime_to_int(idle)
    kernel_value = filetime_to_int(kernel)
    user_value = filetime_to_int(user)
    previous_idle = CPU_TIME_STATE.get("idle")
    previous_kernel = CPU_TIME_STATE.get("kernel")
    previous_user = CPU_TIME_STATE.get("user")
    CPU_TIME_STATE["idle"] = idle_value
    CPU_TIME_STATE["kernel"] = kernel_value
    CPU_TIME_STATE["user"] = user_value
    if previous_idle is None or previous_kernel is None or previous_user is None:
        return 0.0
    idle_delta = idle_value - previous_idle
    kernel_delta = kernel_value - previous_kernel
    user_delta = user_value - previous_user
    total_delta = kernel_delta + user_delta
    if total_delta <= 0:
        return None
    busy_delta = total_delta - idle_delta
    return clamp_percent((busy_delta / total_delta) * 100)


class CoreTempSharedData(ctypes.Structure):
    _fields_ = [
        ("uiLoad", ctypes.c_uint * 256),
        ("uiTjMax", ctypes.c_uint * 128),
        ("uiCoreCnt", ctypes.c_uint),
        ("uiCPUCnt", ctypes.c_uint),
        ("fTemp", ctypes.c_float * 256),
        ("fVID", ctypes.c_float),
        ("fCPUSpeed", ctypes.c_float),
        ("fFSBSpeed", ctypes.c_float),
        ("fMultiplier", ctypes.c_float),
        ("sCPUName", ctypes.c_char * 100),
        ("ucFahrenheit", ctypes.c_ubyte),
        ("ucDeltaToTjMax", ctypes.c_ubyte),
    ]


def get_coretemp_snapshot():
    if os.name != "nt":
        return {}
    kernel32 = ctypes.windll.kernel32
    file_map_read = 0x0004
    handle = kernel32.OpenFileMappingW(file_map_read, False, "CoreTempMappingObject")
    if not handle:
        return {}
    kernel32.MapViewOfFile.restype = ctypes.c_void_p
    view = kernel32.MapViewOfFile(handle, file_map_read, 0, 0, ctypes.sizeof(CoreTempSharedData))
    if not view:
        kernel32.CloseHandle(handle)
        return {}
    try:
        data = CoreTempSharedData.from_address(view)
        core_count = int(data.uiCoreCnt or 0)
        cpu_count = int(data.uiCPUCnt or 1)
        sample_count = max(1, min(256, core_count * cpu_count if core_count else cpu_count))
        loads = [float(data.uiLoad[index]) for index in range(sample_count) if 0 <= data.uiLoad[index] <= 100]
        temps = [float(data.fTemp[index]) for index in range(sample_count) if -20 <= data.fTemp[index] <= 130]
        snapshot = {}
        if loads:
            snapshot["cpu"] = clamp_percent(sum(loads) / len(loads))
        if temps:
            snapshot["temp"] = max(temps)
        return snapshot
    except Exception:
        return {}
    finally:
        kernel32.UnmapViewOfFile(view)
        kernel32.CloseHandle(handle)


def get_cpu_usage_percent():
    if psutil is not None:
        try:
            value = psutil.cpu_percent(interval=None)
            if value == 0:
                value = psutil.cpu_percent(interval=0.05)
            return clamp_percent(value)
        except Exception:
            pass
    value = powershell_float("(Get-CimInstance Win32_PerfFormattedData_PerfOS_Processor -Filter \"Name='_Total'\").PercentProcessorTime", timeout=2)
    if value is None:
        value = powershell_float("(Get-Counter '\\Processor(_Total)\\% Processor Time').CounterSamples.CookedValue", timeout=2)
    if value is None:
        value = get_cpu_usage_percent_from_system_times()
    return clamp_percent(value)


def get_gpu_usage_percent():
    now = time_module.monotonic()
    if PERFORMANCE_IO_STATE.get("gpu") is not None and now - PERFORMANCE_IO_STATE.get("gpu_time", 0) < 2:
        return PERFORMANCE_IO_STATE.get("gpu")
    scripts = [
        (
            "$samples=(Get-Counter '\\GPU Engine(*)\\Utilization Percentage' -ErrorAction SilentlyContinue).CounterSamples | "
            "Where-Object {$_.InstanceName -match 'engtype_(3D|Copy|Compute|VideoDecode|VideoEncode)'}; "
            "if ($samples) {($samples | Measure-Object CookedValue -Maximum).Maximum}"
        ),
        (
            "$samples=(Get-Counter '\\GPU Engine(*)\\Utilization Percentage' -ErrorAction SilentlyContinue).CounterSamples; "
            "if ($samples) {($samples | Measure-Object CookedValue -Maximum).Maximum}"
        ),
        (
            "$samples=Get-CimInstance Win32_PerfFormattedData_GPUPerformanceCounters_GPUEngine -ErrorAction SilentlyContinue; "
            "if ($samples) {($samples | Measure-Object UtilizationPercentage -Maximum).Maximum}"
        ),
    ]
    value = None
    for script in scripts:
        value = powershell_float(script, timeout=3)
        if value is not None:
            value = clamp_percent(value)
            break
    PERFORMANCE_IO_STATE["gpu_time"] = now
    PERFORMANCE_IO_STATE["gpu"] = value
    return value


def get_core_temperature_celsius():
    scripts = [
        "$s=Get-CimInstance -Namespace root/OpenHardwareMonitor -Class Sensor -ErrorAction SilentlyContinue | Where-Object {$_.SensorType -eq 'Temperature' -and ($_.Name -like '*CPU*' -or $_.Name -like '*Core*' -or $_.Name -like '*Package*')} | Sort-Object Value -Descending | Select-Object -First 1; if ($s) {$s.Value}",
    ]
    for script in scripts:
        value = powershell_float(script)
        if value is not None and -20 <= value <= 130:
            return float(value)
    return None


def get_ram_usage_percent():
    return get_memory_usage_percent()


def collect_performance_snapshot():
    snapshot = {}
    for key, getter in (
        ("cpu", get_cpu_usage_percent),
        ("memory", get_memory_usage_percent),
        ("disk", get_disk_usage_percent),
        ("network", get_network_usage_percent),
        ("gpu", get_gpu_usage_percent),
    ):
        try:
            snapshot[key] = getter()
        except Exception:
            snapshot[key] = None
    try:
        coretemp = get_coretemp_snapshot()
    except Exception:
        coretemp = {}
    try:
        snapshot["temp"] = coretemp.get("temp") if coretemp.get("temp") is not None else get_core_temperature_celsius()
    except Exception:
        snapshot["temp"] = None
    try:
        onedrive_snapshot = one_drive_activity_snapshot()
        snapshot["onedrive_upload"] = onedrive_rate_to_activity_percent(onedrive_snapshot.get("upload_rate", 0))
        snapshot["onedrive_download"] = onedrive_rate_to_activity_percent(onedrive_snapshot.get("download_rate", 0))
        snapshot["onedrive_upload_label"] = onedrive_snapshot.get("upload_label", "No active file detected")
        snapshot["onedrive_download_label"] = onedrive_snapshot.get("download_label", "No active file detected")
    except Exception:
        snapshot["onedrive_upload"] = 0
        snapshot["onedrive_download"] = 0
        snapshot["onedrive_upload_label"] = "No active file detected"
        snapshot["onedrive_download_label"] = "No active file detected"
    return snapshot


PERFORMANCE_METER_OPTIONS = {
    "cpu": {"title": "Processor", "unit": "%", "maximum": 100, "color": "#2f6f73"},
    "memory": {"title": "Memory", "unit": "%", "maximum": 100, "color": "#7b5f2a"},
    "disk": {"title": "Disk", "unit": "%", "maximum": 100, "color": "#4f6f8f"},
    "network": {"title": "Network", "unit": "%", "maximum": 100, "color": "#5b6770"},
    "gpu": {"title": "GPU", "unit": "%", "maximum": 100, "color": "#4f6f8f"},
    "temp": {"title": "Core temp", "unit": "C", "maximum": 100, "color": "#8d3f3f"},
    "onedrive_upload": {"title": "Upload activity", "unit": "%", "maximum": 100, "color": "#8d3f3f"},
    "onedrive_download": {"title": "Download activity", "unit": "%", "maximum": 100, "color": "#2f6f73"},
}


def format_bytes(size):
    value = float(size)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if value < 1024 or unit == "TB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{int(size)} B"


def explorer_window_handles():
    handles = []
    try:
        user32 = ctypes.windll.user32

        def callback(hwnd, _lparam):
            if not user32.IsWindowVisible(hwnd):
                return True
            class_name = ctypes.create_unicode_buffer(256)
            user32.GetClassNameW(hwnd, class_name, 256)
            if class_name.value in ("CabinetWClass", "ExploreWClass"):
                handles.append(hwnd)
            return True

        enum_proc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)(callback)
        user32.EnumWindows(enum_proc, 0)
    except Exception:
        return []
    return handles


def move_explorer_windows_side_by_side(before_handles):
    try:
        time_module.sleep(0.8)
        after_handles = explorer_window_handles()
        new_handles = [hwnd for hwnd in after_handles if hwnd not in before_handles]
        handles = new_handles[-2:] if len(new_handles) >= 2 else after_handles[-2:]
        if len(handles) < 2:
            return
        user32 = ctypes.windll.user32
        screen_w = user32.GetSystemMetrics(0)
        screen_h = user32.GetSystemMetrics(1)
        half_w = max(640, screen_w // 2)
        height = max(520, screen_h - 80)
        positions = [(0, 0, half_w, height), (half_w, 0, max(640, screen_w - half_w), height)]
        for hwnd, (x, y, width, win_height) in zip(handles, positions):
            user32.ShowWindow(hwnd, 9)
            user32.MoveWindow(hwnd, x, y, width, win_height, True)
    except Exception:
        pass


def open_dual_file_explorer(folder_path):
    path = Path(folder_path)
    if not path.exists():
        messagebox.showerror("Folder not found", f"This folder does not exist:\n\n{path}")
        return
    before_handles = explorer_window_handles()
    for _side in ("left", "right"):
        try:
            subprocess.Popen(["explorer.exe", str(path)], shell=False)
        except Exception as exc:
            messagebox.showerror("Could not open File Explorer", str(exc))
            return
    threading.Thread(target=move_explorer_windows_side_by_side, args=(before_handles,), daemon=True).start()


class SimpleTooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.window = None
        widget.bind("<Enter>", self.show, add="+")
        widget.bind("<Leave>", self.hide, add="+")

    def show(self, _event=None):
        if self.window is not None:
            return
        x = self.widget.winfo_rootx() + 10
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 8
        self.window = tk.Toplevel(self.widget)
        self.window.wm_overrideredirect(True)
        self.window.wm_geometry(f"+{x}+{y}")
        label = ttk.Label(self.window, text=self.text, padding=(10, 6), style="CardText.TLabel")
        label.pack()

    def hide(self, _event=None):
        if self.window is not None:
            self.window.destroy()
            self.window = None


class ToolsPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.cache_status_var = tk.StringVar()
        self.cache_progress_var = tk.StringVar()
        self.system_tiles = {}
        self.meter_background_paths = find_tile_backgrounds()
        self.meter_jobs = []
        self.performance_canvases = {}
        self.performance_display_values = {}
        self.meter_availability = {key: True for key in PERFORMANCE_METER_OPTIONS}
        self.performance_running = True
        self.performance_collecting = False
        self.meter_vars = {}
        self.meter_menu = None
        self.meter_menu_indices = {}
        self.performance_container = None
        self.system_grid = None
        self.web_grid = None
        self.onedrive_activity_job = None
        self._build()

    def _build(self):
        outer = ScrollFrame(self)
        outer.pack(fill="both", expand=True, pady=(0, 14))
        outer.canvas.configure(bg=self.app.colors["panel_bg"], highlightthickness=0)
        outer.body.configure(style="Panel.TFrame")
        outer.set_auto_scrollbar(True)

        main = ttk.Frame(outer.body, style="Panel.TFrame", padding=18)
        main.pack(fill="both", expand=True)
        title_row = ttk.Frame(main, style="Panel.TFrame")
        title_row.pack(fill="x", pady=(0, 14))
        ttk.Label(title_row, text="Shortcuts", style="CardTitle.TLabel").pack(side="left")

        system_title = ttk.Frame(main, style="Panel.TFrame")
        system_title.pack(fill="x", pady=(0, 12))
        ttk.Label(system_title, text="System actions", style="CardTitle.TLabel").pack(side="left")
        system_grid = tk.Canvas(main, height=260, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.system_grid = system_grid
        system_grid.pack(fill="x", pady=(0, 18))
        for col in range(3):
            system_grid.columnconfigure(col, weight=1, uniform="systemtiles")
        for row in range(2):
            system_grid.rowconfigure(row, weight=1, uniform="systemrows")
        system_grid.bind("<Configure>", lambda _event: self.draw_pane_background(system_grid))
        self.create_action_tile(system_grid, 0, 0, "OneDrive", "#5b6770", "Checking...", "Refresh", self.refresh, "onedrive", "cloud", "microsoft.com")
        self.create_action_tile(system_grid, 0, 1, "Outlook", "#8d3f3f", "Close outlook.exe", "Click to kill", self.app.kill_outlook, "outlook", "mail", "outlook.live.com")
        self.create_action_tile(system_grid, 0, 2, "Media cache", "#7b5f2a", "Checking...", "Click to delete", self.confirm_delete_cache, "cache", "trash", "microsoft.com")
        self.create_action_tile(system_grid, 1, 0, "YT Downloader", "#d93025", "Download videos", "Open tool", self.app.show_youtube_downloader, "youtube", "play", "youtube.com")
        self.create_action_tile(system_grid, 1, 1, "LukeStrom MeloVerse Explorer", "#6f42c1", "Open creations", "Open two explorers", self.open_universe_browser, "universe", "tree", "lukestrom.com")
        self.layout_system_actions()
        self.cache_progress = ttk.Progressbar(main, mode="determinate", maximum=100)
        self.cache_progress_label = ttk.Label(main, textvariable=self.cache_progress_var, style="CardText.TLabel")

        ttk.Label(main, text="Web shortcuts", style="CardTitle.TLabel").pack(anchor="w", pady=(22, 12))

        grid = tk.Canvas(main, height=420, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.web_grid = grid
        grid.pack(fill="both", expand=True)
        for col in range(3):
            grid.columnconfigure(col, weight=1, uniform="webtiles")
        for row in range(4):
            grid.rowconfigure(row, weight=1, uniform="webrows")
        grid.bind("<Configure>", lambda _event: self.draw_pane_background(grid))
        shortcuts = [
            {"title": "Open all", "icon_domain": "google.com", "icon": "links", "color": "#202124", "command": self.app.open_all_links, "action": "Open core links"},
            {"title": "ChatGPT", "icon_domain": "chatgpt.com", "icon": "note", "color": "#10a37f", "command": self.open_chatgpt_links, "action": "Open projects"},
            *WEB_SHORTCUTS,
        ]
        for index, shortcut in enumerate(shortcuts):
            self.create_web_tile(grid, index // 3, index % 3, shortcut)
        self.refresh()
        self.refresh_onedrive_activity()

    def destroy(self):
        self.performance_running = False
        for job in self.meter_jobs:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        if self.onedrive_activity_job is not None:
            try:
                self.after_cancel(self.onedrive_activity_job)
            except Exception:
                pass
        super().destroy()

    def layout_performance_meters(self):
        if self.performance_container is None:
            return
        for canvas in self.performance_canvases.values():
            canvas.grid_forget()
        for canvas in (getattr(self, "audio_left_canvas", None), getattr(self, "audio_right_canvas", None)):
            if canvas is not None:
                canvas.grid_forget()
        for column in range(3):
            self.performance_container.columnconfigure(column, weight=1, uniform="meters", minsize=0)
        for row in range(6):
            self.performance_container.rowconfigure(row, weight=0, minsize=0)
        visible = [key for key in PERFORMANCE_METER_OPTIONS.keys() if self.meter_vars.get(key, tk.BooleanVar(value=True)).get()]
        row_count = (len(visible) + 2) // 3
        for row in range(row_count):
            self.performance_container.rowconfigure(row, weight=1 if row < row_count else 0)
        for index, key in enumerate(visible):
            canvas = self.performance_canvases.get(key)
            if canvas is None:
                continue
            row = index // 3
            col = index % 3
            canvas.grid(row=row, column=col, sticky="nsew", padx=(0 if col == 0 else 8, 0), pady=(0 if row == 0 else 8, 0))
        self.performance_container.update_idletasks()

    def draw_pane_background(self, canvas):
        canvas.delete("pane_background")
        width = max(1, canvas.winfo_width())
        height = max(1, canvas.winfo_height())
        rounded_rect(canvas, 0, 0, width - 1, height - 1, radius=10, fill=self.app.colors["panel_bg"], outline=self.app.colors["border"], width=1, tags=("pane_background",))
        canvas.tag_lower("pane_background")

    def open_universe_browser(self):
        open_dual_file_explorer(UNIVERSE_DIR)

    def layout_system_actions(self):
        if self.system_grid is None:
            return
        self.system_grid.configure(height=260)
        positions = {
            "onedrive": (0, 0),
            "outlook": (0, 1),
            "cache": (0, 2),
            "youtube": (1, 0),
            "universe": (1, 1),
        }
        for key, position in positions.items():
            canvas = self.system_tiles.get(key)
            if canvas is not None and hasattr(canvas, "tile_card"):
                row, column = position
                canvas.tile_card.grid(row=row, column=column, sticky="nsew", padx=6, pady=6)

    def rotate_meter_background(self, canvas):
        if not canvas.winfo_exists():
            return
        if self.meter_background_paths:
            canvas.meter_background_path = random.choice(self.meter_background_paths)
        else:
            canvas.meter_background_path = None
        canvas.meter_face_signature = None
        self.draw_vu_meter(canvas, self.performance_display_values.get(getattr(canvas, "slot_key", "")))
        interval_ms = get_tile_picture_interval_seconds() * 1000
        job = self.after(interval_ms, lambda: self.rotate_meter_background(canvas))
        self.meter_jobs.append(job)

    def draw_vu_meter(self, canvas, value):
        canvas.update_idletasks()
        width = max(180, canvas.winfo_width())
        height = max(getattr(canvas, "meter_min_height", 220), canvas.winfo_height())
        title = getattr(canvas, "meter_title", "")
        subtext = getattr(canvas, "meter_subtext", "")
        unit = getattr(canvas, "meter_unit", "")
        maximum = getattr(canvas, "meter_maximum", 100)
        display = "n/a" if value is None else f"{value:.0f}{unit}"
        ratio = 0 if value is None else max(0, min(1, float(value) / maximum))

        if self.app.dark_mode:
            outer_bg = "#101010"
            face = "#202020"
            inner_face = "#262626"
            face_shadow = "#171717"
            border = "#3a3a3a"
            ink = "#f3f0e8"
            muted_ink = "#a8a095"
            needle = "#f3f0e8" if value is not None else "#777777"
        else:
            outer_bg = "#f2f0ea"
            face = "#ffffff"
            inner_face = "#ffffff"
            face_shadow = "#e8e3d8"
            border = "#b8b0a2"
            ink = "#1c1710"
            muted_ink = "#756a5c"
            needle = "#2a2118" if value is not None else "#8a8173"
        red = "#d23a2e"

        if value is None:
            canvas.delete("all")
            background_path = getattr(canvas, "meter_background_path", None)
            rounded_rect(canvas, 0, 0, width, height, radius=10, fill=inner_face, outline=border, width=1, tags=("meter_face",))
            if background_path and Image is not None and ImageTk is not None:
                try:
                    bg_image = image_for_meter_full(background_path, width, height)
                    canvas.meter_background_ref = bg_image
                    canvas.create_image(0, 0, image=bg_image, anchor="nw", tags=("meter_face",))
                    rounded_rect(canvas, 0, 0, width - 1, height - 1, radius=10, outline=border, width=1, fill="", tags=("meter_face",))
                except Exception:
                    canvas.meter_background_ref = None
            return

        center_x = width / 2
        center_y = height * 1.16
        radius = min(width * 0.50, height * 1.06)
        start_angle = 132
        end_angle = 48
        sweep = start_angle - end_angle

        def point_for(percent, extra_radius=0):
            angle = math.radians(start_angle - sweep * percent)
            r = radius + extra_radius
            return center_x + math.cos(angle) * r, center_y - math.sin(angle) * r

        background_path = getattr(canvas, "meter_background_path", None)
        face_signature = (width, height, self.app.dark_mode, title, subtext, str(background_path) if background_path else "")
        if getattr(canvas, "meter_face_signature", None) != face_signature:
            canvas.delete("all")
            canvas.meter_face_signature = face_signature
            rounded_rect(canvas, 0, 0, width, height, radius=10, fill=outer_bg, outline="", tags=("meter_face",))
            rounded_rect(canvas, 6, 6, width - 6, height - 6, radius=10, fill=face_shadow, outline=border, width=1, tags=("meter_face", "meter_frame"))
            rounded_rect(canvas, 14, 14, width - 14, height - 14, radius=10, fill=face, outline=border, tags=("meter_face", "meter_frame"))
            rounded_rect(canvas, 20, 20, width - 20, height - 20, radius=10, fill=inner_face, outline="", tags=("meter_face",))
            if background_path and Image is not None and ImageTk is not None:
                try:
                    bg_image = image_for_meter(background_path, width - 40, height - 40, inner_face)
                    canvas.meter_background_ref = bg_image
                    canvas.create_image(20, 20, image=bg_image, anchor="nw", tags=("meter_face",))
                except Exception:
                    canvas.meter_background_ref = None
            canvas.create_text(26, 24, text=title, fill=ink, anchor="nw", font=("Segoe UI", 11), tags=("meter_face",))
            if subtext:
                canvas.create_text(26, height - 30, text=ellipsize_middle(subtext, 78), fill=muted_ink, anchor="sw", font=("Segoe UI", 8), tags=("meter_face",))

            for idx in range(0, 41):
                percent = idx / 40
                is_major = idx % 10 == 0
                is_mid = idx % 5 == 0
                tick_len = 24 if is_major else 18 if is_mid else 10
                outer = point_for(percent)
                inner = point_for(percent, -tick_len)
                tick_color = red if percent >= 0.72 else ink
                canvas.create_line(inner[0], inner[1], outer[0], outer[1], fill=tick_color, width=2 if is_major else 1, tags=("meter_face",))

            arc_box = (
                center_x - radius,
                center_y - radius,
                center_x + radius,
                center_y + radius,
            )
            canvas.create_arc(arc_box, start=end_angle, extent=sweep * 0.72, style="arc", outline=muted_ink, width=1, tags=("meter_face",))
            canvas.create_arc(arc_box, start=end_angle, extent=sweep * 0.28, style="arc", outline=red, width=5, tags=("meter_face",))

        canvas.delete("meter_dynamic")
        canvas.create_text(width - 26, 24, text=display, fill=ink, anchor="ne", font=("Segoe UI", 17), tags=("meter_dynamic",))
        if value is not None:
            now = time_module.monotonic()
            peak_time = getattr(canvas, "meter_peak_time", 0)
            peak_ratio = getattr(canvas, "meter_peak_ratio", 0)
            if now - peak_time >= 60 or ratio > peak_ratio:
                canvas.meter_peak_ratio = ratio
                canvas.meter_peak_time = now
        peak_ratio = getattr(canvas, "meter_peak_ratio", 0)
        if peak_ratio > 0:
            peak_outer = point_for(peak_ratio, 12)
            peak_inner = point_for(peak_ratio, -18)
            canvas.create_line(peak_inner[0], peak_inner[1], peak_outer[0], peak_outer[1], fill=red, width=3, tags=("meter_dynamic", "meter_peak"))

        jitter = 0 if value is None else random.uniform(-0.012, 0.012)
        needle_percent = max(0, min(1, ratio + jitter))
        needle_tip = point_for(needle_percent, -10)
        pivot_x, pivot_y = center_x, height + 18
        canvas.create_line(pivot_x, pivot_y, needle_tip[0], needle_tip[1], fill=needle, width=2, tags=("meter_dynamic", "meter_needle"))

        canvas.create_rectangle(0, height - 24, width, height, fill=outer_bg, outline="", tags=("meter_dynamic", "meter_frame_overlay"))
        canvas.create_rectangle(6, height - 24, width - 6, height - 6, fill=face_shadow, outline="", tags=("meter_dynamic", "meter_frame_overlay"))
        canvas.create_rectangle(14, height - 24, width - 14, height - 14, fill=face, outline="", tags=("meter_dynamic", "meter_frame_overlay"))
        canvas.create_rectangle(20, height - 24, width - 20, height - 20, fill=inner_face, outline="", tags=("meter_dynamic", "meter_frame_overlay"))
        rounded_rect(canvas, 6, 6, width - 6, height - 6, radius=10, outline=border, width=1, fill="", tags=("meter_dynamic", "meter_frame_overlay"))
        rounded_rect(canvas, 14, 14, width - 14, height - 14, radius=10, outline=border, width=1, fill="", tags=("meter_dynamic", "meter_frame_overlay"))

    def refresh_performance(self):
        if not self.winfo_exists() or not self.performance_running:
            return
        if self.performance_collecting:
            self.after(500, self.refresh_performance)
            return
        self.performance_collecting = True

        def worker():
            try:
                snapshot = collect_performance_snapshot()
            except Exception:
                snapshot = {}
            def finish():
                self.performance_collecting = False
                if self.winfo_exists() and self.performance_running:
                    self.update_performance(snapshot)
            if self.winfo_exists() and self.performance_running:
                self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()
        self.after(500, self.refresh_performance)

    def update_performance(self, snapshot):
        for slot_key, canvas in self.performance_canvases.items():
            if slot_key in ("onedrive_upload", "onedrive_download"):
                canvas.meter_subtext = snapshot.get(f"{slot_key}_label", "No active file detected")
            new_value = snapshot.get(slot_key)
            previous = self.performance_display_values.get(slot_key)
            if new_value is None:
                display_value = previous
            elif previous is None:
                display_value = new_value
            else:
                display_value = previous + (new_value - previous) * 0.68
            self.performance_display_values[slot_key] = display_value
            self.draw_vu_meter(canvas, display_value)

    def create_action_tile(self, parent, row, column, title, color, status, action_text, command, key, icon, icon_domain=None):
        card = tk.Frame(parent, bg=self.app.colors["panel_bg"], highlightthickness=0, cursor="hand2")
        card.grid(row=row, column=column, sticky="nsew", padx=6, pady=6)
        canvas = tk.Canvas(card, height=118, highlightthickness=0, bg=self.app.colors["panel_bg"], cursor="hand2")
        canvas.pack(fill="both", expand=True)
        canvas.tile_card = card
        canvas.tile_title = title
        canvas.tile_color = color
        canvas.tile_status = status
        canvas.tile_action = action_text
        canvas.tile_command = command
        canvas.tile_icon = icon
        canvas.tile_icon_url = favicon_url_for_domain(icon_domain, 64) if icon_domain else ""
        self.system_tiles[key] = canvas

        def run_action(_event=None):
            action = getattr(canvas, "tile_command", None)
            if action is not None:
                action()

        canvas.bind("<Configure>", lambda _event=None, c=canvas: self.draw_action_tile(c))
        canvas.bind("<Button-1>", run_action)
        card.bind("<Button-1>", run_action)
        self.draw_action_tile(canvas)

    def refresh_onedrive_activity(self):
        if not self.winfo_exists() or not self.performance_running:
            return
        self.update_onedrive_widgets()
        self.onedrive_activity_job = self.after(2500, self.refresh_onedrive_activity)

    def update_onedrive_widgets(self):
        snapshot = one_drive_activity_snapshot()
        status = snapshot.get("status", "unknown")
        if status == "Stopped":
            self.update_action_tile(
                "onedrive",
                status="Stopped",
                action_text="Click to start",
                command=lambda: self.app.start_onedrive(on_change=self.refresh),
            )
            return

        if status == "starting":
            activity = "Activity starting"
        elif status == "unknown":
            activity = "Activity unknown"
        else:
            activity = f"D {format_bytes(snapshot.get('download_rate', 0))}/s | U {format_bytes(snapshot.get('upload_rate', 0))}/s"
        self.update_action_tile(
            "onedrive",
            status="Running",
            action_text=activity,
            command=lambda: self.app.stop_onedrive(on_change=self.refresh),
        )

    def draw_action_tile(self, canvas):
        canvas.delete("all")
        width = max(220, canvas.winfo_width())
        height = max(118, canvas.winfo_height())
        color = getattr(canvas, "tile_color", "#5b6770")
        title = getattr(canvas, "tile_title", "")
        status = getattr(canvas, "tile_status", "")
        action = getattr(canvas, "tile_action", "")
        icon = getattr(canvas, "tile_icon", "tool")
        icon_url = getattr(canvas, "tile_icon_url", "")
        rounded_rect(canvas, 0, 0, width, height, radius=10, fill=self.app.colors["panel_bg"], outline="")
        rounded_rect(canvas, 0, 0, width, 34, radius=10, fill=color, outline="")
        canvas.create_rectangle(0, 22, width, 34, fill=color, outline="")
        rounded_rect(canvas, 12, 48, 92, height - 16, radius=10, fill=color, outline="")
        if not self.draw_tile_image_icon(canvas, icon_url, 52, (height + 32) / 2, 56):
            self.draw_tile_icon(canvas, icon, 52, (height + 32) / 2, 56, color)
        canvas.create_text(108, 50, text=title, fill=self.app.colors["text"], anchor="nw", font=("Segoe UI", 12, "bold"))
        canvas.create_text(108, 73, text=status, fill=self.app.colors["text"], anchor="nw", font=("Segoe UI", 8), width=max(120, width - 120))
        canvas.create_text(108, height - 12, text=action, fill=self.app.colors["muted"], anchor="sw", font=("Segoe UI", 8), width=max(120, width - 120))

    def update_action_tile(self, key, status=None, action_text=None, command=None):
        canvas = self.system_tiles.get(key)
        if canvas is None:
            return
        if status is not None:
            canvas.tile_status = status
        if action_text is not None:
            canvas.tile_action = action_text
        if command is not None:
            canvas.tile_command = command
        self.draw_action_tile(canvas)

    def confirm_delete_cache(self):
        if not CACHECLIP_DIR.exists():
            messagebox.showinfo("Cache not found", f"This cache folder does not exist:\n\n{CACHECLIP_DIR}")
            self.refresh()
            return
        confirmed = messagebox.askyesno(
            "Delete cache",
            f"Delete this entire cache folder?\n\n{CACHECLIP_DIR}\n\nThis cannot be undone.",
        )
        if not confirmed:
            return
        self.start_cache_delete()

    def start_cache_delete(self):
        self.cache_progress_var.set("Preparing media cache delete...")
        self.cache_progress.configure(value=0, maximum=100)
        self.cache_progress.pack(fill="x", pady=(0, 6))
        self.cache_progress_label.pack(anchor="w", pady=(0, 14))
        self.update_action_tile("cache", status="Deleting...", action_text="Please wait", command=lambda: None)

        def worker():
            errors = []
            total_files = 0
            deleted_files = 0
            try:
                total_files, _total_size = folder_stats(CACHECLIP_DIR)
            except Exception:
                total_files = 0

            def report_progress(force=False):
                if total_files <= 0:
                    percent = 0
                else:
                    percent = min(100, int(deleted_files * 100 / total_files))
                if force or deleted_files % 25 == 0:
                    self.after(0, lambda d=deleted_files, t=total_files, p=percent: self.update_cache_delete_progress(d, t, p))

            if CACHECLIP_DIR.exists():
                make_tree_writable(CACHECLIP_DIR)
                try:
                    for root, dirs, files in os.walk(CACHECLIP_DIR, topdown=False):
                        for filename in files:
                            file_path = Path(root) / filename
                            try:
                                os.chmod(file_path, 0o700)
                            except Exception:
                                pass
                            try:
                                file_path.unlink()
                                deleted_files += 1
                                report_progress()
                            except Exception as exc:
                                errors.append(f"{file_path}: {exc}")
                        for dirname in dirs:
                            dir_path = Path(root) / dirname
                            try:
                                dir_path.rmdir()
                            except Exception:
                                pass
                    try:
                        CACHECLIP_DIR.rmdir()
                    except Exception:
                        pass
                except Exception as exc:
                    errors.append(str(exc))

            if CACHECLIP_DIR.exists():
                result = powershell_remove_tree(CACHECLIP_DIR)
                if result.returncode != 0:
                    details = (result.stderr or result.stdout or "").strip()
                    if details:
                        errors.append(details)

            if CACHECLIP_DIR.exists():
                remaining = folder_stats(CACHECLIP_DIR)
                success = False
            else:
                remaining = (0, 0)
                success = True
            report_progress(force=True)
            self.after(0, lambda: self.finish_cache_delete(success, remaining, errors))

        threading.Thread(target=worker, daemon=True).start()

    def update_cache_delete_progress(self, deleted_files, total_files, percent):
        if total_files <= 0:
            self.cache_progress.configure(value=0, maximum=100)
            self.cache_progress_var.set("Deleting media cache...")
        else:
            self.cache_progress.configure(value=percent, maximum=100)
            self.cache_progress_var.set(f"Deleting media cache: {deleted_files:,} / {total_files:,} files")

    def finish_cache_delete(self, success, remaining, errors):
        self.cache_progress.pack_forget()
        self.cache_progress_label.pack_forget()
        if success:
            self.cache_progress_var.set("")
            messagebox.showinfo("Cache deleted", "Media cache folder deleted.")
        else:
            file_count, total_size = remaining
            messagebox.showwarning(
                "Cache partly deleted",
                "Some cache files could not be deleted. They may be locked by OneDrive or another app.\n\n"
                f"Remaining: {file_count:,} files | {format_bytes(total_size)}\n\n"
                "Tip: stop OneDrive in Tools, close video/editing apps, then try Delete cache again.\n\n"
                + ("\n\n".join(errors[-2:]) if errors else ""),
            )
        self.refresh()

    def draw_tile_image_icon(self, canvas, icon_url, cx, cy, size):
        radius = size / 2
        canvas.create_oval(cx - radius, cy - radius, cx + radius, cy + radius, fill="#ffffff", outline="", width=0)
        image = load_tile_icon_image(icon_url, int(size * 0.68))
        if image is None:
            return False
        if not hasattr(canvas, "tile_icon_refs"):
            canvas.tile_icon_refs = []
        canvas.tile_icon_refs = [image]
        canvas.create_image(cx, cy, image=image, anchor="center")
        return True

    def draw_tile_icon(self, canvas, icon, cx, cy, size, color):
        half = size / 2
        canvas.create_oval(cx - half, cy - half, cx + half, cy + half, fill="#ffffff", outline="", width=0)
        size = size * 0.70
        half = size / 2
        if icon == "cloud":
            canvas.create_oval(cx - 24, cy - 2, cx + 2, cy + 24, outline=color, width=3)
            canvas.create_oval(cx - 8, cy - 18, cx + 22, cy + 16, outline=color, width=3)
            canvas.create_oval(cx - 28, cy - 14, cx - 2, cy + 14, outline=color, width=3)
            canvas.create_line(cx - 28, cy + 14, cx + 24, cy + 14, fill=color, width=3)
        elif icon == "mail":
            canvas.create_rectangle(cx - 24, cy - 16, cx + 24, cy + 16, outline=color, width=3)
            canvas.create_line(cx - 24, cy - 16, cx, cy + 4, cx + 24, cy - 16, fill=color, width=3)
            canvas.create_line(cx - 24, cy + 16, cx - 4, cy, fill=color, width=2)
            canvas.create_line(cx + 24, cy + 16, cx + 4, cy, fill=color, width=2)
        elif icon == "trash":
            canvas.create_rectangle(cx - 18, cy - 10, cx + 18, cy + 24, outline=color, width=3)
            canvas.create_line(cx - 24, cy - 16, cx + 24, cy - 16, fill=color, width=3)
            canvas.create_line(cx - 8, cy - 22, cx + 8, cy - 22, fill=color, width=3)
            for offset in (-8, 0, 8):
                canvas.create_line(cx + offset, cy - 4, cx + offset, cy + 18, fill=color, width=2)
        elif icon == "play":
            canvas.create_polygon(cx - 16, cy - 22, cx - 16, cy + 22, cx + 24, cy, fill=color, outline=color)
        elif icon == "refresh":
            canvas.create_arc(cx - 24, cy - 24, cx + 24, cy + 24, start=35, extent=285, outline=color, width=3, style="arc")
            canvas.create_polygon(cx + 18, cy - 18, cx + 28, cy - 6, cx + 12, cy - 5, fill=color, outline=color)
        elif icon == "music":
            canvas.create_oval(cx - 22, cy + 10, cx - 4, cy + 28, fill=color, outline=color)
            canvas.create_line(cx - 4, cy + 18, cx - 4, cy - 24, fill=color, width=4)
            canvas.create_line(cx - 4, cy - 24, cx + 22, cy - 18, fill=color, width=4)
        elif icon == "chart":
            for index, bar_h in enumerate((20, 34, 48)):
                x = cx - 24 + index * 18
                canvas.create_rectangle(x, cy + 24 - bar_h, x + 10, cy + 24, fill=color, outline=color)
        elif icon == "infinity":
            canvas.create_text(cx, cy, text="âˆž", fill=color, font=("Segoe UI Symbol", 34, "bold"))
        elif icon == "note":
            canvas.create_oval(cx - 18, cy + 8, cx, cy + 26, fill=color, outline=color)
            canvas.create_oval(cx + 8, cy + 2, cx + 26, cy + 20, fill=color, outline=color)
            canvas.create_line(cx, cy + 16, cx, cy - 24, fill=color, width=4)
            canvas.create_line(cx + 26, cy + 10, cx + 26, cy - 18, fill=color, width=4)
            canvas.create_line(cx, cy - 24, cx + 26, cy - 18, fill=color, width=4)
        elif icon == "tree":
            canvas.create_line(cx, cy - 24, cx, cy + 24, fill=color, width=4)
            for ox, oy in ((-18, -8), (18, -8), (-18, 16), (18, 16)):
                canvas.create_oval(cx + ox - 7, cy + oy - 7, cx + ox + 7, cy + oy + 7, fill=color, outline=color)
                canvas.create_line(cx, cy + oy, cx + ox, cy + oy, fill=color, width=3)
        elif icon == "monitor":
            canvas.create_rectangle(cx - 26, cy - 20, cx + 26, cy + 14, outline=color, width=3)
            canvas.create_line(cx - 10, cy + 24, cx + 10, cy + 24, fill=color, width=3)
            canvas.create_line(cx, cy + 14, cx, cy + 24, fill=color, width=3)
        elif icon == "compass":
            canvas.create_oval(cx - 24, cy - 24, cx + 24, cy + 24, outline=color, width=3)
            canvas.create_polygon(cx + 4, cy - 20, cx - 8, cy + 6, cx + 8, cy + 2, fill=color, outline=color)
        elif icon == "links":
            canvas.create_oval(cx - 26, cy - 16, cx - 2, cy + 8, outline=color, width=3)
            canvas.create_oval(cx + 2, cy - 8, cx + 26, cy + 16, outline=color, width=3)
            canvas.create_line(cx - 8, cy + 4, cx + 8, cy - 4, fill=color, width=3)
        else:
            canvas.create_oval(cx - half / 2, cy - half / 2, cx + half / 2, cy + half / 2, outline=color, width=3)

    def create_web_tile(self, parent, row, column, shortcut):
        card = tk.Frame(parent, bg=self.app.colors["panel_bg"], highlightthickness=0, cursor="hand2")
        card.grid(row=row, column=column, sticky="nsew", padx=6, pady=6)
        canvas = tk.Canvas(card, height=118, highlightthickness=0, bg=self.app.colors["panel_bg"], cursor="hand2")
        canvas.pack(fill="both", expand=True)

        def draw(_event=None):
            canvas.delete("all")
            width = max(220, canvas.winfo_width())
            height = max(118, canvas.winfo_height())
            color = shortcut["color"]
            icon_url = shortcut.get("icon_url") or favicon_url_for_domain(shortcut.get("icon_domain") or urllib.parse.urlparse(shortcut.get("url", "")).netloc, 64)
            rounded_rect(canvas, 0, 0, width, height, radius=10, fill=self.app.colors["panel_bg"], outline="")
            rounded_rect(canvas, 0, 0, width, 34, radius=10, fill=color, outline="")
            canvas.create_rectangle(0, 22, width, 34, fill=color, outline="")
            rounded_rect(canvas, 12, 48, 92, height - 16, radius=10, fill=color, outline="")
            if not self.draw_tile_image_icon(canvas, icon_url, 52, (height + 32) / 2, 56):
                self.draw_tile_icon(canvas, shortcut.get("icon", "tool"), 52, (height + 32) / 2, 56, color)
            canvas.create_text(108, 58, text=shortcut["title"], fill=self.app.colors["text"], anchor="nw", font=("Segoe UI", 12, "bold"))
            canvas.create_text(108, 84, text=shortcut.get("action", "Open in Chrome"), fill=self.app.colors["muted"], anchor="nw", font=("Segoe UI", 9))

        def open_shortcut(_event=None):
            command = shortcut.get("command")
            if command is not None:
                command()
            else:
                self.app.open_in_chrome(shortcut["url"])

        canvas.bind("<Configure>", draw)
        canvas.bind("<Button-1>", open_shortcut)
        card.bind("<Button-1>", open_shortcut)

    def open_chatgpt_links(self):
        self.app.clear_content()
        self.app.set_page_title("ChatGPT")
        self.app.current_page = ChatGPTLinksPage(self.app.content, self.app, self.app.show_tools)
        self.app.current_page.pack(fill="both", expand=True)

    def refresh(self):
        self.update_onedrive_widgets()

        if CACHECLIP_DIR.exists():
            file_count, total_size = folder_stats(CACHECLIP_DIR)
            self.cache_status_var.set(f"{file_count:,} files | {format_bytes(total_size)}")
        else:
            self.cache_status_var.set("Folder not found")
        self.update_action_tile("cache", status=self.cache_status_var.get(), action_text="Click to delete")
        self.update_action_tile("outlook", status="Close outlook.exe", action_text="Click to kill")


class VuPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.meter_background_paths = find_tile_backgrounds()
        self.meter_jobs = []
        self.performance_canvases = {}
        self.performance_display_values = {}
        self.meter_availability = {key: True for key in PERFORMANCE_METER_OPTIONS}
        self.performance_running = True
        self.performance_collecting = False
        self.meter_vars = {}
        self.meter_menu = None
        self.meter_menu_indices = {}
        self.audio_left_var = tk.BooleanVar(value=True)
        self.audio_right_var = tk.BooleanVar(value=True)
        self.performance_container = None
        self.empty_meter_label = None
        self.audio_page = None
        self._build()
        self.refresh_performance()

    def _build(self):
        main = ttk.Frame(self, style="Panel.TFrame", padding=18)
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(1, weight=1)
        header = ttk.Frame(main, style="Panel.TFrame")
        header.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        ttk.Label(header, text="Tools", style="CardTitle.TLabel").pack(side="left")
        meter_button = ttk.Menubutton(header, text="VU meters")
        meter_button.pack(side="left", padx=(10, 0))
        meter_menu = tk.Menu(meter_button, tearoff=False)
        self.meter_menu = meter_menu
        meter_button.configure(menu=meter_menu)
        for key, config in PERFORMANCE_METER_OPTIONS.items():
            var = tk.BooleanVar(value=True)
            self.meter_vars[key] = var
            meter_menu.add_checkbutton(label=config["title"], variable=var, command=self.layout_meters)
            self.meter_menu_indices[key] = meter_menu.index("end")
        meter_menu.add_separator()
        meter_menu.add_checkbutton(label="Audio L", variable=self.audio_left_var, command=self.layout_meters)
        meter_menu.add_checkbutton(label="Audio R", variable=self.audio_right_var, command=self.layout_meters)
        actions = ttk.Frame(header, style="Panel.TFrame")
        actions.pack(side="left", fill="x", expand=True, padx=(14, 0))
        self.create_header_action(actions, "OneDrive", self.toggle_onedrive, "Start or stop OneDrive syncing")
        self.create_header_action(actions, "Outlook", self.app.kill_outlook, "Close outlook.exe")
        self.create_header_action(actions, "Cache", self.app.delete_cacheclip, "Delete the media cache folder")
        self.create_header_action(actions, "YT", self.app.show_youtube_downloader, "Open YouTube downloader")
        self.create_header_action(actions, "MeloVerse", self.open_meloverse_explorer, "Open the LukeStrom MeloVerse Explorer")

        performance = ttk.Frame(main, style="Panel.TFrame")
        self.performance_container = performance
        performance.grid(row=1, column=0, sticky="nsew")
        for key, config in PERFORMANCE_METER_OPTIONS.items():
            canvas = tk.Canvas(performance, height=220, bg=self.app.colors["panel_bg"], highlightthickness=1, highlightbackground=self.app.colors["border"])
            canvas.slot_key = key
            canvas.meter_title = config["title"]
            canvas.meter_unit = config["unit"]
            canvas.meter_maximum = config["maximum"]
            canvas.meter_color = config["color"]
            canvas.meter_peak_ratio = 0
            canvas.meter_peak_time = 0
            canvas.meter_subtext = ""
            self.performance_canvases[key] = canvas
            canvas.bind("<Configure>", lambda _event=None, c=canvas: self.draw_vu_meter(c, self.performance_display_values.get(c.slot_key)))
            self.draw_vu_meter(canvas, None)
            self.rotate_meter_background(canvas)

        self.audio_left_canvas = tk.Canvas(performance, height=220, bg=self.app.colors["panel_bg"], highlightthickness=0, bd=0)
        self.audio_right_canvas = tk.Canvas(performance, height=220, bg=self.app.colors["panel_bg"], highlightthickness=0, bd=0)
        self.audio_left_canvas.slot_key = "audio_left"
        self.audio_right_canvas.slot_key = "audio_right"
        self.audio_page = AudioVuPage(
            self,
            self.app,
            embedded_canvases=(self.audio_left_canvas, self.audio_right_canvas),
            controls_parent=None,
            show_source_controls=False,
        )
        self.layout_meters()

    def create_header_action(self, parent, text, command, tooltip):
        button = ttk.Button(parent, text=text, command=command, style="NavLinks.TButton")
        button.pack(side="left", padx=(0, 6))
        SimpleTooltip(button, tooltip)
        return button

    def toggle_onedrive(self):
        if is_onedrive_running():
            self.app.stop_onedrive(on_change=self.refresh_performance)
        else:
            self.app.start_onedrive(on_change=self.refresh_performance)

    def open_meloverse_explorer(self):
        open_dual_file_explorer(UNIVERSE_DIR)

    def destroy(self):
        self.performance_running = False
        for job in self.meter_jobs:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        if self.audio_page is not None and self.audio_page.winfo_exists():
            try:
                self.audio_page.destroy()
            except Exception:
                pass
        super().destroy()

    def layout_meters(self):
        if self.performance_container is None:
            return
        if self.empty_meter_label is not None and self.empty_meter_label.winfo_exists():
            self.empty_meter_label.destroy()
            self.empty_meter_label = None
        for canvas in self.performance_canvases.values():
            canvas.grid_forget()
        for canvas in (self.audio_left_canvas, self.audio_right_canvas):
            canvas.grid_forget()
        for column in range(6):
            self.performance_container.columnconfigure(column, weight=0, uniform="", minsize=0)
        for row in range(6):
            self.performance_container.rowconfigure(row, weight=0, uniform="", minsize=0)
        visible = [
            key for key in PERFORMANCE_METER_OPTIONS
            if self.meter_availability.get(key, True) and self.meter_vars.get(key, tk.BooleanVar(value=True)).get()
        ]
        if self.audio_left_var.get():
            visible.append("audio_left")
        if self.audio_right_var.get():
            visible.append("audio_right")
        if not visible:
            self.empty_meter_label = ttk.Label(self.performance_container, text="No meters selected.", style="Muted.TLabel")
            self.empty_meter_label.grid(row=0, column=0, sticky="nsew")
            self.performance_container.columnconfigure(0, weight=1)
            self.performance_container.rowconfigure(0, weight=1)
            return
        pairs = (("onedrive_upload", "onedrive_download"), ("audio_left", "audio_right"))
        paired_keys = {key for pair in pairs if pair[0] in visible and pair[1] in visible for key in pair}
        blocks = []
        consumed = set()
        for key in visible:
            if key in consumed:
                continue
            pair = next((pair for pair in pairs if pair[0] == key and pair[1] in visible), None)
            if pair is not None:
                blocks.append(pair)
                consumed.update(pair)
            else:
                blocks.append((key,))
                consumed.add(key)
        column_count = 1 if len(visible) == 1 else 2 if paired_keys else 3
        placements = []
        row = 0
        col = 0
        for block in blocks:
            if len(block) == 2:
                if col != 0:
                    row += 1
                    col = 0
                placements.append((block[0], row, 0, 1))
                placements.append((block[1], row, 1, 1))
                row += 1
                col = 0
                continue
            key = block[0]
            span = column_count if column_count == 2 and col == 0 and row == len(blocks) - 1 else 1
            placements.append((key, row, col, span))
            col += span
            if col >= column_count:
                row += 1
                col = 0
        row_count = max(1, max((item[1] for item in placements), default=0) + 1)
        for column in range(column_count):
            self.performance_container.columnconfigure(column, weight=1, uniform="meters", minsize=0)
        for row in range(row_count):
            self.performance_container.rowconfigure(row, weight=1, uniform="meterrows", minsize=0)
        for key, row, col, span in placements:
            canvas = self.performance_canvases.get(key)
            if key == "audio_left":
                canvas = self.audio_left_canvas
            elif key == "audio_right":
                canvas = self.audio_right_canvas
            if canvas is None:
                continue
            canvas.grid(row=row, column=col, columnspan=span, sticky="nsew", padx=(0 if col == 0 else 8, 0), pady=(0 if row == 0 else 8, 0))

    def update_meter_menu_state(self, key, available):
        self.meter_availability[key] = available
        if self.meter_menu is None or key not in self.meter_menu_indices:
            return
        index = self.meter_menu_indices[key]
        if available:
            self.meter_menu.entryconfig(index, state="normal")
        else:
            self.meter_menu.entryconfig(index, state="disabled")

    def rotate_meter_background(self, canvas):
        if not self.performance_running or not canvas.winfo_exists():
            return
        canvas.meter_background_path = random.choice(self.meter_background_paths) if self.meter_background_paths else None
        canvas.meter_face_signature = None
        self.draw_vu_meter(canvas, self.performance_display_values.get(canvas.slot_key))
        job = self.after(get_tile_picture_interval_seconds() * 1000, lambda: self.rotate_meter_background(canvas))
        self.meter_jobs.append(job)

    def draw_vu_meter(self, canvas, value):
        ToolsPage.draw_vu_meter(self, canvas, value)

    def refresh_performance(self):
        if not self.performance_running or not self.winfo_exists():
            return
        if self.performance_collecting:
            self.after(500, self.refresh_performance)
            return
        self.performance_collecting = True

        def worker():
            try:
                snapshot = collect_performance_snapshot()
            except Exception:
                snapshot = {}

            def finish():
                self.performance_collecting = False
                if self.performance_running and self.winfo_exists():
                    self.update_performance(snapshot)

            if self.performance_running and self.winfo_exists():
                self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()
        self.after(500, self.refresh_performance)

    def update_performance(self, snapshot):
        for slot_key, canvas in self.performance_canvases.items():
            if slot_key in ("onedrive_upload", "onedrive_download"):
                canvas.meter_subtext = snapshot.get(f"{slot_key}_label", "No active file detected")
            new_value = snapshot.get(slot_key)
            self.update_meter_menu_state(slot_key, new_value is not None)
            previous = self.performance_display_values.get(slot_key)
            if previous is None or new_value is None:
                display_value = new_value
            else:
                display_value = previous + (new_value - previous) * 0.55
            self.performance_display_values[slot_key] = display_value
            self.draw_vu_meter(canvas, display_value)
        self.layout_meters()


class FolderBrowserPage(ttk.Frame):
    def __init__(self, parent, app, root_path, back_command):
        super().__init__(parent)
        self.app = app
        self.root_path = Path(root_path)
        self.current_path = self.root_path
        self.back_command = back_command
        self.view_mode = tk.StringVar(value="list")
        self.selected_path = None
        self.selected_widget = None
        self._build()
        self.show_folder(self.current_path)

    def _build(self):
        shell = ttk.Frame(self, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True)
        header = ttk.Frame(shell, style="Panel.TFrame")
        header.pack(fill="x", pady=(0, 12))
        ttk.Label(header, text="LukeStrom MeloVerse Explorer", style="CardTitle.TLabel").pack(side="left")
        ttk.Button(header, text="Open in File Explorer", command=self.open_current_in_explorer).pack(side="right", padx=(8, 0))
        ttk.Button(header, text="Back", command=self.back_command).pack(side="right")
        ttk.Button(header, text="Copy selected", command=self.copy_selected).pack(side="right", padx=(8, 0))
        ttk.Button(header, text="Big icons", command=lambda: self.set_view_mode("big")).pack(side="right", padx=(8, 0))
        ttk.Button(header, text="List", command=lambda: self.set_view_mode("list")).pack(side="right", padx=(8, 0))
        self.path_var = tk.StringVar()
        ttk.Label(shell, textvariable=self.path_var, style="Muted.TLabel").pack(anchor="w", pady=(0, 10))
        self.scroller = ScrollFrame(shell)
        self.scroller.pack(fill="both", expand=True)
        self.scroller.canvas.configure(bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.scroller.body.configure(style="Panel.TFrame")
        self.scroller.set_auto_scrollbar(True)
        self.list_frame = self.scroller.body

    def set_view_mode(self, mode):
        self.view_mode.set(mode)
        self.show_folder(self.current_path)

    def show_folder(self, path):
        self.current_path = Path(path)
        if not str(self.current_path).lower().startswith(str(self.root_path).lower()):
            self.current_path = self.root_path
        self.selected_path = None
        self.selected_widget = None
        self.path_var.set(str(self.current_path))
        for widget in self.list_frame.winfo_children():
            widget.destroy()
        if self.current_path != self.root_path:
            self.add_item("..", self.current_path.parent, True, 0)
        if not self.current_path.exists():
            ttk.Label(self.list_frame, text="Folder not found.", style="Muted.TLabel").pack(anchor="w", pady=8)
            return
        try:
            folders = sorted([item for item in self.current_path.iterdir() if item.is_dir()], key=lambda item: item.name.lower())
            files = sorted([item for item in self.current_path.iterdir() if item.is_file()], key=lambda item: item.name.lower())
        except Exception as exc:
            ttk.Label(self.list_frame, text=f"Could not read folder: {exc}", style="Muted.TLabel").pack(anchor="w", pady=8)
            return
        items = [(item, True) for item in folders] + [(item, False) for item in files]
        if not items and self.current_path == self.root_path:
            ttk.Label(self.list_frame, text="No folders or files here.", style="Muted.TLabel").pack(anchor="w", pady=8)
            return
        for index, (item, is_folder) in enumerate(items, start=1 if self.current_path != self.root_path else 0):
            self.add_item(item.name, item, is_folder, index)

    def add_item(self, title, path, is_folder, index):
        if self.view_mode.get() == "list":
            row = ttk.Frame(self.list_frame, style="Panel.TFrame", padding=(8, 4))
            row.pack(fill="x", pady=2)
            label = ttk.Label(row, text=("Folder  " if is_folder else "File      ") + title, style="NavText.TLabel", cursor="hand2")
            label.pack(side="left", fill="x", expand=True)
        else:
            row = ttk.Frame(self.list_frame, style="Panel.TFrame", padding=10)
            row.grid(row=index // 4, column=index % 4, sticky="nsew", padx=6, pady=6)
            self.list_frame.columnconfigure(index % 4, weight=1, uniform="universe")
            badge = tk.Canvas(row, width=86, height=64, bg=self.app.colors["panel_bg"], highlightthickness=0, cursor="hand2")
            badge.pack(anchor="center")
            color = "#6f42c1" if is_folder else "#2f6f73"
            rounded_rect(badge, 8, 6, 78, 58, radius=10, fill=color, outline="")
            badge.create_text(43, 32, text="DIR" if is_folder else "FILE", fill="#ffffff", font=("Segoe UI", 12, "bold"))
            label = ttk.Label(row, text=ellipsize_middle(title, 28), style="CardText.TLabel", cursor="hand2", anchor="center")
            label.pack(fill="x")
            badge.bind("<Button-1>", lambda event=None, widget=row, p=path: self.select_item(widget, p))
            badge.bind("<Double-Button-1>", lambda event=None, p=path, folder=is_folder: self.open_item(p, folder))
        for widget in (row, label):
            widget.bind("<Button-1>", lambda event=None, widget=row, p=path: self.select_item(widget, p))
            widget.bind("<Double-Button-1>", lambda event=None, p=path, folder=is_folder: self.open_item(p, folder))

    def select_item(self, widget, path):
        if self.selected_widget is not None and self.selected_widget.winfo_exists():
            self.selected_widget.configure(style="Panel.TFrame")
        self.selected_widget = widget
        self.selected_path = Path(path)
        try:
            widget.configure(style="Selected.TFrame")
        except Exception:
            pass

    def open_item(self, path, is_folder):
        path = Path(path)
        if is_folder:
            self.show_folder(path)
        else:
            try:
                os.startfile(str(path))
            except Exception as exc:
                messagebox.showerror("Could not open file", str(exc))

    def copy_selected(self):
        if self.selected_path is None:
            messagebox.showinfo("Nothing selected", "Select a file or folder first.")
            return
        destination = filedialog.askdirectory(title="Copy selected item to")
        if not destination:
            return
        destination = Path(destination) / self.selected_path.name
        try:
            if self.selected_path.is_dir():
                shutil.copytree(self.selected_path, destination, dirs_exist_ok=True)
            else:
                shutil.copy2(self.selected_path, destination)
        except Exception as exc:
            messagebox.showerror("Could not copy", str(exc))

    def open_current_in_explorer(self):
        try:
            os.startfile(str(self.current_path))
        except Exception as exc:
            messagebox.showerror("Could not open File Explorer", str(exc))


class ExplorerPane(ttk.Frame):
    def __init__(self, parent, app, root_path, title):
        super().__init__(parent, style="Panel.TFrame", padding=10)
        self.app = app
        self.root_path = Path(root_path)
        self.current_path = self.root_path
        self.selected_path = None
        self.selected_widget = None
        self.path_var = tk.StringVar(value=str(self.current_path))
        ttk.Label(self, text=title, style="CardTitle.TLabel").pack(anchor="w", pady=(0, 8))
        controls = ttk.Frame(self, style="Panel.TFrame")
        controls.pack(fill="x", pady=(0, 8))
        ttk.Button(controls, text="Back", command=self.go_up).pack(side="left")
        ttk.Button(controls, text="Copy selected", command=self.copy_selected).pack(side="left", padx=(8, 0))
        ttk.Button(controls, text="Open in File Explorer", command=self.open_current_in_explorer).pack(side="right")
        ttk.Label(self, textvariable=self.path_var, style="Muted.TLabel").pack(anchor="w", pady=(0, 8))
        self.scroller = ScrollFrame(self)
        self.scroller.pack(fill="both", expand=True)
        self.scroller.canvas.configure(bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.scroller.body.configure(style="Panel.TFrame")
        self.scroller.set_auto_scrollbar(True)
        self.list_frame = self.scroller.body
        self.show_folder(self.current_path)

    def go_up(self):
        if self.current_path != self.root_path:
            self.show_folder(self.current_path.parent)

    def show_folder(self, path):
        path = Path(path)
        if not str(path).lower().startswith(str(self.root_path).lower()):
            path = self.root_path
        self.current_path = path
        self.selected_path = None
        self.selected_widget = None
        self.path_var.set(str(path))
        for widget in self.list_frame.winfo_children():
            widget.destroy()
        if not path.exists():
            ttk.Label(self.list_frame, text="Folder not found.", style="Muted.TLabel").pack(anchor="w", pady=8)
            return
        try:
            folders = sorted([item for item in path.iterdir() if item.is_dir()], key=lambda item: item.name.lower())
            files = sorted([item for item in path.iterdir() if item.is_file()], key=lambda item: item.name.lower())
        except Exception as exc:
            ttk.Label(self.list_frame, text=f"Could not read folder: {exc}", style="Muted.TLabel").pack(anchor="w", pady=8)
            return
        for item in folders + files:
            self.add_item(item, item.is_dir())

    def add_item(self, path, is_folder):
        row = ttk.Frame(self.list_frame, style="Panel.TFrame", padding=(8, 5))
        row.pack(fill="x", pady=2)
        kind = "Folder" if is_folder else "File"
        label = ttk.Label(row, text=f"{kind}  {path.name}", style="NavText.TLabel", cursor="hand2")
        label.pack(side="left", fill="x", expand=True)
        for widget in (row, label):
            widget.bind("<Button-1>", lambda _event=None, w=row, p=path: self.select_item(w, p))
            widget.bind("<Double-Button-1>", lambda _event=None, p=path, folder=is_folder: self.open_item(p, folder))

    def select_item(self, widget, path):
        if self.selected_widget is not None and self.selected_widget.winfo_exists():
            self.selected_widget.configure(style="Panel.TFrame")
        self.selected_widget = widget
        self.selected_path = Path(path)
        try:
            widget.configure(style="Selected.TFrame")
        except Exception:
            pass

    def open_item(self, path, is_folder):
        if is_folder:
            self.show_folder(path)
            return
        try:
            os.startfile(str(path))
        except Exception as exc:
            messagebox.showerror("Could not open file", str(exc))

    def copy_selected(self):
        if self.selected_path is None:
            messagebox.showinfo("Nothing selected", "Select a file or folder first.")
            return
        destination = filedialog.askdirectory(title="Copy selected item to")
        if not destination:
            return
        target = Path(destination) / self.selected_path.name
        try:
            if self.selected_path.is_dir():
                shutil.copytree(self.selected_path, target, dirs_exist_ok=True)
            else:
                shutil.copy2(self.selected_path, target)
        except Exception as exc:
            messagebox.showerror("Could not copy", str(exc))

    def open_current_in_explorer(self):
        try:
            os.startfile(str(self.current_path))
        except Exception as exc:
            messagebox.showerror("Could not open File Explorer", str(exc))


class DualFolderBrowserPage(ttk.Frame):
    def __init__(self, parent, app, root_path, back_command):
        super().__init__(parent)
        self.app = app
        self.root_path = Path(root_path)
        self.back_command = back_command
        self._build()

    def _build(self):
        shell = ttk.Frame(self, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True)
        header = ttk.Frame(shell, style="Panel.TFrame")
        header.pack(fill="x", pady=(0, 12))
        ttk.Label(header, text="LukeStrom MeloVerse Explorer", style="CardTitle.TLabel").pack(side="left")
        ttk.Button(header, text="Back", command=self.back_command).pack(side="right")
        panes = ttk.Frame(shell, style="Panel.TFrame")
        panes.pack(fill="both", expand=True)
        panes.columnconfigure(0, weight=1, uniform="meloverse")
        panes.columnconfigure(1, weight=1, uniform="meloverse")
        panes.rowconfigure(0, weight=1)
        ExplorerPane(panes, self.app, self.root_path, "Left pane").grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ExplorerPane(panes, self.app, self.root_path, "Right pane").grid(row=0, column=1, sticky="nsew", padx=(6, 0))


class ChatGPTLinksPage(ttk.Frame):
    def __init__(self, parent, app, back_command):
        super().__init__(parent)
        self.app = app
        self.back_command = back_command
        self._build()

    def _build(self):
        shell = ttk.Frame(self, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True)
        header = ttk.Frame(shell, style="Panel.TFrame")
        header.pack(fill="x", pady=(0, 14))
        ttk.Label(header, text="ChatGPT links", style="CardTitle.TLabel").pack(side="left")
        ttk.Button(header, text="Back", command=self.back_command).pack(side="right")

        scroller = ScrollFrame(shell)
        scroller.pack(fill="both", expand=True)
        scroller.canvas.configure(bg=self.app.colors["panel_bg"], highlightthickness=0)
        scroller.body.configure(style="Panel.TFrame")
        scroller.set_auto_scrollbar(True)

        grid = scroller.body
        for col in range(3):
            grid.columnconfigure(col, weight=1, uniform="chatgpt")
        for index, (title, url) in enumerate(CHATGPT_LINKS):
            self.create_chatgpt_tile(grid, index // 3, index % 3, title, url)

    def create_chatgpt_tile(self, parent, row, column, title, url):
        card = tk.Canvas(parent, height=126, highlightthickness=0, bg=self.app.colors["panel_bg"], cursor="hand2")
        card.grid(row=row, column=column, sticky="nsew", padx=6, pady=6)

        def draw(_event=None):
            card.delete("all")
            width = max(230, card.winfo_width())
            height = max(126, card.winfo_height())
            color = "#10a37f"
            rounded_rect(card, 0, 0, width, height, radius=10, fill=self.app.colors["panel_bg"], outline=self.app.colors["border"])
            rounded_rect(card, 0, 0, width, 34, radius=10, fill=color, outline="")
            card.create_rectangle(0, 22, width, 34, fill=color, outline="")
            rounded_rect(card, 16, 50, 82, height - 18, radius=10, fill=color, outline="")
            icon = load_tile_icon_image(favicon_url_for_domain("chatgpt.com", 64), 42)
            if icon is not None:
                card.icon_ref = icon
                card.create_oval(26, 58, 72, 104, fill="#ffffff", outline="")
                card.create_image(49, 81, image=icon, anchor="center")
            else:
                card.create_text(49, 81, text="AI", fill="#ffffff", font=("Segoe UI", 17, "bold"))
            card.create_text(102, 54, text=title, fill=self.app.colors["text"], anchor="nw", font=("Segoe UI", 12, "bold"), width=max(120, width - 118))
            card.create_text(102, 84, text="Open in Chrome", fill=self.app.colors["muted"], anchor="nw", font=("Segoe UI", 9))

        card.bind("<Configure>", draw)
        card.bind("<Button-1>", lambda _event=None: self.app.open_in_chrome(url))
        draw()


class PerformanceVuWindow(tk.Toplevel):
    def __init__(self, app):
        super().__init__(app)
        self.app = app
        self.title(f"{APP_TITLE} - Performance VU")
        self.geometry("1280x620")
        self.minsize(920, 460)
        self.configure(bg=self.app.colors["app_bg"])
        self.meter_background_paths = find_tile_backgrounds()
        self.meter_jobs = []
        self.performance_canvases = {}
        self.performance_display_values = {}
        self.performance_running = True
        self.performance_collecting = False
        self.protocol("WM_DELETE_WINDOW", self.close)
        self._build()
        self.refresh_performance()

    def _build(self):
        shell = ttk.Frame(self, padding=18)
        shell.pack(fill="both", expand=True)
        ttk.Label(shell, text="Performance VU", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 12))
        grid = ttk.Frame(shell)
        grid.pack(fill="both", expand=True)
        keys = ["cpu", "memory", "disk", "network", "gpu", "temp"]
        for row in range(2):
            grid.rowconfigure(row, weight=1, uniform="vurows")
        for col in range(3):
            grid.columnconfigure(col, weight=1, uniform="vucols")
        for index, key in enumerate(keys):
            config = PERFORMANCE_METER_OPTIONS[key]
            canvas = tk.Canvas(grid, height=240, bg=self.app.colors["panel_bg"], highlightthickness=1, highlightbackground=self.app.colors["border"])
            canvas.grid(row=index // 3, column=index % 3, sticky="nsew", padx=6, pady=6)
            canvas.slot_key = key
            canvas.meter_title = config["title"]
            canvas.meter_unit = config["unit"]
            canvas.meter_maximum = config["maximum"]
            canvas.meter_color = config["color"]
            canvas.meter_peak_ratio = 0
            canvas.meter_peak_time = 0
            self.performance_canvases[key] = canvas
            canvas.bind("<Configure>", lambda _event, c=canvas: self.draw_vu_meter(c, self.performance_display_values.get(c.slot_key)))
            self.draw_vu_meter(canvas, None)
            self.rotate_meter_background(canvas)

    def close(self):
        self.performance_running = False
        for job in self.meter_jobs:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        self.destroy()

    def rotate_meter_background(self, canvas):
        if not self.performance_running or not canvas.winfo_exists():
            return
        canvas.meter_background_path = random.choice(self.meter_background_paths) if self.meter_background_paths else None
        canvas.meter_face_signature = None
        self.draw_vu_meter(canvas, self.performance_display_values.get(canvas.slot_key))
        job = self.after(get_tile_picture_interval_seconds() * 1000, lambda: self.rotate_meter_background(canvas))
        self.meter_jobs.append(job)

    def draw_vu_meter(self, canvas, value):
        ToolsPage.draw_vu_meter(self, canvas, value)

    def refresh_performance(self):
        if not self.performance_running or not self.winfo_exists():
            return
        if self.performance_collecting:
            self.after(500, self.refresh_performance)
            return
        self.performance_collecting = True

        def worker():
            try:
                snapshot = collect_performance_snapshot()
            except Exception:
                snapshot = {}

            def finish():
                self.performance_collecting = False
                if self.performance_running and self.winfo_exists():
                    self.update_performance(snapshot)

            if self.performance_running and self.winfo_exists():
                self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()
        self.after(500, self.refresh_performance)

    def update_performance(self, snapshot):
        for key, canvas in self.performance_canvases.items():
            new_value = snapshot.get(key)
            previous = self.performance_display_values.get(key)
            if previous is None or new_value is None:
                display_value = new_value
            else:
                display_value = previous + (new_value - previous) * 0.45
            self.performance_display_values[key] = display_value
            self.draw_vu_meter(canvas, display_value)


class HomePage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.background_paths = []
        self.tile_canvases = []
        self.tile_jobs = []
        self._build()
        self.after(120, self.load_backgrounds)

    def _build(self):
        tiles = ttk.Frame(self)
        tiles.pack(fill="both", expand=True)
        tiles.columnconfigure((0, 1, 2), weight=1, uniform="tiles")
        tiles.rowconfigure((0, 1), weight=1, uniform="tiles")

        self._tile(tiles, 0, 0, "Song Analyzer", self.app.show_song_analyzer, "#2f6f73")
        self._tile(tiles, 0, 1, "Reel Design", self.app.show_reel_design, "#7b5f2a")
        self._tile(tiles, 0, 2, "Campaign Planner", self.app.show_post_planner, "#4f6f8f")
        self._tile(tiles, 1, 0, "Tools", self.app.show_audio_vu, "#8d3f3f")
        self._tile(tiles, 1, 1, "Shortcuts", self.app.show_tools, "#5b6770")
        self._tile(tiles, 1, 2, "Metrics", self.app.show_metrics, "#5b6770")

    def _tile(self, parent, row, column, title, command, color):
        card = tk.Frame(parent, bg=self.app.colors["panel_bg"], highlightthickness=0, cursor="hand2")
        card.grid(row=row, column=column, sticky="nsew", padx=8, pady=8)
        card.grid_propagate(False)

        canvas = tk.Canvas(card, highlightthickness=0, bg=self.app.colors["panel_bg"])
        canvas.pack(fill="both", expand=True)
        canvas.bind("<Configure>", lambda event=None: self.on_tile_resize(canvas))
        click_command = lambda: self.app.run_with_loading(command)
        canvas.bind("<Button-1>", lambda _event=None: click_command())
        canvas.bind("<Enter>", lambda _event: canvas.configure(cursor="hand2"))
        canvas.bind("<Leave>", lambda _event: canvas.configure(cursor=""))
        rounded_rect(canvas, 0, 0, 2000, 10, radius=10, fill=color, outline=color, tags=("accent",))
        shadow_item = canvas.create_text(0, 0, text=title, fill="#ffffff", font=("Segoe UI", 18, "bold"), anchor="center", justify="center", width=280, tags=("title_shadow",))
        title_item = canvas.create_text(0, 0, text=title, fill="#111111", font=("Segoe UI", 18, "bold"), anchor="center", justify="center", width=280, tags=("title",))
        canvas.title_items = [shadow_item, title_item]
        canvas.fallback_color = color
        self.tile_canvases.append(canvas)

        self.rotate_tile_background(canvas, color)
        self.rotate_tile_font(canvas)


    def _display_tile(self, parent, row, column, title, color):
        card = tk.Frame(parent, bg=self.app.colors["panel_bg"], highlightthickness=0)
        card.grid(row=row, column=column, sticky="nsew", padx=8, pady=8)
        card.grid_propagate(False)
        canvas = tk.Canvas(card, highlightthickness=0, bg=self.app.colors["panel_bg"])
        canvas.pack(fill="both", expand=True)
        canvas.bind("<Configure>", lambda event=None: self.on_tile_resize(canvas))
        rounded_rect(canvas, 0, 0, 2000, 10, radius=10, fill=color, outline=color, tags=("accent",))
        shadow_item = canvas.create_text(0, 0, text=title, fill="#ffffff", font=("Segoe UI", 18, "bold"), anchor="center", justify="center", width=280, tags=("title_shadow",))
        title_item = canvas.create_text(0, 0, text=title, fill="#111111", font=("Segoe UI", 18, "bold"), anchor="center", justify="center", width=280, tags=("title",))
        canvas.title_items = [shadow_item, title_item]
        canvas.fallback_color = color
        self.tile_canvases.append(canvas)
        self.rotate_tile_background(canvas, color)
        self.rotate_tile_font(canvas)

    def load_backgrounds(self):
        if not self.winfo_exists():
            return
        self.background_paths = find_tile_backgrounds()
        if not self.background_paths:
            return
        for canvas in list(self.tile_canvases):
            if canvas.winfo_exists():
                self.rotate_tile_background(canvas, getattr(canvas, "fallback_color", "#5b6770"))

    def rotate_tile_background(self, canvas, fallback_color):
        if not canvas.winfo_exists():
            return

        if self.background_paths:
            path = random.choice(self.background_paths)
            try:
                canvas.background_path = path
                image = image_for_tile(path, canvas.winfo_width(), canvas.winfo_height())
                canvas.image_ref = image
                if hasattr(canvas, "background_item"):
                    canvas.itemconfigure(canvas.background_item, image=image)
                else:
                    canvas.background_item = canvas.create_image(0, 0, image=image, anchor="nw", tags=("background",))
                    canvas.tag_lower(canvas.background_item)
            except Exception:
                canvas.configure(bg=self.app.colors["panel_bg"])
        else:
            canvas.configure(bg=self.app.colors["panel_bg"])

        interval_ms = get_tile_picture_interval_seconds() * 1000
        old_job = getattr(canvas, "background_job", None)
        if old_job is not None:
            try:
                self.after_cancel(old_job)
            except Exception:
                pass
        canvas.background_job = self.after(interval_ms, lambda: self.rotate_tile_background(canvas, fallback_color))
        self.tile_jobs.append(canvas.background_job)

    def rotate_tile_font(self, canvas):
        if not canvas.winfo_exists():
            return

        font_name = random.choice(tile_title_fonts())
        base_size = 16
        min_percent, max_percent = get_tile_font_size_range()
        if min_percent == max_percent:
            percent = min_percent
        else:
            percent = random.randint(min_percent, max_percent)
        font_size = max(8, int(base_size * percent / 100))
        for item in getattr(canvas, "title_items", []):
            canvas.itemconfigure(item, font=(font_name, font_size, "bold"))

        interval_ms = get_tile_font_interval_seconds() * 1000
        old_job = getattr(canvas, "font_job", None)
        if old_job is not None:
            try:
                self.after_cancel(old_job)
            except Exception:
                pass
        canvas.font_job = self.after(interval_ms, lambda: self.rotate_tile_font(canvas))
        self.tile_jobs.append(canvas.font_job)

    def on_tile_resize(self, canvas):
        self.resize_tile_background(canvas)
        self.position_tile_content(canvas)

    def position_tile_content(self, canvas):
        width = max(1, canvas.winfo_width())
        height = max(1, canvas.winfo_height())
        center_x = width / 2
        center_y = height / 2
        text_width = max(160, min(width - 40, 520))

        if hasattr(canvas, "button_window"):
            title_y = center_y - 46
            canvas.coords(canvas.button_window, center_x, center_y + 24)
        else:
            title_y = center_y

        items = getattr(canvas, "title_items", [])
        if len(items) >= 2:
            canvas.coords(items[0], center_x + 2, title_y + 2)
            canvas.itemconfigure(items[0], width=text_width)
            canvas.coords(items[1], center_x, title_y)
            canvas.itemconfigure(items[1], width=text_width)
        elif items:
            canvas.coords(items[0], center_x, title_y)
            canvas.itemconfigure(items[0], width=text_width)

    def resize_tile_background(self, canvas):
        path = getattr(canvas, "background_path", None)
        if not path or not canvas.winfo_exists():
            return
        try:
            image = image_for_tile(path, canvas.winfo_width(), canvas.winfo_height())
            canvas.image_ref = image
            if hasattr(canvas, "background_item"):
                canvas.itemconfigure(canvas.background_item, image=image)
        except Exception:
            pass


class AudioVuPage(ttk.Frame):
    def __init__(self, parent, app, embedded_canvases=None, controls_parent=None, show_source_controls=True):
        super().__init__(parent)
        self.app = app
        self.embedded_canvases = embedded_canvases
        self.controls_parent = controls_parent
        self.show_source_controls = show_source_controls
        self.background_paths = find_tile_backgrounds()
        self.running = True
        self.stream = None
        self.pycaw_meter = None
        self.soundcard_thread = None
        self.audio_stop_event = threading.Event()
        self.left_level = 0.0
        self.right_level = 0.0
        self.raw_left_level = 0.0
        self.raw_right_level = 0.0
        self.display_left = 0.0
        self.display_right = 0.0
        self.audio_source_name = ""
        self.audio_start_note = ""
        self.audio_sources = []
        self.audio_source_signals = {}
        self.selected_audio_source_index = None
        self.last_source_label_update = 0
        self.background_job = None
        self.update_job = None
        self._build()

    def _build(self):
        if self.embedded_canvases is not None:
            self.left_canvas, self.right_canvas = self.embedded_canvases
            self.left_canvas.vu_title = "L"
            self.right_canvas.vu_title = "R"
            self.audio_title_var = tk.StringVar(value="Audio VU")
            self.left_canvas.bind("<Configure>", lambda _event=None: self.draw_audio_meter(self.left_canvas, self.display_left))
            self.right_canvas.bind("<Configure>", lambda _event=None: self.draw_audio_meter(self.right_canvas, self.display_right))
            controls = self.controls_parent or self
            self.status_var = tk.StringVar(value="")
            self.engine_var = tk.StringVar(value="")
            self.source_list = tk.Listbox(controls, height=1, activestyle="none", exportselection=False, font=("Segoe UI", 9), width=38)
            self.source_list.configure(
                bg=self.app.colors["entry_bg"],
                fg=self.app.colors["entry_fg"],
                selectbackground="#4f6f8f",
                selectforeground="#ffffff",
                highlightbackground=self.app.colors["border"],
            )
            if self.show_source_controls:
                ttk.Label(controls, text="Audio source", style="CardText.TLabel").pack(side="left", padx=(0, 6))
                self.source_list.pack(side="left", fill="x", expand=True)
            self.source_list.bind("<<ListboxSelect>>", self.on_audio_source_select)
            self.rotate_backgrounds()
            self.start_audio()
            self.update_meters()
            return

        shell = ttk.Frame(self, style="Panel.TFrame", padding=18)
        shell.pack(fill="both", expand=True, pady=(0, 14))
        header = ttk.Frame(shell, style="Panel.TFrame")
        header.pack(fill="x", pady=(0, 14))
        self.audio_title_var = tk.StringVar(value="Audio VU")
        ttk.Label(header, textvariable=self.audio_title_var, style="CardTitle.TLabel").pack(side="left")

        meter_frame = ttk.Frame(shell, style="Panel.TFrame")
        meter_frame.pack(fill="both", expand=True)
        meter_frame.columnconfigure(0, weight=1)
        meter_frame.rowconfigure((0, 1), weight=1, uniform="audiovu")
        self.left_canvas = tk.Canvas(meter_frame, bg=self.app.colors["panel_bg"], highlightthickness=0, bd=0)
        self.right_canvas = tk.Canvas(meter_frame, bg=self.app.colors["panel_bg"], highlightthickness=0, bd=0)
        self.left_canvas.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        self.right_canvas.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        self.left_canvas.vu_title = "L"
        self.right_canvas.vu_title = "R"
        self.left_canvas.bind("<Configure>", lambda _event=None: self.draw_audio_meter(self.left_canvas, self.display_left))
        self.right_canvas.bind("<Configure>", lambda _event=None: self.draw_audio_meter(self.right_canvas, self.display_right))

        self.status_var = tk.StringVar(value="")
        self.engine_var = tk.StringVar(value="")
        self.source_list = tk.Listbox(shell, height=1, activestyle="none", exportselection=False, font=("Segoe UI", 9))
        self.source_list.configure(
            bg=self.app.colors["entry_bg"],
            fg=self.app.colors["entry_fg"],
            selectbackground="#4f6f8f",
            selectforeground="#ffffff",
            highlightbackground=self.app.colors["border"],
        )
        self.source_list.bind("<<ListboxSelect>>", self.on_audio_source_select)
        ttk.Label(shell, text="Audio source", style="CardText.TLabel").pack(anchor="w", pady=(0, 4))
        self.source_list.pack(fill="x", pady=(0, 12))
        self.rotate_backgrounds()
        self.start_audio()
        self.update_meters()

    def destroy(self):
        self.running = False
        self.stop_audio()
        for job in (self.background_job, self.update_job):
            if job is not None:
                try:
                    self.after_cancel(job)
                except Exception:
                    pass
        super().destroy()

    def start_audio(self):
        self.stop_audio()
        self.audio_stop_event.clear()
        if self.start_windows_peak_meter():
            return
        if np is None:
            self.status_var.set("Audio VU needs numpy for audio capture fallback.")
            return
        if sc is not None:
            try:
                microphones = self.soundcard_loopback_microphones()
                self.audio_sources = microphones
                self.refresh_audio_source_list()
                self.soundcard_thread = threading.Thread(
                    target=self.soundcard_scan_loop,
                    args=(microphones,),
                    daemon=True,
                )
                self.soundcard_thread.start()
                self.audio_source_name = "scanning loopback sources"
                prefix = f"{self.audio_start_note} | " if self.audio_start_note else ""
                self.status_var.set(f"{prefix}Scanning {len(microphones)} loopback source(s) for system audio...")
                return
            except Exception as exc:
                self.status_var.set(f"Soundcard loopback failed, trying fallback: {exc}")

        if sd is None:
            self.status_var.set("Audio VU needs soundcard or sounddevice. For YouTube, install soundcard: py -m pip install soundcard")
            return
        errors = []
        for candidate in self.audio_capture_candidates():
            try:
                self.stream = sd.InputStream(
                    device=candidate["device"],
                    channels=candidate["channels"],
                    samplerate=candidate["samplerate"],
                    blocksize=1024,
                    extra_settings=candidate["extra"],
                    callback=self.audio_callback,
                )
                self.stream.start()
                self.audio_source_name = candidate["name"]
                self.status_var.set(f"Listening: {candidate['name']}")
                return
            except Exception as exc:
                self.stream = None
                errors.append(f"{candidate['name']}: {exc}")
        detail = errors[-1] if errors else "No usable audio capture device found."
        self.status_var.set(f"Audio capture not available: {detail}. For YouTube, try: py -m pip install soundcard")

    def start_windows_peak_meter(self):
        if AudioUtilities is None or IAudioMeterInformation is None or CLSCTX_ALL is None:
            detail = f": {PYCAW_IMPORT_ERROR}" if PYCAW_IMPORT_ERROR else ""
            self.audio_start_note = f"Windows peak meter unavailable: pycaw/comtypes import failed{detail}"
            self.engine_var.set(self.audio_start_note)
            return False
        try:
            if comtypes is not None:
                try:
                    comtypes.CoInitialize()
                except Exception:
                    pass
            speakers = AudioUtilities.GetSpeakers()
            self.pycaw_meter = self.audio_meter_from_pycaw_device(speakers)
            self.audio_source_name = "Windows output peak meter"
            self.audio_sources = []
            self.source_list.delete(0, tk.END)
            self.source_list.insert(tk.END, "Windows output peak meter")
            self.source_list.selection_set(0)
            self.engine_var.set("Windows peak meter: active")
            self.status_var.set("Listening: Windows output peak meter")
            return True
        except Exception as exc:
            self.pycaw_meter = None
            self.audio_start_note = f"Windows peak meter failed: {exc}"
            self.engine_var.set(self.audio_start_note)
            if hasattr(self, "source_list") and self.source_list.winfo_exists():
                self.source_list.delete(0, tk.END)
                self.source_list.insert(tk.END, self.audio_start_note)
            return False

    def audio_meter_from_pycaw_device(self, device):
        direct_meter = self.existing_audio_meter(device)
        if direct_meter is not None:
            return direct_meter

        for activator in self.pycaw_activators(device):
            try:
                interface = activator(IAudioMeterInformation._iid_, CLSCTX_ALL, None)
                meter = interface.QueryInterface(IAudioMeterInformation)
                if hasattr(meter, "GetPeakValue"):
                    return meter
            except Exception:
                continue

        useful_attrs = []
        try:
            useful_attrs = [name for name in dir(device) if "activ" in name.lower() or "meter" in name.lower() or "device" in name.lower()]
        except Exception:
            useful_attrs = []
        raise RuntimeError(f"{type(device).__name__} has no usable output peak meter route; attrs: {', '.join(useful_attrs[:12])}")

    def existing_audio_meter(self, obj):
        for attr_name in ("AudioMeterInformation", "audio_meter", "AudioMeter", "meter", "_audio_meter", "_meter"):
            try:
                candidate = getattr(obj, attr_name, None)
                if candidate is None:
                    continue
                if callable(candidate):
                    candidate = candidate()
                if hasattr(candidate, "GetPeakValue"):
                    return candidate
            except Exception:
                continue
        return None

    def pycaw_activators(self, device):
        seen = set()

        def add_object(obj):
            if obj is None or id(obj) in seen:
                return
            seen.add(id(obj))
            for method_name in ("Activate", "activate"):
                try:
                    method = getattr(obj, method_name, None)
                    if callable(method):
                        yield method
                except Exception:
                    pass

        for activator in add_object(device):
            yield activator

        for attr_name in ("_device", "device", "_dev", "dev", "endpoint", "_endpoint", "_AudioDevice__device"):
            try:
                inner = getattr(device, attr_name, None)
            except Exception:
                inner = None
            for activator in add_object(inner):
                yield activator

    def on_audio_source_select(self, _event=None):
        selection = self.source_list.curselection()
        if not selection or selection[0] == 0:
            self.selected_audio_source_index = None
            self.status_var.set("Audio source: Auto select strongest")
            return
        self.selected_audio_source_index = selection[0] - 1
        if 0 <= self.selected_audio_source_index < len(self.audio_sources):
            name = getattr(self.audio_sources[self.selected_audio_source_index], "name", "selected source")
            self.status_var.set(f"Audio source selected: {name}")

    def refresh_audio_source_list(self):
        if not hasattr(self, "source_list") or not self.source_list.winfo_exists():
            return
        current_selection = self.selected_audio_source_index
        self.source_list.delete(0, tk.END)
        self.source_list.insert(tk.END, "Auto select strongest source")
        for index, microphone in enumerate(self.audio_sources):
            name = getattr(microphone, "name", str(microphone))
            signal = self.audio_source_signals.get(index, 0.0)
            self.source_list.insert(tk.END, f"{name} | {signal * 100:.3f}%")
        select_row = 0 if current_selection is None else current_selection + 1
        if self.source_list.size() > select_row:
            self.source_list.selection_set(select_row)

    def soundcard_loopback_microphones(self):
        speaker = sc.default_speaker()
        microphones = []
        try:
            microphones.append(sc.get_microphone(speaker.name, include_loopback=True))
        except Exception:
            pass
        try:
            microphones.extend(sc.all_microphones(include_loopback=True))
        except Exception:
            pass

        unique = []
        seen = set()
        for microphone in microphones:
            key = getattr(microphone, "id", None) or getattr(microphone, "name", repr(microphone))
            if key in seen:
                continue
            seen.add(key)
            unique.append(microphone)

        if not unique:
            raise RuntimeError("No soundcard loopback microphone found.")
        return unique

    def soundcard_scan_loop(self, microphones):
        last_status = 0
        while self.running and not self.audio_stop_event.is_set():
            best_data = None
            best_signal = -1.0
            best_name = ""
            errors = 0
            selected = self.selected_audio_source_index
            if selected is not None and 0 <= selected < len(microphones):
                scan_items = [(selected, microphones[selected])]
            else:
                scan_items = list(enumerate(microphones))
            for index, microphone in scan_items:
                if self.audio_stop_event.is_set():
                    break
                try:
                    with microphone.recorder(samplerate=48000, channels=2) as recorder:
                        data = recorder.record(numframes=4096)
                    signal = self.audio_raw_signal(data)
                    self.audio_source_signals[index] = signal
                    if signal > best_signal:
                        best_signal = signal
                        best_data = data
                        best_name = getattr(microphone, "name", str(microphone))
                except Exception:
                    errors += 1
            if best_data is not None:
                self.audio_source_name = f"scan:{best_name}"
                self.update_audio_levels(best_data)
            now = time_module.monotonic()
            if now - last_status > 1.0 and self.running and self.winfo_exists():
                source = best_name or "no usable source"
                signal = max(0.0, best_signal)
                mode = "Selected" if selected is not None else "Auto"
                note = f"{self.audio_start_note} | " if self.audio_start_note else ""
                self.after(0, lambda src=source, sig=signal, count=len(scan_items), err=errors, mode=mode, note=note: (
                    self.status_var.set(f"{note}{mode}: {src} | signal {sig * 100:.3f}% | errors {err}"),
                    self.refresh_audio_source_list(),
                ))
                last_status = now

    def audio_raw_signal(self, data):
        try:
            data = np.asarray(data, dtype=float)
            if data.size == 0:
                return 0.0
            return float(np.sqrt(np.mean(np.square(data))))
        except Exception:
            return 0.0

    def audio_capture_candidates(self):
        devices = list(sd.query_devices())
        hostapis = sd.query_hostapis()
        wasapi_indices = {
            index for index, hostapi in enumerate(hostapis)
            if "wasapi" in str(hostapi.get("name", "")).lower()
        }
        default_output_index = None
        default_input_index = None
        try:
            default_output_index = sd.default.device[1]
            default_input_index = sd.default.device[0]
        except Exception:
            pass
        default_output_name = ""
        if isinstance(default_output_index, int) and 0 <= default_output_index < len(devices):
            default_output_name = str(devices[default_output_index].get("name", "")).lower()

        candidates = []
        for index, device in enumerate(devices):
            name = str(device.get("name", ""))
            name_lower = name.lower()
            hostapi = int(device.get("hostapi", -1))
            input_channels = int(device.get("max_input_channels", 0) or 0)
            score = 0
            if hostapi in wasapi_indices:
                score += 20
            if "loopback" in name_lower or "what u hear" in name_lower or "stereo mix" in name_lower:
                score += 40
            if default_output_name and default_output_name[:18] in name_lower:
                score += 15
            if input_channels > 0 and score > 0:
                candidates.append((score, index, device, input_channels, name))

        if isinstance(default_output_index, int) and 0 <= default_output_index < len(devices):
            device = devices[default_output_index]
            hostapi = int(device.get("hostapi", -1))
            output_channels = int(device.get("max_output_channels", 0) or 0)
            if output_channels > 0:
                candidates.append((35, default_output_index, device, min(2, output_channels), f"{device.get('name', 'Default output')} (system output)"))

        if isinstance(default_input_index, int) and 0 <= default_input_index < len(devices):
            device = devices[default_input_index]
            input_channels = int(device.get("max_input_channels", 0) or 0)
            if input_channels > 0:
                candidates.append((5, default_input_index, device, min(2, input_channels), f"{device.get('name', 'Default input')} (input fallback)"))

        candidates.sort(reverse=True, key=lambda item: item[0])
        seen = set()
        result = []
        for _score, index, device, channels, name in candidates:
            key = (index, channels, name)
            if key in seen:
                continue
            seen.add(key)
            hostapi = int(device.get("hostapi", -1))
            extra = sd.WasapiSettings(exclusive=False) if hasattr(sd, "WasapiSettings") and hostapi in wasapi_indices else None
            result.append({
                "device": index,
                "channels": max(1, min(2, channels)),
                "samplerate": int(device.get("default_samplerate") or 44100),
                "extra": extra,
                "name": name,
            })
        return result

    def stop_audio(self):
        self.audio_stop_event.set()
        self.pycaw_meter = None
        if self.stream is not None:
            try:
                self.stream.stop()
                self.stream.close()
            except Exception:
                pass
            self.stream = None

    def update_windows_peak_meter(self):
        if self.pycaw_meter is None:
            return
        try:
            left = right = float(self.pycaw_meter.GetPeakValue())
            try:
                channel_count = int(self.pycaw_meter.GetMeteringChannelCount())
                if channel_count > 0:
                    peaks = self.pycaw_meter.GetChannelsPeakValues(channel_count)
                    if peaks:
                        left = float(peaks[0])
                        right = float(peaks[1] if len(peaks) > 1 else peaks[0])
            except Exception:
                pass
            self.raw_left_level = left
            self.raw_right_level = right
            self.left_level = self.windows_peak_to_meter(left)
            self.right_level = self.windows_peak_to_meter(right)
        except Exception as exc:
            self.pycaw_meter = None
            self.status_var.set(f"Windows peak meter stopped: {exc}")

    def update_audio_title(self):
        self.audio_title_var.set("Audio VU")

    def audio_callback(self, indata, _frames, _time, status):
        if status:
            message = str(status)
            if self.running and self.winfo_exists():
                self.after(0, lambda: self.status_var.set(message))
        self.update_audio_levels(indata)

    def update_audio_levels(self, data):
        try:
            data = np.asarray(data, dtype=float)
            if data.ndim == 1:
                data = data.reshape((-1, 1))
            levels = np.sqrt(np.mean(np.square(data), axis=0))
            left_raw = float(levels[0])
            right_raw = float(levels[1] if len(levels) > 1 else levels[0])
            self.raw_left_level = left_raw
            self.raw_right_level = right_raw
            self.left_level = self.audio_level_to_meter(left_raw)
            self.right_level = self.audio_level_to_meter(right_raw)
        except Exception:
            pass

    def audio_level_to_meter(self, raw_level):
        if raw_level <= 0.00005:
            return 0.0
        linear = raw_level * 30
        db = 20 * math.log10(max(raw_level, 0.000001))
        db_scaled = (db + 55) / 40
        return max(0.0, min(1.0, max(linear, db_scaled)))

    def windows_peak_to_meter(self, raw_level):
        if raw_level <= 0.002:
            return 0.0
        # Windows exposes a very fast peak value. Keep headroom so loud browser
        # audio moves musically instead of pinning the needle to 100%.
        compressed = math.pow(min(1.0, max(0.0, raw_level)), 0.72)
        return max(0.0, min(0.96, compressed * 0.78))

    def rotate_backgrounds(self):
        if not self.running:
            return
        for canvas in (self.left_canvas, self.right_canvas):
            if self.background_paths:
                canvas.background_path = random.choice(self.background_paths)
            else:
                canvas.background_path = None
            canvas.background_signature = None
        self.draw_audio_meter(self.left_canvas, self.display_left)
        self.draw_audio_meter(self.right_canvas, self.display_right)
        self.background_job = self.after(get_tile_picture_interval_seconds() * 1000, self.rotate_backgrounds)

    def update_meters(self):
        if not self.running:
            return
        self.update_windows_peak_meter()
        left_speed = 0.48 if self.left_level > self.display_left else 0.16
        right_speed = 0.48 if self.right_level > self.display_right else 0.16
        self.display_left += (self.left_level - self.display_left) * left_speed
        self.display_right += (self.right_level - self.display_right) * right_speed
        if self.stream is None and sd is not None and np is not None:
            self.left_level *= 0.9
            self.right_level *= 0.9
        self.left_canvas.audio_raw_level = self.raw_left_level
        self.right_canvas.audio_raw_level = self.raw_right_level
        self.draw_audio_meter(self.left_canvas, self.display_left)
        self.draw_audio_meter(self.right_canvas, self.display_right)
        self.update_audio_title()
        if self.audio_source_name and not self.audio_source_name.startswith("scan:"):
            self.status_var.set(
                f"Listening to system audio: {self.audio_source_name} | "
                f"L {self.raw_left_level * 100:.2f}% / R {self.raw_right_level * 100:.2f}%"
            )
        self.update_job = self.after(33, self.update_meters)

    def draw_audio_meter(self, canvas, level):
        canvas.update_idletasks()
        width = max(320, canvas.winfo_width())
        height = max(220, canvas.winfo_height())
        title = getattr(canvas, "vu_title", "")
        face = "#202020" if self.app.dark_mode else "#ffffff"
        border = "#3d3d3d" if self.app.dark_mode else "#b8b0a2"
        ink = "#f3f0e8" if self.app.dark_mode else "#1c1710"
        red = "#d23a2e"
        center_x = width / 2
        center_y = height * 1.08
        radius = min(width * 0.50, height * 0.95)
        start_angle = 132
        end_angle = 48
        sweep = start_angle - end_angle

        def point_for(percent, extra_radius=0):
            angle = math.radians(start_angle - sweep * percent)
            r = radius + extra_radius
            return center_x + math.cos(angle) * r, center_y - math.sin(angle) * r

        background_path = getattr(canvas, "background_path", None)
        signature = (
            width,
            height,
            self.app.dark_mode,
            get_vu_artwork_opacity_percent(),
            str(background_path) if background_path else "",
        )
        if getattr(canvas, "background_signature", None) != signature:
            canvas.delete("all")
            canvas.background_signature = signature
            canvas.configure(bg=face)
            canvas.create_rectangle(0, 0, width, height, fill=face, outline="")
            if background_path and Image is not None and ImageTk is not None:
                try:
                    image = image_for_meter(background_path, width, height, face)
                    canvas.background_ref = image
                    canvas.create_image(0, 0, image=image, anchor="nw")
                except Exception:
                    canvas.background_ref = None
            for idx in range(0, 41):
                percent = idx / 40
                tick_len = 34 if idx % 10 == 0 else 22 if idx % 5 == 0 else 12
                outer = point_for(percent)
                inner = point_for(percent, -tick_len)
                canvas.create_line(inner[0], inner[1], outer[0], outer[1], fill=red if percent >= 0.72 else ink, width=2 if idx % 10 == 0 else 1)
            canvas.create_text(22, 20, text=title, fill=ink, anchor="nw", font=("Segoe UI", 14))

        canvas.delete("audio_dynamic")
        level = max(0, min(1, level))
        raw_level = max(0.0, float(getattr(canvas, "audio_raw_level", 0.0) or 0.0))
        dbfs_text = "-inf dBFS" if raw_level <= 0.000001 else f"{20 * math.log10(min(1.0, raw_level)):.1f} dBFS"
        now = time_module.time()
        peak_ratio = getattr(canvas, "audio_peak_ratio", 0.0)
        peak_time = getattr(canvas, "audio_peak_time", 0.0)
        if now - peak_time >= 5 or level > peak_ratio:
            canvas.audio_peak_ratio = level
            canvas.audio_peak_time = now
            peak_ratio = level
        if peak_ratio > 0:
            peak_outer = point_for(peak_ratio, 14)
            peak_inner = point_for(peak_ratio, -20)
            canvas.create_line(peak_inner[0], peak_inner[1], peak_outer[0], peak_outer[1], fill=red, width=3, tags=("audio_dynamic", "audio_peak"))
        needle_tip = point_for(level, -10)
        canvas.create_line(center_x, height + 24, needle_tip[0], needle_tip[1], fill=ink, width=3, tags=("audio_dynamic",))
        canvas.create_text(width - 22, 22, text=dbfs_text, fill=ink, anchor="ne", font=("Segoe UI", 17), tags=("audio_dynamic",))
        rounded_rect(canvas, 1, 1, width - 2, height - 2, radius=10, outline=border, width=1, fill="", tags=("audio_dynamic",))


class SongAnalyzerPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.module = None
        self.results = None
        self.spectrum_canvas = None
        self.band_canvas = None
        self._build()

    def _build(self):
        try:
            self.module = load_song_analyzer_module()
        except Exception as exc:
            ttk.Label(self, text="Song Analyzer could not be loaded.", style="PageTitle.TLabel").pack(anchor="w")
            ttk.Label(self, text=str(exc), style="Muted.TLabel", wraplength=900).pack(anchor="w", pady=(10, 0))
            return

        input_panel = ttk.Frame(self, style="Panel.TFrame", padding=16)
        input_panel.pack(fill="x", pady=(0, 14))
        ttk.Label(input_panel, text="Select an audio file to analyze.", style="CardText.TLabel").pack(side="left")
        ttk.Button(input_panel, text="Select Audio File", command=self.select_file).pack(side="left", padx=(16, 0))
        ttk.Button(input_panel, text="Export PDF", command=self.export_pdf).pack(side="left", padx=(8, 0))
        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(input_panel, textvariable=self.status_var, style="CardText.TLabel").pack(side="left", padx=(14, 0))

        panes = ttk.Panedwindow(self, orient="horizontal")
        panes.pack(fill="both", expand=True)
        left = ttk.Frame(panes)
        right = ttk.Frame(panes)
        panes.add(left, weight=1)
        panes.add(right, weight=2)

        self.summary = tk.Text(left, width=44, height=32, wrap="word", font=("Segoe UI", 10), padx=10, pady=10)
        self.app.style_text_widget(self.summary)
        self.summary.pack(fill="both", expand=True)
        self.plot_frame = ttk.Frame(right)
        self.plot_frame.pack(fill="both", expand=True)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select audio file",
            filetypes=[("Audio files", "*.wav *.mp3 *.flac *.m4a *.aac *.ogg"), ("All files", "*.*")],
        )
        if file_path:
            self.load_audio_file(file_path)

    def load_audio_file(self, file_path):
        path = Path(file_path)
        if not path.exists():
            messagebox.showerror("File not found", str(path))
            return
        if path.suffix.lower() not in self.module.AUDIO_EXTS:
            messagebox.showerror("Unsupported file", "Please select a WAV, MP3, FLAC, M4A, AAC or OGG file.")
            return
        self.status_var.set("Analyzing...")
        self.summary.delete("1.0", tk.END)
        self.summary.insert(tk.END, f"Analyzing:\n{path}\n\nPlease wait...")
        threading.Thread(target=self._run_analysis, args=(str(path),), daemon=True).start()

    def _run_analysis(self, file_path):
        try:
            results = self.module.analyze_file(file_path)
            self.after(0, lambda: self.show_results(results))
        except Exception as exc:
            self.after(0, lambda: self.show_error(exc))

    def show_error(self, error):
        self.status_var.set("Error.")
        messagebox.showerror("Analysis error", str(error))

    def show_results(self, results):
        self.results = results
        self.status_var.set("Analysis complete.")
        self.summary.delete("1.0", tk.END)
        lufs_txt = f"{results['lufs_integrated']} LUFS" if results["lufs_integrated"] is not None else "n/a"
        true_peak_txt = f"{results['true_peak']} dBTP" if results.get("true_peak") is not None else "n/a"
        tempo_txt = f"{results['tempo_bpm']} BPM" if results["tempo_bpm"] else "n/a"
        key_txt = results["key"] if results.get("key") else "n/a"
        key_conf = f"{results['key_confidence']:.3f}" if results.get("key_confidence") is not None else "n/a"
        lines = [
            "YOUTUBE CHECK", "=" * 30, results["youtube_verdict"], results["youtube_note"],
            f"Loudness: {lufs_txt}", f"True Peak: {true_peak_txt}", "",
            "SUMMARY", "=" * 30, f"File: {results['file']}", f"Duration: {results['duration_sec']} sec",
            f"Sample rate: {results['samplerate']} Hz", f"Tempo: {tempo_txt}", f"Estimated key: {key_txt}",
            f"Key confidence: {key_conf}", f"Tonal tilt: {results['tilt_note']}", "",
            "EQ NOTES", "=" * 30, "Positive = too much in that area.", "Negative = too little in that area.", "",
        ]
        for row in results["bands"]:
            lines.append(f"{row['Band']} ({row['Range']}): {row['AvgDev']} -> {row['EQ']}")
        self.summary.insert(tk.END, "\n".join(lines))
        self.draw_plots(results)

    def draw_plots(self, results):
        for widget in self.plot_frame.winfo_children():
            widget.destroy()
        spec_fig = self.module.build_spectrum_figure(results)
        band_fig = self.module.build_band_figure(results)
        spec_canvas = self.module.FigureCanvasTkAgg(spec_fig, master=self.plot_frame)
        spec_canvas.draw()
        spec_canvas.get_tk_widget().pack(fill="both", expand=True)
        band_canvas = self.module.FigureCanvasTkAgg(band_fig, master=self.plot_frame)
        band_canvas.draw()
        band_canvas.get_tk_widget().pack(fill="both", expand=True)
        self.spectrum_canvas = spec_canvas
        self.band_canvas = band_canvas

    def export_pdf(self):
        if not self.results:
            messagebox.showinfo("No analysis", "Analyze a song first.")
            return
        original = Path(self.results["file_path"])
        output_path = filedialog.asksaveasfilename(
            title="Save PDF report",
            defaultextension=".pdf",
            initialdir=str(DOWNLOADS_PATH),
            initialfile=f"{original.stem}_analysis.pdf",
            filetypes=[("PDF files", "*.pdf")],
        )
        if output_path:
            self.module.write_pdf(self.results, Path(output_path))
            messagebox.showinfo("PDF saved", output_path)


class ExpandingText(tk.Text):
    def __init__(self, parent, min_height=1, max_height=5, **kwargs):
        super().__init__(parent, height=min_height, **kwargs)
        self.min_height = min_height
        self.max_height = max_height
        self.bind("<KeyRelease>", self._resize)
        self.bind("<FocusOut>", self._resize)

    def _resize(self, _event=None):
        line_count = int(self.index("end-1c").split(".")[0])
        self.configure(height=max(self.min_height, min(self.max_height, line_count)))


class ReelDesignPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.creative_boxes = []
        self.song_title_var = tk.StringVar()
        self.post_count_var = tk.IntVar(value=4)
        self._build()

    def _build(self):
        scroller = ScrollFrame(self)
        scroller.pack(fill="both", expand=True)
        main = scroller.body
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=1)

        top = ttk.Frame(main, style="Panel.TFrame", padding=16)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        ttk.Label(top, text="Song title", style="CardText.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.song_title_var, width=54).grid(row=1, column=0, sticky="ew", pady=(5, 12))
        ttk.Label(top, text="Number of posts", style="CardText.TLabel").grid(row=2, column=0, sticky="w")
        post_row = ttk.Frame(top, style="Panel.TFrame")
        post_row.grid(row=3, column=0, sticky="w", pady=(5, 0))
        ttk.Spinbox(post_row, from_=1, to=20, textvariable=self.post_count_var, width=6).pack(side="left")
        ttk.Button(post_row, text="Update boxes", command=self.rebuild_creative_boxes).pack(side="left", padx=(8, 0))
        top.columnconfigure(0, weight=1)

        self.creative = ttk.Frame(main, style="Panel.TFrame", padding=16)
        self.creative.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        self.creative.columnconfigure(1, weight=1)
        self.creative.columnconfigure(2, weight=1)
        self.rebuild_creative_boxes()

        buttons = ttk.Frame(main)
        buttons.grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 12))
        ttk.Button(buttons, text="Export Reel Design", command=self.export_reel_design).pack(side="left")

        self.result_box = tk.Text(main, height=13, wrap="word", font=("Segoe UI", 10), padx=10, pady=10)
        self.app.style_text_widget(self.result_box)
        self.result_box.grid(row=4, column=0, columnspan=2, sticky="nsew", pady=(0, 12))

    def rebuild_creative_boxes(self):
        for widget in self.creative.winfo_children():
            widget.destroy()
        self.creative_boxes = []

        post_count = max(1, min(20, int(self.post_count_var.get())))
        self.post_count_var.set(post_count)

        ttk.Label(self.creative, text="Creative reel design", style="CardTitle.TLabel").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        guideline = "Reel guideline: one striking visual, immediate branding, quick montage, then settle into the story."
        ttk.Label(self.creative, text=guideline, style="CardText.TLabel", wraplength=950).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 12))
        for col, header in enumerate(["Post", "First 2 seconds", "Opening text"]):
            ttk.Label(self.creative, text=header, style="CardText.TLabel").grid(row=2, column=col, sticky="w", padx=(0, 10))

        for post_number in range(1, post_count + 1):
            row = post_number + 2
            ttk.Label(self.creative, text=f"Post {post_number}", style="CardText.TLabel").grid(row=row, column=0, sticky="nw", padx=(0, 12), pady=(8, 0))
            first_seconds = ExpandingText(self.creative, width=38, min_height=1, max_height=5, wrap="word", font=("Segoe UI", 10), padx=8, pady=8)
            opening = ExpandingText(self.creative, width=54, min_height=1, max_height=5, wrap="word", font=("Segoe UI", 10), padx=8, pady=8)
            self.app.style_text_widget(first_seconds)
            self.app.style_text_widget(opening)
            first_seconds.grid(row=row, column=1, sticky="ew", padx=(0, 10), pady=(8, 0))
            opening.grid(row=row, column=2, sticky="ew", pady=(8, 0))
            self.creative_boxes.append((first_seconds, opening))

    def get_song_title(self):
        song_title = self.song_title_var.get().strip()
        if not song_title:
            messagebox.showerror("Error", "Enter a song title.")
            return None
        return song_title

    def build_creative_items(self, song_title):
        filename_song = display_song_filename(song_title)
        items = []
        for index, (first_seconds_box, opening_text_box) in enumerate(self.creative_boxes, start=1):
            items.append({
                "label": f"Post {index}",
                "filename": f"post {index} - {filename_song}.mp4",
                "first_seconds": first_seconds_box.get("1.0", tk.END).strip(),
                "opening_text": opening_text_box.get("1.0", tk.END).strip(),
            })
        return items

    def export_reel_design(self):
        song_title = self.get_song_title()
        if song_title is None:
            return
        filenames = get_output_filenames(song_title)
        design_txt_path = DOWNLOADS_PATH / filenames["design_txt_filename"]
        DOWNLOADS_PATH.mkdir(parents=True, exist_ok=True)
        design_txt_path.write_text(build_reel_design_text(song_title, self.build_creative_items(song_title)), encoding="utf-8")
        self.result_box.delete("1.0", tk.END)
        self.result_box.insert(tk.END, f"Created design TXT: {design_txt_path}\n\nReel design exported only.")
        messagebox.showinfo("Done", f"Reel design file created:\n\n{design_txt_path}")
        os.startfile(str(DOWNLOADS_PATH))


class CampaignPlannerPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.song_title_var = tk.StringVar()
        self.post_count_var = tk.IntVar(value=4)
        self.repost_count_var = tk.IntVar(value=4)
        self.start_date_var = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))
        self.default_intervals_var = tk.BooleanVar(value=True)
        self.post_offsets = list(FIRST_POST_OFFSETS)
        self.repost_offsets = list(REPOST_START_OFFSETS)
        self._build()

    def _build(self):
        scroller = ScrollFrame(self)
        scroller.pack(fill="both", expand=True)
        main = scroller.body
        main.columnconfigure(0, weight=1)

        top = ttk.Frame(main, style="Panel.TFrame", padding=16)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        top.columnconfigure(0, weight=1)

        ttk.Label(top, text="Song title", style="CardText.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.song_title_var, width=54).grid(row=1, column=0, sticky="ew", pady=(5, 12))

        counts = ttk.Frame(top, style="Panel.TFrame")
        counts.grid(row=2, column=0, sticky="w", pady=(0, 12))
        ttk.Label(counts, text="Posts", style="CardText.TLabel").pack(side="left")
        ttk.Spinbox(counts, from_=1, to=20, textvariable=self.post_count_var, width=6).pack(side="left", padx=(6, 16))
        ttk.Label(counts, text="Repost rounds", style="CardText.TLabel").pack(side="left")
        ttk.Spinbox(counts, from_=0, to=20, textvariable=self.repost_count_var, width=6).pack(side="left", padx=(6, 0))

        schedule_row = ttk.Frame(top, style="Panel.TFrame")
        schedule_row.grid(row=3, column=0, sticky="w", pady=(0, 12))
        ttk.Label(schedule_row, text="Start date", style="CardText.TLabel").pack(side="left")
        ttk.Entry(schedule_row, textvariable=self.start_date_var, width=14).pack(side="left", padx=(8, 8))
        ttk.Button(schedule_row, text="Calendar", command=self.open_start_calendar).pack(side="left", padx=(0, 14))
        ttk.Checkbutton(schedule_row, text="Default intervals", variable=self.default_intervals_var, command=self.on_default_intervals_changed).pack(side="left")
        ttk.Button(schedule_row, text="Edit intervals", command=self.open_interval_editor).pack(side="left", padx=(8, 0))

        ttk.Label(top, text="Song Description", style="CardText.TLabel").grid(row=4, column=0, sticky="w")
        self.song_description_box = tk.Text(top, height=4, wrap="word", font=("Segoe UI", 10), padx=8, pady=8)
        self.app.style_text_widget(self.song_description_box)
        self.song_description_box.grid(row=5, column=0, sticky="ew", pady=(5, 12))
        self.song_description_box.insert("1.0", "A song about...")

        ttk.Label(top, text="Hashtags", style="CardText.TLabel").grid(row=6, column=0, sticky="w")
        self.hashtags_box = tk.Text(top, height=5, wrap="word", font=("Segoe UI", 10), padx=8, pady=8)
        self.app.style_text_widget(self.hashtags_box)
        self.hashtags_box.grid(row=7, column=0, sticky="ew", pady=(5, 0))
        self.hashtags_box.insert("1.0", DEFAULT_HASHTAGS)

        buttons = ttk.Frame(main)
        buttons.grid(row=1, column=0, sticky="w", pady=(0, 12))
        ttk.Button(buttons, text="Generate Campaign", command=self.generate_campaign).pack(side="left")

        self.result_box = tk.Text(main, height=20, wrap="word", font=("Segoe UI", 10), padx=10, pady=10)
        self.app.style_text_widget(self.result_box)
        self.result_box.grid(row=2, column=0, sticky="nsew")

    def get_song_title(self):
        song_title = self.song_title_var.get().strip()
        if not song_title:
            messagebox.showerror("Error", "Enter a song title.")
            return None
        return song_title

    def selected_start_date(self):
        return datetime.strptime(self.start_date_var.get().strip(), "%Y-%m-%d").date()

    def open_start_calendar(self):
        try:
            current = self.selected_start_date()
        except Exception:
            current = date.today()

        popup = tk.Toplevel(self)
        popup.title("Choose campaign start")
        popup.transient(self)
        popup.resizable(False, False)

        year_var = tk.IntVar(value=current.year)
        month_var = tk.IntVar(value=current.month)
        shell = ttk.Frame(popup, padding=12)
        shell.pack(fill="both", expand=True)
        header = ttk.Frame(shell)
        header.pack(fill="x", pady=(0, 10))

        def redraw():
            for widget in grid.winfo_children():
                widget.destroy()
            title.configure(text=f"{calendar.month_name[month_var.get()]} {year_var.get()}")
            for col, label in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
                ttk.Label(grid, text=label, font=("Segoe UI", 9, "bold")).grid(row=0, column=col, padx=3, pady=3)
            cal = calendar.Calendar(firstweekday=0)
            for row_idx, week in enumerate(cal.monthdatescalendar(year_var.get(), month_var.get()), start=1):
                for col_idx, day in enumerate(week):
                    if day.month != month_var.get():
                        ttk.Label(grid, text="").grid(row=row_idx, column=col_idx, padx=3, pady=3)
                        continue
                    ttk.Button(grid, text=str(day.day), width=4, command=lambda d=day: choose_day(d)).grid(row=row_idx, column=col_idx, padx=3, pady=3)

        def change_month(delta):
            month = month_var.get() + delta
            year = year_var.get()
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            month_var.set(month)
            year_var.set(year)
            redraw()

        def choose_day(day):
            self.start_date_var.set(day.strftime("%Y-%m-%d"))
            popup.destroy()

        ttk.Button(header, text="<", width=3, command=lambda: change_month(-1)).pack(side="left")
        title = ttk.Label(header, text="", font=("Segoe UI", 11, "bold"))
        title.pack(side="left", expand=True)
        ttk.Button(header, text=">", width=3, command=lambda: change_month(1)).pack(side="right")
        grid = ttk.Frame(shell)
        grid.pack()
        redraw()

    def on_default_intervals_changed(self):
        if not self.default_intervals_var.get():
            self.open_interval_editor()

    def open_interval_editor(self):
        post_count = max(1, min(20, int(self.post_count_var.get())))
        repost_count = max(0, min(20, int(self.repost_count_var.get())))
        while len(self.post_offsets) < post_count:
            self.post_offsets.append(self.post_offsets[-1] + 7)
        while len(self.repost_offsets) < repost_count:
            self.repost_offsets.append(self.repost_offsets[-1] + 35 if self.repost_offsets else 21)

        window = tk.Toplevel(self)
        window.title("Campaign intervals")
        window.geometry("760x560")
        window.transient(self)

        shell = ttk.Frame(window, style="Panel.TFrame", padding=14)
        shell.pack(fill="both", expand=True)
        ttk.Label(shell, text="Intervals in days from start date", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 10))

        body = ttk.Frame(shell, style="Panel.TFrame")
        body.pack(fill="both", expand=True)
        body.columnconfigure(2, weight=1)
        body.columnconfigure(6, weight=1)
        post_vars = []
        repost_vars = []

        ttk.Label(body, text="Posts", style="CardText.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 6))
        for index in range(post_count):
            ttk.Label(body, text=f"Post {index + 1}", style="CardText.TLabel").grid(row=index + 1, column=0, sticky="w", pady=3)
            var = tk.IntVar(value=self.post_offsets[index])
            ttk.Spinbox(body, from_=0, to=365, textvariable=var, width=7).grid(row=index + 1, column=1, sticky="w", pady=3)
            scale = ttk.Scale(body, from_=0, to=365, orient="horizontal", variable=var, length=260, command=lambda value, v=var: v.set(int(float(value))))
            scale.grid(row=index + 1, column=2, sticky="ew", pady=3, padx=(8, 0))
            post_vars.append(var)

        col_offset = 4
        ttk.Label(body, text="Repost rounds", style="CardText.TLabel").grid(row=0, column=col_offset, sticky="w", pady=(0, 6), padx=(22, 0))
        for index in range(repost_count):
            ttk.Label(body, text=f"Repost {index + 1}", style="CardText.TLabel").grid(row=index + 1, column=col_offset, sticky="w", padx=(22, 0), pady=3)
            var = tk.IntVar(value=self.repost_offsets[index])
            ttk.Spinbox(body, from_=0, to=730, textvariable=var, width=7).grid(row=index + 1, column=col_offset + 1, sticky="w", pady=3)
            scale = ttk.Scale(body, from_=0, to=730, orient="horizontal", variable=var, length=260, command=lambda value, v=var: v.set(int(float(value))))
            scale.grid(row=index + 1, column=col_offset + 2, sticky="ew", pady=3, padx=(8, 0))
            repost_vars.append(var)

        buttons = ttk.Frame(shell, style="Panel.TFrame")
        buttons.pack(fill="x", pady=(12, 0))

        def save_intervals():
            self.post_offsets = [int(var.get()) for var in post_vars]
            self.repost_offsets = [int(var.get()) for var in repost_vars]
            window.destroy()

        def reset_defaults():
            self.post_offsets = list(FIRST_POST_OFFSETS)
            self.repost_offsets = list(REPOST_START_OFFSETS)
            self.default_intervals_var.set(True)
            window.destroy()

        ttk.Button(buttons, text="OK", command=save_intervals, style="Accent.TButton").pack(side="left")
        ttk.Button(buttons, text="Use defaults", command=reset_defaults).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="Cancel", command=window.destroy).pack(side="left", padx=(8, 0))

    def generate_campaign(self):
        song_title = self.get_song_title()
        if song_title is None:
            return
        song_description = self.song_description_box.get("1.0", tk.END).strip()
        hashtags = self.hashtags_box.get("1.0", tk.END).strip()
        post_count = max(1, min(20, int(self.post_count_var.get())))
        repost_count = max(0, min(20, int(self.repost_count_var.get())))
        self.post_count_var.set(post_count)
        self.repost_count_var.set(repost_count)
        try:
            start_day = self.selected_start_date()
        except Exception:
            messagebox.showerror("Invalid start date", "Use start date as YYYY-MM-DD.")
            return
        post_offsets = None if self.default_intervals_var.get() else self.post_offsets[:post_count]
        repost_offsets = None if self.default_intervals_var.get() else self.repost_offsets[:repost_count]
        filenames = get_output_filenames(song_title)
        ics_path = DOWNLOADS_PATH / filenames["ics_filename"]
        schedule_txt_path = DOWNLOADS_PATH / filenames["schedule_txt_filename"]
        campaign_items = build_campaign_items(song_title, post_count, repost_count, start_day, post_offsets, repost_offsets)
        events = build_events(song_title, filenames["schedule_txt_filename"], filenames["design_txt_filename"], len(campaign_items), start_day)
        DOWNLOADS_PATH.mkdir(parents=True, exist_ok=True)
        ics_path.write_text(create_ics(events), encoding="utf-8")
        schedule_txt_path.write_text(build_post_schedule_text(song_title, song_description, campaign_items, hashtags), encoding="utf-8")
        lines = [f"Created ICS: {ics_path}", f"Created schedule TXT: {schedule_txt_path}", "", "Campaign generated.", "", "Calendar events:", ""]
        lines.extend(f"{event['datetime'].strftime('%Y-%m-%d %H:%M')} - {event['title']}" for event in events)
        lines.extend(["", "Post schedule:", ""])
        lines.extend(f"{item['date'].strftime('%d/%m/%Y')} 20:00 - {item['label']} - {item['filename']}" for item in campaign_items)
        self.result_box.delete("1.0", tk.END)
        self.result_box.insert(tk.END, "\n".join(lines))
        messagebox.showinfo("Done", f"Campaign files created:\n\n{ics_path}\n{schedule_txt_path}")


class MetricsPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.entries = {}
        self.rows = []
        self.totals = []
        self.monthly_rows = []
        self._build()

    def _build(self):
        scroller = ScrollFrame(self)
        scroller.pack(fill="both", expand=True)
        scroller.canvas.configure(bg=self.app.colors["app_bg"], highlightthickness=0)
        scroller.body.configure(style="TFrame")
        scroller.set_auto_scrollbar(True)

        main = ttk.Frame(scroller.body, style="TFrame")
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, weight=1)

        toolbar = ttk.Frame(main, style="Panel.TFrame", padding=14)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        ttk.Label(toolbar, text="Weekly reel metrics", style="CardTitle.TLabel").pack(side="left")
        ttk.Button(toolbar, text="Refresh", command=self.refresh).pack(side="right")
        ttk.Button(toolbar, text="Open Excel", command=lambda: os.startfile(str(get_metrics_workbook_path()))).pack(side="right", padx=(0, 8))

        self.summary = ttk.Frame(main, style="TFrame")
        self.summary.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        self.metrics_body = ttk.Frame(main, style="TFrame")
        self.metrics_body.grid(row=2, column=0, sticky="nsew")
        self.metrics_body.columnconfigure(0, weight=0)
        self.metrics_body.columnconfigure(1, weight=1)
        self.metrics_body.rowconfigure(0, weight=1)
        self.metrics_body.bind("<Configure>", self.layout_metrics_body)

        left = ttk.Frame(self.metrics_body, style="TFrame")
        self.metrics_left = left
        self.dashboard = ttk.Frame(self.metrics_body, style="TFrame")
        self.dashboard.bind("<Configure>", self.layout_dashboard)

        form = ttk.Frame(left, style="Panel.TFrame", padding=14)
        form.pack(fill="x", pady=(0, 12))
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)
        ttk.Label(form, text="Add week", style="CardTitle.TLabel").grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 10))

        platform_fields = [
            ("TikTok", "tt_views", "tt_followers"),
            ("Instagram", "ig_views", "ig_followers"),
            ("YouTube", "yt_views", "yt_followers"),
            ("Facebook", "fb_views", "fb_followers"),
        ]

        for row, (platform, views_key, followers_key) in enumerate(platform_fields, start=1):
            ttk.Label(form, text=f"{platform} views", style="CardText.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=4)
            views_var = tk.StringVar()
            ttk.Entry(form, textvariable=views_var, width=12).grid(row=row, column=1, sticky="ew", pady=4)
            self.entries[views_key] = views_var

            ttk.Label(form, text=f"{platform} followers", style="CardText.TLabel").grid(row=row, column=2, sticky="w", padx=(14, 8), pady=4)
            followers_var = tk.StringVar()
            ttk.Entry(form, textvariable=followers_var, width=12).grid(row=row, column=3, sticky="ew", pady=4)
            self.entries[followers_key] = followers_var

        date_row = ttk.Frame(form, style="Panel.TFrame")
        date_row.grid(row=5, column=0, columnspan=4, sticky="ew", pady=(8, 0))
        self.entries["week"] = tk.StringVar()
        self.entries["date"] = tk.StringVar()
        self.period_var = tk.StringVar(value="")
        ttk.Label(date_row, text="Week", style="CardText.TLabel").pack(side="left")
        ttk.Entry(date_row, textvariable=self.entries["week"], width=10).pack(side="left", padx=(8, 18))
        ttk.Label(date_row, text="Update Sunday", style="CardText.TLabel").pack(side="left")
        ttk.Entry(date_row, textvariable=self.entries["date"], width=14).pack(side="left", padx=(8, 8))
        ttk.Button(date_row, text="Calendar", command=self.open_week_calendar).pack(side="left")
        ttk.Label(form, textvariable=self.period_var, style="CardText.TLabel").grid(row=6, column=0, columnspan=4, sticky="w", pady=(6, 0))

        ttk.Button(form, text="Save week", command=self.save_week, style="Accent.TButton").grid(row=7, column=0, columnspan=4, sticky="ew", pady=(12, 0))

        views_table_frame = ttk.Frame(left, style="Panel.TFrame", padding=8)
        views_table_frame.pack(fill="both", expand=True)
        self.table = tk.Text(views_table_frame, height=7, wrap="none", font=("Consolas", 9), padx=8, pady=8)
        self.app.style_text_widget(self.table)
        views_scrollbar = ttk.Scrollbar(views_table_frame, orient="vertical", command=self.table.yview)
        self.table.configure(yscrollcommand=views_scrollbar.set)
        self.table.pack(side="left", fill="both", expand=True)
        views_scrollbar.pack(side="right", fill="y")

        followers_table_frame = ttk.Frame(left, style="Panel.TFrame", padding=8)
        followers_table_frame.pack(fill="both", expand=True, pady=(12, 0))
        self.followers_table = tk.Text(followers_table_frame, height=7, wrap="none", font=("Consolas", 9), padx=8, pady=8)
        self.app.style_text_widget(self.followers_table)
        followers_scrollbar = ttk.Scrollbar(followers_table_frame, orient="vertical", command=self.followers_table.yview)
        self.followers_table.configure(yscrollcommand=followers_scrollbar.set)
        self.followers_table.pack(side="left", fill="both", expand=True)
        followers_scrollbar.pack(side="right", fill="y")

        self.views_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.views_canvas)
        self.followers_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.followers_canvas)
        self.platform_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.platform_canvas)
        observations = ttk.Frame(self.dashboard, style="Panel.TFrame", padding=12)
        self.weekly_observations_panel = observations
        ttk.Label(observations, text="Weekly observations", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 8))
        self.weekly_observations_text = tk.Text(observations, height=7, wrap="word", font=("Segoe UI", 10), padx=10, pady=10)
        self.app.style_text_widget(self.weekly_observations_text)
        self.weekly_observations_text.bind("<Key>", lambda _event: "break")
        self.weekly_observations_text.pack(fill="both", expand=True)
        self.monthly_impressions_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.monthly_impressions_canvas)
        self.monthly_followers_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.monthly_followers_canvas)
        self.monthly_posts_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.monthly_posts_canvas)
        self.monthly_interactions_canvas = tk.Canvas(self.dashboard, height=230, bg=self.app.colors["panel_bg"], highlightthickness=0)
        self.app.style_chart_canvas(self.monthly_interactions_canvas)

        recommendations = ttk.Frame(self.dashboard, style="Panel.TFrame", padding=12)
        self.recommendations_panel = recommendations
        ttk.Label(recommendations, text="Monthly summary", style="CardTitle.TLabel").pack(anchor="w", pady=(0, 8))
        rec_body = ttk.Frame(recommendations, style="Panel.TFrame")
        rec_body.pack(fill="both", expand=True)
        self.recommendations_text = tk.Text(rec_body, height=8, wrap="word", font=("Segoe UI", 10), padx=10, pady=10)
        self.app.style_text_widget(self.recommendations_text)
        self.recommendations_text.bind("<Key>", lambda _event: "break")
        self.recommendations_text.bind("<Button-1>", self.handle_recommendation_click)
        rec_scrollbar = ttk.Scrollbar(rec_body, orient="vertical", command=self.recommendations_text.yview)
        self.recommendations_text.configure(yscrollcommand=rec_scrollbar.set)
        self.recommendations_text.pack(side="left", fill="both", expand=True)
        rec_scrollbar.pack(side="right", fill="y")

        self.layout_metrics_body()
        self.layout_dashboard()
        self.refresh()

    def layout_metrics_body(self, _event=None):
        if not hasattr(self, "metrics_left"):
            return
        width = max(1, self.winfo_width())
        for widget in (self.metrics_left, self.dashboard):
            widget.grid_forget()
        if width < 1180:
            self.metrics_body.columnconfigure(0, weight=1)
            self.metrics_body.columnconfigure(1, weight=0)
            self.metrics_left.grid(row=0, column=0, sticky="ew", pady=(0, 12))
            self.dashboard.grid(row=1, column=0, sticky="nsew")
        else:
            self.metrics_body.columnconfigure(0, weight=0, minsize=430)
            self.metrics_body.columnconfigure(1, weight=1)
            self.metrics_left.grid(row=0, column=0, sticky="nsw", padx=(0, 12))
            self.dashboard.grid(row=0, column=1, sticky="nsew")

    def layout_dashboard(self, _event=None):
        if not hasattr(self, "recommendations_panel"):
            return
        widgets = [
            self.views_canvas,
            self.followers_canvas,
            self.platform_canvas,
            self.weekly_observations_panel,
            self.monthly_impressions_canvas,
            self.monthly_followers_canvas,
            self.monthly_posts_canvas,
            self.monthly_interactions_canvas,
            self.recommendations_panel,
        ]
        for widget in widgets:
            widget.grid_forget()
        width = max(1, self.dashboard.winfo_width())
        for index in range(len(widgets)):
            self.dashboard.rowconfigure(index, weight=0)
            self.dashboard.columnconfigure(index, weight=0)
        if width < 720:
            self.dashboard.columnconfigure(0, weight=1)
            for row in range(len(widgets)):
                self.dashboard.rowconfigure(row, weight=1)
            for row, widget in enumerate(widgets):
                widget.grid(row=row, column=0, sticky="nsew", pady=(0, 10))
        else:
            self.dashboard.columnconfigure(0, weight=1)
            self.dashboard.columnconfigure(1, weight=1)
            for row in range(4):
                self.dashboard.rowconfigure(row, weight=1)
            for index, widget in enumerate(widgets):
                row = index // 2
                col = index % 2
                widget.grid(row=row, column=col, sticky="nsew", padx=(0 if col == 0 else 6, 0 if col == 1 else 6), pady=(0, 10))

    def growth_text(self, current, previous):
        try:
            current = float(current)
            previous = float(previous)
        except Exception:
            return "no previous week"
        if previous <= 0:
            return "new baseline" if current > 0 else "no change"
        growth = (current - previous) * 100 / previous
        sign = "+" if growth >= 0 else ""
        if 0 < abs(growth) < 1:
            growth_text = f"{sign}{growth:.1f}%"
        else:
            growth_text = f"{sign}{growth:.0f}%"
        return f"{growth_text} vs last week"

    def platform_mvp(self):
        if not self.rows:
            return ("Platform MVP", "n/a", "no data yet")
        latest = self.rows[-1]
        platforms = [
            ("TikTok", latest["tt_views"]),
            ("Instagram", latest["ig_views"]),
            ("YouTube", latest["yt_views"]),
            ("Facebook", latest["fb_views"]),
        ]
        name, value = max(platforms, key=lambda item: item[1])
        total = sum(value for _name, value in platforms)
        share = 0 if total <= 0 else value * 100 / total
        return ("Platform MVP", name, f"{value:,} impressions | {share:.0f}% of weekly traffic")

    def monthly_kpis(self):
        if not self.monthly_rows:
            return []
        latest = self.monthly_rows[-1]
        previous = self.monthly_rows[-2] if len(self.monthly_rows) >= 2 else None
        platform_values = [
            ("Facebook", latest["fb_impressions"]),
            ("Instagram", latest["ig_impressions"]),
            ("TikTok", latest["tt_impressions"]),
            ("YouTube", latest["yt_impressions"]),
        ]
        best_name, best_value = max(platform_values, key=lambda item: item[1])
        total_impressions = sum(value for _name, value in platform_values)
        total_followers = latest["fb_followers"] + latest["ig_followers"] + latest["tt_followers"] + latest["yt_followers"]
        growth_label = "no previous month"
        if previous is not None:
            growths = []
            for prefix, name in (("fb", "Facebook"), ("ig", "Instagram"), ("tt", "TikTok"), ("yt", "YouTube")):
                old = previous[f"{prefix}_impressions"]
                new = latest[f"{prefix}_impressions"]
                if old > 0:
                    growths.append(((new - old) * 100 / old, name))
            if growths:
                growth, name = max(growths, key=lambda item: item[0])
                growth_label = f"{name} {growth:+.0f}%"
        return [
            ("Monthly impressions", total_impressions, latest["month"]),
            ("Monthly followers", total_followers, latest["month"]),
            ("Best platform", best_name, f"{best_value:,} impressions"),
            ("Largest MoM growth", growth_label, "impressions"),
        ]

    def refresh(self):
        for widget in self.summary.winfo_children():
            widget.destroy()
        try:
            self.rows = load_metrics_rows()
            self.totals = metrics_totals(self.rows)
            self.monthly_rows = load_monthly_rows()
        except Exception as exc:
            ttk.Label(self.summary, text=str(exc), style="Muted.TLabel").pack(anchor="w")
            return

        self.prefill_next_week()
        latest = self.totals[-1] if self.totals else None
        previous = self.totals[-2] if len(self.totals) >= 2 else None
        latest_week_views = sum((self.rows[-1][f"{p}_views"] for p in ("tt", "ig", "yt", "fb")), 0) if self.rows else 0
        previous_week_views = sum((self.rows[-2][f"{p}_views"] for p in ("tt", "ig", "yt", "fb")), 0) if len(self.rows) >= 2 else 0
        _mvp_title, mvp_name, mvp_detail = self.platform_mvp()
        cards = [
            ("Weeks", len(self.rows), "tracked entries"),
            ("Weekly impressions", latest_week_views, self.growth_text(latest_week_views, previous_week_views)),
            ("Followers", latest["total_followers"] if latest else 0, self.growth_text(latest["total_followers"] if latest else 0, previous["total_followers"] if previous else 0)),
            ("Platform MVP", mvp_name, mvp_detail),
        ]
        cards.extend(self.monthly_kpis())
        for col in range(len(cards)):
            self.summary.columnconfigure(col % 4, weight=1, uniform="summary")
        for index, (title, value, detail) in enumerate(cards):
            card = ttk.Frame(self.summary, style="Panel.TFrame", padding=12)
            card.grid(row=index // 4, column=index % 4, sticky="ew", padx=6, pady=(0, 8))
            ttk.Label(card, text=title, style="CardText.TLabel").pack(anchor="w")
            value_text = f"{value:,}" if isinstance(value, (int, float)) else str(value)
            ttk.Label(card, text=value_text, font=("Segoe UI", 18, "bold"), background=self.app.colors["panel_bg"], foreground=self.app.colors["text"]).pack(anchor="w")
            ttk.Label(card, text=detail, style="CardText.TLabel").pack(anchor="w")

        self.draw_table()
        self.draw_followers_table()
        self.draw_views_chart()
        self.draw_followers_chart()
        self.draw_platform_chart()
        self.load_weekly_observations()
        self.draw_monthly_charts()
        self.load_monthly_recommendations()

    def load_weekly_observations(self):
        if not hasattr(self, "weekly_observations_text"):
            return
        lines = []
        if not self.rows:
            lines = ["waiting for weekly data..."]
        else:
            latest = self.rows[-1]
            previous = self.rows[-2] if len(self.rows) >= 2 else None
            platforms = [
                ("TikTok", latest["tt_views"], previous["tt_views"] if previous else None),
                ("Instagram", latest["ig_views"], previous["ig_views"] if previous else None),
                ("YouTube", latest["yt_views"], previous["yt_views"] if previous else None),
                ("Facebook", latest["fb_views"], previous["fb_views"] if previous else None),
            ]
            best_name, best_value, _previous_value = max(platforms, key=lambda item: item[1])
            total = sum(value for _name, value, _old in platforms)
            lines.append(f"Best platform this week: {best_name} with {best_value:,} impressions.")
            if previous is not None:
                previous_total = sum(previous[f"{prefix}_views"] for prefix in ("tt", "ig", "yt", "fb"))
                lines.append(f"Weekly movement: {self.growth_text(total, previous_total)}.")
                movers = []
                for name, value, old_value in platforms:
                    if old_value is None:
                        continue
                    movers.append((value - old_value, name))
                if movers:
                    change, name = max(movers, key=lambda item: abs(item[0]))
                    direction = "up" if change >= 0 else "down"
                    lines.append(f"Biggest platform movement: {name} {direction} {abs(change):,} impressions.")
            else:
                lines.append("Add one more week to unlock week-over-week observations.")
        self.weekly_observations_text.configure(state="normal")
        self.weekly_observations_text.delete("1.0", tk.END)
        self.weekly_observations_text.insert(tk.END, "\n\n".join(lines))
        self.weekly_observations_text.configure(state="disabled")

    def load_monthly_recommendations(self):
        if not hasattr(self, "recommendations_text"):
            return
        if self.monthly_rows:
            latest = self.monthly_rows[-1]
            content = (
                f"Month: {latest['month']}\n\n"
                f"Dashboard Summary\n{latest.get('dashboard_summary') or 'waiting for info...'}\n\n"
                f"Next Month Focus\n{latest.get('next_month_focus') or 'waiting for info...'}"
            )
        else:
            content = "waiting for monthly info..."
        self.recommendations_text.configure(state="normal")
        self.recommendations_text.delete("1.0", tk.END)
        self.recommendations_text.insert(tk.END, content)
        self.recommendations_text.configure(state="disabled")

    def handle_recommendation_click(self, event):
        try:
            line_number = int(self.recommendations_text.index(f"@{event.x},{event.y}").split(".")[0])
            report_path = getattr(self.recommendations_text, "report_paths_by_line", {}).get(line_number)
            if report_path is not None:
                self.open_monthly_report(report_path)
                return "break"
        except Exception:
            pass
        return None

    def open_monthly_report(self, path):
        try:
            if path.exists():
                os.startfile(str(path))
            else:
                messagebox.showinfo("Report not found", f"I could not find this report:\n\n{path}")
        except Exception as exc:
            messagebox.showerror("Could not open report", str(exc))

    def prefill_next_week(self):
        if not self.rows:
            return
        latest = self.rows[-1]
        try:
            self.entries["week"].set(str(int(latest["week"]) + 1))
        except Exception:
            self.entries["week"].set("")
        latest_date = latest["date"]
        if isinstance(latest_date, datetime):
            next_date = latest_date.date() + timedelta(days=7)
        elif isinstance(latest_date, date):
            next_date = latest_date + timedelta(days=7)
        else:
            next_date = date.today()
        self.entries["date"].set(next_date.strftime("%Y-%m-%d"))
        self.update_period_label()
        for key in ("tt_views", "ig_views", "yt_views", "fb_views"):
            self.entries[key].set("0")
        for key in ("tt_followers", "ig_followers", "yt_followers", "fb_followers"):
            self.entries[key].set(str(latest.get(key, 0)))

    def selected_update_date(self):
        return datetime.strptime(self.entries["date"].get().strip(), "%Y-%m-%d").date()

    def update_period_label(self):
        try:
            update_day = self.selected_update_date()
            start_day = update_day - timedelta(days=6)
            self.period_var.set(f"Period: {start_day.strftime('%d-%m')} to {update_day.strftime('%d-%m')}  |  Sunday 20:00 update")
        except Exception:
            self.period_var.set("Period: choose a Sunday update date")

    def open_week_calendar(self):
        try:
            current = self.selected_update_date()
        except Exception:
            current = date.today()

        popup = tk.Toplevel(self)
        popup.title("Choose update Sunday")
        popup.transient(self)
        popup.resizable(False, False)

        year_var = tk.IntVar(value=current.year)
        month_var = tk.IntVar(value=current.month)
        shell = ttk.Frame(popup, padding=12)
        shell.pack(fill="both", expand=True)
        header = ttk.Frame(shell)
        header.pack(fill="x", pady=(0, 10))

        def redraw():
            for widget in grid.winfo_children():
                widget.destroy()
            month_name = calendar.month_name[month_var.get()]
            title.configure(text=f"{month_name} {year_var.get()}")
            for col, label in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
                ttk.Label(grid, text=label, font=("Segoe UI", 9, "bold")).grid(row=0, column=col, padx=3, pady=3)
            cal = calendar.Calendar(firstweekday=0)
            for row_idx, week in enumerate(cal.monthdatescalendar(year_var.get(), month_var.get()), start=1):
                for col_idx, day in enumerate(week):
                    if day.month != month_var.get():
                        ttk.Label(grid, text="").grid(row=row_idx, column=col_idx, padx=3, pady=3)
                        continue
                    state = "normal" if day.weekday() == 6 else "disabled"
                    button = ttk.Button(grid, text=str(day.day), width=4, state=state, command=lambda d=day: choose_day(d))
                    button.grid(row=row_idx, column=col_idx, padx=3, pady=3)

        def change_month(delta):
            month = month_var.get() + delta
            year = year_var.get()
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            month_var.set(month)
            year_var.set(year)
            redraw()

        def choose_day(day):
            self.entries["week"].set(str(day.isocalendar().week))
            self.entries["date"].set(day.strftime("%Y-%m-%d"))
            self.update_period_label()
            popup.destroy()

        ttk.Button(header, text="<", width=3, command=lambda: change_month(-1)).pack(side="left")
        title = ttk.Label(header, text="", font=("Segoe UI", 11, "bold"))
        title.pack(side="left", expand=True)
        ttk.Button(header, text=">", width=3, command=lambda: change_month(1)).pack(side="right")
        grid = ttk.Frame(shell)
        grid.pack()
        ttk.Label(shell, text="Only Sundays are selectable.", style="Muted.TLabel").pack(anchor="w", pady=(10, 0))
        redraw()

    def parse_int(self, key):
        text = self.entries[key].get().strip()
        return int(text) if text else 0

    def save_week(self):
        try:
            self.update_period_label()
            values = {
                "week": self.parse_int("week"),
                "date": datetime.combine(self.selected_update_date(), time(20, 0)),
                "tt_views": self.parse_int("tt_views"),
                "tt_followers": self.parse_int("tt_followers"),
                "ig_views": self.parse_int("ig_views"),
                "ig_followers": self.parse_int("ig_followers"),
                "yt_views": self.parse_int("yt_views"),
                "yt_followers": self.parse_int("yt_followers"),
                "fb_views": self.parse_int("fb_views"),
                "fb_followers": self.parse_int("fb_followers"),
            }
            append_metrics_row(values)
            self.refresh()
            messagebox.showinfo("Saved", "Weekly metrics saved to Excel.")
        except Exception as exc:
            messagebox.showerror("Could not save", str(exc))

    def draw_table(self):
        self.table.delete("1.0", tk.END)
        self.table.insert(tk.END, "Impressions\n")
        self.table.insert(tk.END, "Week   Date        TT     IG     YT     FB     Total\n")
        self.table.insert(tk.END, "-------------------------------------------------------\n")
        for row, total in reversed(list(zip(self.rows, self.totals))):
            row_date = row["date"].strftime("%Y-%m-%d") if hasattr(row["date"], "strftime") else str(row["date"])
            line = f"{str(row['week']):<6} {row_date:<10} {row['tt_views']:>6} {row['ig_views']:>6} {row['yt_views']:>6} {row['fb_views']:>6} {total['total_views']:>9}\n"
            self.table.insert(tk.END, line)

    def draw_followers_table(self):
        self.followers_table.delete("1.0", tk.END)
        self.followers_table.insert(tk.END, "Followers\n")
        self.followers_table.insert(tk.END, "Week   Date        TT     IG     YT     FB     Total\n")
        self.followers_table.insert(tk.END, "-------------------------------------------------------\n")
        for row, total in reversed(list(zip(self.rows, self.totals))):
            row_date = row["date"].strftime("%Y-%m-%d") if hasattr(row["date"], "strftime") else str(row["date"])
            line = f"{str(row['week']):<6} {row_date:<10} {row['tt_followers']:>6} {row['ig_followers']:>6} {row['yt_followers']:>6} {row['fb_followers']:>6} {total['total_followers']:>9}\n"
            self.followers_table.insert(tk.END, line)

    def draw_total_line_chart(self, canvas, title, value_key, color):
        canvas.delete("all")
        canvas.update_idletasks()
        width = max(260, canvas.winfo_width())
        height = max(190, canvas.winfo_height())
        rounded_rect(canvas, 0, 0, width, height, radius=10, fill=self.app.colors["chart_bg"], outline=self.app.colors["border"])
        canvas.create_text(16, 14, text=title, anchor="nw", font=("Segoe UI", 12, "bold"), fill=self.app.colors["text"])
        data = self.totals[-16:]
        if len(data) < 2:
            return
        values = [row[value_key] for row in data]
        max_value = max(values) or 1
        y_max = max_value
        left, top, right, bottom = 50, 45, width - 18, height - 34
        for step in range(0, 5):
            value = y_max * step / 4
            y = bottom - (bottom - top) * step / 4
            canvas.create_line(left - 4, y, right, y, fill=self.app.colors["grid"])
            canvas.create_text(left - 8, y, text=f"{int(value):,}", anchor="e", font=("Segoe UI", 8), fill=self.app.colors["muted"])
        points = []
        for idx, value in enumerate(values):
            x = left + (right - left) * idx / (len(values) - 1)
            y = bottom - (bottom - top) * value / max_value
            points.extend([x, y])
        canvas.create_line(left, bottom, right, bottom, fill=self.app.colors["border"])
        canvas.create_line(left, top, left, bottom, fill=self.app.colors["border"])
        canvas.create_line(*points, fill=color, width=3, smooth=True)
        previous_month = None
        for idx, row in enumerate(data):
            label = metric_month_label(row["date"])
            show_label = idx == 0 or idx == len(data) - 1 or (label and label != previous_month)
            if show_label:
                x = left + (right - left) * idx / (len(data) - 1)
                canvas.create_text(x, bottom + 14, text=label or str(row["week"]), font=("Segoe UI", 8), fill=self.app.colors["muted"])
            if label:
                previous_month = label

    def draw_views_chart(self):
        self.draw_total_line_chart(self.views_canvas, "Weekly impressions over time", "total_views", "#2f6f73")

    def draw_followers_chart(self):
        self.draw_total_line_chart(self.followers_canvas, "Total followers over time", "total_followers", "#7b5f2a")

    def draw_platform_chart(self):
        canvas = self.platform_canvas
        canvas.delete("all")
        canvas.update_idletasks()
        width = max(260, canvas.winfo_width())
        height = max(190, canvas.winfo_height())
        rounded_rect(canvas, 0, 0, width, height, radius=10, fill=self.app.colors["chart_bg"], outline=self.app.colors["border"])
        canvas.create_text(16, 14, text="Latest week impressions by platform", anchor="nw", font=("Segoe UI", 12, "bold"), fill=self.app.colors["text"])
        if not self.rows:
            return
        latest = self.rows[-1]
        labels = ["TikTok", "Instagram", "YouTube", "Facebook"]
        values = [latest["tt_views"], latest["ig_views"], latest["yt_views"], latest["fb_views"]]
        colors = ["#111111", "#7b5f2a", "#4f6f8f", "#8d3f3f"]
        max_value = max(values) or 1
        left, top, right, bar_h, gap = 86, 50, width - 24, 22, 14
        for idx, (label, value, color) in enumerate(zip(labels, values, colors)):
            y = top + idx * (bar_h + gap)
            canvas.create_text(16, y + bar_h / 2, text=label, anchor="w", font=("Segoe UI", 9), fill=self.app.colors["text"])
            rounded_rect(canvas, left, y, right, y + bar_h, radius=10, fill=self.app.colors["grid"], outline="")
            bar_w = (right - left) * value / max_value
            rounded_rect(canvas, left, y, left + bar_w, y + bar_h, radius=10, fill=color, outline="")
            canvas.create_text(left + bar_w + 8, y + bar_h / 2, text=f"{value:,}", anchor="w", font=("Segoe UI", 9, "bold"), fill=self.app.colors["text"])

    def draw_monthly_charts(self):
        self.draw_monthly_line_chart(
            self.monthly_impressions_canvas,
            "Monthly impressions",
            ["fb_impressions", "ig_impressions", "tt_impressions", "yt_impressions"],
        )
        self.draw_monthly_line_chart(
            self.monthly_followers_canvas,
            "Monthly followers",
            ["fb_followers", "ig_followers", "tt_followers", "yt_followers"],
        )
        self.draw_monthly_line_chart(
            self.monthly_posts_canvas,
            "Monthly posts",
            ["fb_posts", "ig_posts", "tt_posts", "yt_posts"],
        )
        self.draw_monthly_line_chart(
            self.monthly_interactions_canvas,
            "Monthly interactions",
            ["fb_interactions", "ig_interactions", "tt_interactions", "yt_interactions"],
        )

    def draw_monthly_line_chart(self, canvas, title, keys):
        canvas.delete("all")
        canvas.update_idletasks()
        width = max(260, canvas.winfo_width())
        height = max(190, canvas.winfo_height())
        rounded_rect(canvas, 0, 0, width, height, radius=10, fill=self.app.colors["chart_bg"], outline=self.app.colors["border"])
        canvas.create_text(16, 14, text=title, anchor="nw", font=("Segoe UI", 12, "bold"), fill=self.app.colors["text"])
        data = self.monthly_rows[-12:]
        if not data:
            canvas.create_text(16, 52, text="waiting for monthly data...", anchor="nw", font=("Segoe UI", 9), fill=self.app.colors["muted"])
            return
        platform_labels = ["Facebook", "Instagram", "TikTok", "YouTube"]
        colors = ["#8d3f3f", "#7b5f2a", "#111111", "#4f6f8f"]
        max_value = max(max(row[key] for key in keys) for row in data) or 1
        left, top, right, bottom = 54, 48, width - 18, height - 38
        for step in range(0, 5):
            value = max_value * step / 4
            y = bottom - (bottom - top) * step / 4
            canvas.create_line(left - 4, y, right, y, fill=self.app.colors["grid"])
            canvas.create_text(left - 8, y, text=f"{int(value):,}", anchor="e", font=("Segoe UI", 8), fill=self.app.colors["muted"])
        if len(data) == 1:
            row = data[0]
            bar_gap = 10
            usable_height = max(40, bottom - top - bar_gap * 3)
            bar_height = usable_height / 4
            label_width = 76
            bar_left = left + label_width
            for idx, (key, label, color) in enumerate(zip(keys, platform_labels, colors)):
                y1 = top + idx * (bar_height + bar_gap)
                y2 = y1 + bar_height
                value = row[key]
                bar_right = bar_left + (right - bar_left - 44) * value / max_value
                canvas.create_text(left, (y1 + y2) / 2, text=label, anchor="w", font=("Segoe UI", 8), fill=self.app.colors["text"])
                rounded_rect(canvas, bar_left, y1, right - 44, y2, radius=6, fill=self.app.colors["grid"], outline="")
                rounded_rect(canvas, bar_left, y1, max(bar_left + 4, bar_right), y2, radius=6, fill=color, outline="")
                canvas.create_text(right - 38, (y1 + y2) / 2, text=f"{value:,}", anchor="w", font=("Segoe UI", 8), fill=self.app.colors["text"])
            canvas.create_text(left, bottom + 14, text=row["month"], font=("Segoe UI", 8), fill=self.app.colors["muted"], anchor="w")
            return
        for key, label, color in zip(keys, platform_labels, colors):
            points = []
            for idx, row in enumerate(data):
                x = left + (right - left) * idx / (len(data) - 1)
                y = bottom - (bottom - top) * row[key] / max_value
                points.extend([x, y])
            canvas.create_line(*points, fill=color, width=2, smooth=True)
        canvas.create_line(left, bottom, right, bottom, fill=self.app.colors["border"])
        canvas.create_line(left, top, left, bottom, fill=self.app.colors["border"])
        for idx, row in enumerate(data):
            if idx in (0, len(data) - 1) or idx % 3 == 0:
                x = left + (right - left) * idx / (len(data) - 1)
                canvas.create_text(x, bottom + 14, text=row["month"], font=("Segoe UI", 8), fill=self.app.colors["muted"])
        legend_x = left
        for label, color in zip(platform_labels, colors):
            canvas.create_text(legend_x, height - 14, text=label, anchor="w", font=("Segoe UI", 8), fill=color)
            legend_x += max(70, len(label) * 7)


class YouTubeDownloaderPage(ttk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.cancel_download = False
        self._build()

    def _build(self):
        panel = ttk.Frame(self, style="Panel.TFrame", padding=18)
        panel.pack(fill="x", pady=(0, 14))
        ttk.Label(panel, text="Paste a YouTube link. Leave times empty for the full video, or enter a segment.", style="CardText.TLabel").pack(anchor="w")
        row = ttk.Frame(panel, style="Panel.TFrame")
        row.pack(fill="x", pady=(12, 0))
        self.url_entry = ttk.Entry(row)
        self.url_entry.pack(side="left", fill="x", expand=True)
        self.download_button = ttk.Button(row, text="Download", command=self.download_video)
        self.download_button.pack(side="left", padx=(8, 0))
        self.stop_button = ttk.Button(row, text="Stop", command=self.stop_download, state="disabled")
        self.stop_button.pack(side="left", padx=(8, 0))
        ttk.Button(row, text="Open Downloads", command=lambda: os.startfile(str(DOWNLOADS_PATH))).pack(side="left", padx=(8, 0))

        segment_row = ttk.Frame(panel, style="Panel.TFrame")
        segment_row.pack(fill="x", pady=(12, 0))
        ttk.Label(segment_row, text="Start", style="CardText.TLabel").pack(side="left")
        self.start_entry = ttk.Entry(segment_row, width=12)
        self.start_entry.pack(side="left", padx=(6, 14))
        ttk.Label(segment_row, text="End", style="CardText.TLabel").pack(side="left")
        self.end_entry = ttk.Entry(segment_row, width=12)
        self.end_entry.pack(side="left", padx=(6, 14))
        ttk.Label(segment_row, text="Use seconds, MM:SS or HH:MM:SS.", style="CardText.TLabel").pack(side="left")

        self.status_text = tk.Text(self, height=24, wrap="word", font=("Consolas", 9), padx=10, pady=10)
        self.app.style_text_widget(self.status_text)
        self.status_text.pack(fill="both", expand=True)

    def write_status(self, text):
        self.after(0, lambda: self.status_text.insert(tk.END, text))
        self.after(0, lambda: self.status_text.see(tk.END))

    def stop_download(self):
        self.cancel_download = True
        self.write_status("\nStopping after the current download step...\n")

    def download_video(self):
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showwarning("Missing link", "Paste a YouTube link first.")
            return
        try:
            from yt_dlp import YoutubeDL
        except Exception as exc:
            messagebox.showerror("yt-dlp missing", f"This function needs yt-dlp:\n\n{exc}")
            return

        try:
            start_seconds = parse_time_to_seconds(self.start_entry.get())
            end_seconds = parse_time_to_seconds(self.end_entry.get())
        except ValueError as exc:
            messagebox.showerror("Invalid time", str(exc))
            return

        if start_seconds is not None and end_seconds is not None and end_seconds <= start_seconds:
            messagebox.showerror("Invalid time", "End must be later than Start.")
            return

        if start_seconds is None and end_seconds is not None:
            start_seconds = 0

        downloads_folder = str(DOWNLOADS_PATH)
        is_segment = start_seconds is not None or end_seconds is not None
        high_quality_segment = is_segment
        if is_segment:
            segment_label = f"segment {format_seconds_for_filename(start_seconds or 0)} to {format_seconds_for_filename(end_seconds)}"
            if high_quality_segment:
                temp_dir = DOWNLOADS_PATH / "_creative_toolbox_temp"
                temp_dir.mkdir(parents=True, exist_ok=True)
                output_template = str(temp_dir / "%(title).160B - full temp.%(ext)s")
            else:
                output_template = os.path.join(downloads_folder, f"%(title).160B - {segment_label} - %(resolution)s.%(ext)s")
        else:
            output_template = os.path.join(downloads_folder, "%(title).200B - %(resolution)s.%(ext)s")

        class GuiLogger:
            def debug(_, msg):
                if msg:
                    self.write_status(str(msg) + "\n")
            def warning(_, msg):
                self.write_status("WARNING: " + str(msg) + "\n")
            def error(_, msg):
                self.write_status("ERROR: " + str(msg) + "\n")

        def progress_hook(data):
            if self.cancel_download:
                raise Exception("Download stopped by user.")
            if data.get("status") == "downloading":
                percent = data.get("_percent_str", "").strip()
                speed = data.get("_speed_str", "").strip()
                eta = data.get("_eta_str", "").strip()
                self.write_status(f"Downloading: {percent} | Speed: {speed} | ETA: {eta}\n")
            elif data.get("status") == "finished":
                self.write_status("Download finished. Processing MP4...\n")

        if high_quality_segment:
            format_selector = "bv*+ba/bestvideo+bestaudio/best"
        else:
            format_selector = "bv*[ext=mp4][vcodec^=avc1]+ba[ext=m4a]/bv*[ext=mp4]+ba[ext=m4a]/b[ext=mp4]/best"

        ydl_opts = {
            "format": format_selector,
            "outtmpl": output_template,
            "noplaylist": True,
            "windowsfilenames": True,
            "socket_timeout": 30,
            "retries": 3,
            "fragment_retries": 3,
            "remuxvideo": "mp4",
            "merge_output_format": "mp4",
            "ffmpeg_location": str(Path(find_ffmpeg_exe()).parent) if find_ffmpeg_exe() else r"C:\ffmpeg\bin",
            "logger": GuiLogger(),
            "progress_hooks": [progress_hook],
        }

        self.download_button.configure(state="disabled")
        self.status_text.delete("1.0", tk.END)
        if not is_segment:
            self.write_status("Starting full download...\n\n")
        else:
            end_text = "end" if end_seconds is None else str(end_seconds)
            self.write_status(f"Starting segment download: {start_seconds} to {end_text} seconds...\n\n")
            self.write_status("Note: the final file may only appear after YouTube download and MP4 processing are finished.\n\n")
            self.write_status("Attempting the highest available video and audio, then cutting locally with ffmpeg.\n\n")
        self.write_status(f"Save folder: {downloads_folder}\n\n")
        started_at = datetime.now().timestamp()
        self.cancel_download = False
        self.stop_button.configure(state="normal")

        def run():
            try:
                if high_quality_segment:
                    self.run_high_quality_segment_download(
                        url,
                        ydl_opts,
                        start_seconds or 0,
                        end_seconds,
                        segment_label,
                    )
                else:
                    with YoutubeDL(ydl_opts) as ydl:
                        ydl.download([url])
                created_files = recent_downloads(started_at)
                self.write_status(f"\nDone.\nSaved in: {downloads_folder}\n")
                if created_files:
                    self.write_status("\nRecent created/changed files:\n")
                    for path in created_files[:6]:
                        self.write_status(f"- {path.name}\n")
                else:
                    self.write_status("\nNo new media file was detected in Downloads. Check the log above for an error or a still-running .part file.\n")
            except Exception as exc:
                self.write_status(f"\nError: {exc}\n")
            finally:
                self.after(0, lambda: self.download_button.configure(state="normal"))
                self.after(0, lambda: self.stop_button.configure(state="disabled"))

        threading.Thread(target=run, daemon=True).start()

    def run_high_quality_segment_download(self, url, ydl_opts, start_seconds, end_seconds, segment_label):
        from yt_dlp import YoutubeDL

        temp_dir = DOWNLOADS_PATH / "_creative_toolbox_temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        started_at = datetime.now().timestamp()

        self.write_status("Step 1/2: downloading temporary high-quality source...\n")
        with YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=True)

        if self.cancel_download:
            raise Exception("Download stopped by user.")

        downloaded_path = None
        requested = info.get("requested_downloads") or []
        for item in requested:
            filepath = item.get("filepath")
            if filepath and Path(filepath).exists():
                downloaded_path = Path(filepath)
                break

        if downloaded_path is None:
            candidates = [
                path for path in temp_dir.iterdir()
                if path.is_file() and path.stat().st_mtime >= started_at and path.suffix.lower() in {".mp4", ".mkv", ".webm"}
            ]
            if candidates:
                downloaded_path = sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]

        if downloaded_path is None:
            raise Exception("The temporary high-quality download finished, but the file could not be found.")

        ffmpeg = find_ffmpeg_exe()
        if not ffmpeg:
            raise Exception("FFmpeg was not found. Expected C:\\ffmpeg\\bin\\ffmpeg.exe or ffmpeg on PATH.")

        title = safe_media_name(info.get("title") or downloaded_path.stem)
        resolution = resolution_from_info(info)
        final_path = DOWNLOADS_PATH / f"{title} - {segment_label} - {resolution}.mp4"

        self.write_status(f"\nTemporary source:\n{downloaded_path}\n")
        self.write_status("Step 2/2: cutting segment locally with ffmpeg...\n")

        cmd = [
            ffmpeg,
            "-hide_banner",
            "-y",
            "-ss",
            str(float(start_seconds)),
        ]
        if end_seconds is not None:
            cmd.extend(["-t", str(float(end_seconds - start_seconds))])
        cmd.extend([
            "-i",
            str(downloaded_path),
            "-map",
            "0:v:0?",
            "-map",
            "0:a:0?",
            "-c",
            "copy",
            "-avoid_negative_ts",
            "make_zero",
            str(final_path),
        ])

        completed = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )

        if completed.returncode != 0:
            self.write_status("\nFast local copy cut failed; retrying with re-encode...\n")
            cmd = [
                ffmpeg,
                "-hide_banner",
                "-y",
                "-ss",
                str(float(start_seconds)),
            ]
            if end_seconds is not None:
                cmd.extend(["-t", str(float(end_seconds - start_seconds))])
            cmd.extend([
                "-i",
                str(downloaded_path),
                "-map",
                "0:v:0?",
                "-map",
                "0:a:0?",
                "-c:v",
                "libx264",
                "-preset",
                "veryfast",
                "-crf",
                "18",
                "-c:a",
                "aac",
                "-b:a",
                "192k",
                str(final_path),
            ])
            completed = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
            )

        if completed.returncode != 0:
            raise Exception((completed.stderr or completed.stdout or "FFmpeg failed.").strip())

        self.write_status(f"\nCreated segment:\n{final_path}\n")


if __name__ == "__main__":
    CreativeToolbox().mainloop()


